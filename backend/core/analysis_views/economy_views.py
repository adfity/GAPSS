from rest_framework.decorators import api_view
from rest_framework.response import Response
from pymongo import MongoClient
import psycopg2
import psycopg2.extras
import requests
import uuid
from datetime import datetime
import os
from dotenv import load_dotenv

# download xlsx
import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Load .env ─────────────────────────────────────────────────────────────────
base_dir    = os.path.dirname(__file__)
dotenv_path = os.path.abspath(os.path.join(base_dir, '..', '..', '.env'))
load_dotenv(dotenv_path)

MONGO_URI     = os.getenv("MONGO_URI")
DB_MONGO_NAME = os.getenv("DB_MONGO_NAME")
BPS_API_KEY   = os.getenv("BPS_WEB_API_KEY")

client   = MongoClient(MONGO_URI)
mongo_db = client[DB_MONGO_NAME]

# ── Koneksi PostgreSQL ────────────────────────────────────────────────────────
def get_pg_connection():
    return psycopg2.connect(
        host     = os.getenv("DB_HOST"),
        port     = int(os.getenv("DB_PORT", "5432")),
        dbname   = os.getenv("DB_NAME"),
        user     = os.getenv("DB_USER"),
        password = os.getenv("DB_PASSWORD"),
    )


# ─── MAPPING TAHUN → KODE BPS ────────────────────────────────────────────────
TAHUN_BPS_MAP = {
    2020: "120",
    2021: "121",
    2022: "122",
    2023: "123",
    2024: "124",
    2025: "125",
    2026: "126",
}

# Untuk auto-detect (urutan terbaru ke terlama)
BPS_TAHUN_URUT = ["126", "125", "124", "123", "122", "121", "120",
                  "119", "118", "117", "116", "115", "114", "113",
                  "112", "111", "110"]

BPS_TAHUN_KAL_MAP = {v: k for k, v in {
    2030: "130", 2029: "129", 2028: "128", 2027: "127",
    2026: "126", 2025: "125", 2024: "124", 2023: "123",
    2022: "122", 2021: "121", 2020: "120", 2019: "119",
    2018: "118", 2017: "117", 2016: "116", 2015: "115",
    2014: "114", 2013: "113", 2012: "112", 2011: "111", 2010: "110",
}.items()}  # th_code → tahun_kalender

# Inverse: th_code → tahun kalender
TH_CODE_TO_TAHUN = {
    "130": 2030, "129": 2029, "128": 2028, "127": 2027,
    "126": 2026, "125": 2025, "124": 2024, "123": 2023,
    "122": 2022, "121": 2021, "120": 2020, "119": 2019,
    "118": 2018, "117": 2017, "116": 2016, "115": 2015,
    "114": 2014, "113": 2013, "112": 2012, "111": 2011, "110": 2010,
}

# ─── MAPPING INDIKATOR → DATASET YANG PERLU DIFETCH ──────────────────────────
INDIKATOR_DATASET_MAP = {
    'ALL':        ['PDRB', 'KEMISKINAN', 'INVESTASI'],
    'PDRB':       ['PDRB'],
    'KEMISKINAN': ['KEMISKINAN'],
    'INVESTASI':  ['INVESTASI'],
}

# ─── LABEL INDIKATOR ─────────────────────────────────────────────────────────
INDIKATOR_LABELS = {
    'ALL':        'Semua Indikator',
    'PDRB':       'PDRB Atas Dasar Harga Berlaku',
    'KEMISKINAN': 'Persentase Penduduk Miskin',
    'INVESTASI':  'Realisasi Investasi PMDN',
}


def get_indikator_config(tahun: int) -> dict:
    """Return konfigurasi indikator ekonomi dengan kode th BPS sesuai tahun."""
    th = TAHUN_BPS_MAP.get(tahun, "124")

    return {
        "PDRB": {
            "url_template": f"https://webapi.bps.go.id/v1/api/list/model/data/lang/ind/domain/0000/var/534/th/{th}/key/{{key}}/",
            "nama":     "PDRB Atas Dasar Harga Berlaku Menurut Pengeluaran",
            "satuan":   "Milyar Rupiah",
            "threshold_tinggi": 75000,
            "threshold_sedang": 50000,
            "bobot":    0.40,
            "reverse":  False,
            "penjelasan": "Produk Domestik Regional Bruto yang mencerminkan kapasitas ekonomi daerah dan output ekonomi total",
        },
        "KEMISKINAN": {
            "url_template": f"https://webapi.bps.go.id/v1/api/list/model/data/lang/ind/domain/0000/var/192/th/{th}/key/{{key}}/",
            "nama":     "Persentase Penduduk Miskin",
            "satuan":   "%",
            "threshold_rendah": 7,
            "threshold_sedang": 12,
            "bobot":    0.40,
            "reverse":  True,
            "penjelasan": "Persentase penduduk yang hidup di bawah garis kemiskinan",
        },
        "INVESTASI": {
            "url_template": f"https://webapi.bps.go.id/v1/api/list/model/data/lang/ind/domain/0000/var/793/th/{th}/key/{{key}}/",
            "nama":     "Realisasi Investasi PMDN",
            "satuan":   "Milyar Rupiah",
            "threshold_tinggi": 10000,
            "threshold_sedang": 5000,
            "bobot":    0.20,
            "reverse":  False,
            "penjelasan": "Investasi Penanaman Modal Dalam Negeri",
        },
    }


# ─── HELPER: CEK DATA KOSONG ─────────────────────────────────────────────────
def _is_data_empty(data):
    if data is None:
        return True
    datacontent = data.get("datacontent", {})
    if not datacontent:
        return True
    valid = [v for v in datacontent.values() if v is not None and v != 0]
    return len(valid) == 0


# ─── ENDPOINT: CEK DATA ──────────────────────────────────────────────────────
@api_view(['POST'])
def check_ekonomi_year_data(request):
    """
    Cek ketersediaan data BPS ekonomi untuk tahun dan indikator yang dipilih.
    Mirip check_health_year_data di health_views.
    """
    if not BPS_API_KEY:
        return Response({"error": "BPS Web API Key belum dikonfigurasi"}, status=500)

    tahun     = request.data.get('tahun', 2024)
    indikator = request.data.get('indikator', 'ALL')

    try:
        tahun = int(tahun)
    except (ValueError, TypeError):
        tahun = 2024

    if tahun not in TAHUN_BPS_MAP:
        return Response({"error": f"Tahun {tahun} tidak didukung. Pilih antara 2020–2026."}, status=400)

    keys_to_check = INDIKATOR_DATASET_MAP.get(indikator, INDIKATOR_DATASET_MAP['ALL'])
    all_config    = get_indikator_config(tahun)
    dataset_status = {}

    for key in keys_to_check:
        config = all_config.get(key)
        if not config:
            continue
        url = config["url_template"].format(key=BPS_API_KEY)
        try:
            resp = requests.get(url, timeout=20)
            if resp.status_code == 200:
                data   = resp.json()
                kosong = _is_data_empty(data)
                dataset_status[key] = {
                    "nama":     config["nama"],
                    "tersedia": not kosong,
                    "status":   "Tersedia" if not kosong else "Kosong / Tidak Tersedia",
                }
            else:
                dataset_status[key] = {
                    "nama":     config["nama"],
                    "tersedia": False,
                    "status":   f"HTTP Error {resp.status_code}",
                }
        except Exception as e:
            dataset_status[key] = {
                "nama":     config["nama"],
                "tersedia": False,
                "status":   f"Gagal ({str(e)[:50]})",
            }

    tersedia_list   = [k for k, v in dataset_status.items() if v["tersedia"]]
    kosong_list     = [k for k, v in dataset_status.items() if not v["tersedia"]]
    semua_kosong    = len(tersedia_list) == 0
    ada_yang_kosong = len(kosong_list) > 0 and not semua_kosong

    return Response({
        "tahun":           tahun,
        "indikator":       indikator,
        "dataset_status":  dataset_status,
        "tersedia":        tersedia_list,
        "kosong":          kosong_list,
        "semua_kosong":    semua_kosong,
        "ada_yang_kosong": ada_yang_kosong,
        "bisa_dieksekusi": not semua_kosong and not ada_yang_kosong,
    })


# ─── HELPER STYLE EXCEL ──────────────────────────────────────────────────────
def _style_header(ws, row_num, col_count, title, subtitle=None):
    COLOR_HEADER    = "0C4A6E"   # biru tua ekonomi
    COLOR_SUBHEADER = "0369A1"
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=col_count)
    cell           = ws.cell(row=row_num, column=1, value=title)
    cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=14)
    cell.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row_num].height = 30
    if subtitle:
        row_num += 1
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=col_count)
        cell           = ws.cell(row=row_num, column=1, value=subtitle)
        cell.font      = Font(name="Arial", italic=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", fgColor=COLOR_SUBHEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_num].height = 20
        row_num += 1
    else:
        row_num += 1
    return row_num


def _style_col_headers(ws, row_num, headers, col_widths=None):
    COLOR_COL = "0369A1"
    thin   = Side(style="thin", color="FFFFFF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_idx, header in enumerate(headers, start=1):
        cell           = ws.cell(row=row_num, column=col_idx, value=header)
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", fgColor=COLOR_COL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
    ws.row_dimensions[row_num].height = 35
    if col_widths:
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    return row_num + 1


def _write_data_rows(ws, start_row, data_rows, number_cols=None):
    COLOR_EVEN = "DBEAFE"   # biru muda
    COLOR_ODD  = "FFFFFF"
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row_offset, row_data in enumerate(data_rows):
        row_num    = start_row + row_offset
        fill_color = COLOR_EVEN if row_offset % 2 == 0 else COLOR_ODD
        fill       = PatternFill("solid", fgColor=fill_color)
        for col_idx, value in enumerate(row_data, start=1):
            cell           = ws.cell(row=row_num, column=col_idx, value=value)
            cell.fill      = fill
            cell.border    = border
            cell.font      = Font(name="Arial", size=10)
            if number_cols and col_idx in number_cols:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if isinstance(value, float):
                    cell.number_format = '#,##0.00'
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row_num].height = 18
    return start_row + len(data_rows)


def _add_source_footer(ws, row_num, col_count, source_text, timestamp=None):
    row_num += 1
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=col_count)
    text = f"Sumber: {source_text}"
    if timestamp:
        text += f"  |  Waktu Pengambilan Data: {timestamp}"
    cell           = ws.cell(row=row_num, column=1, value=text)
    cell.font      = Font(name="Arial", italic=True, color="595959", size=9)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row_num].height = 16


# ─── DOWNLOAD XLSX: PDRB ─────────────────────────────────────────────────────
@api_view(['POST'])
def download_pdrb_xlsx(request):
    try:
        pdrb_data = request.data.get('pdrb_data')
        timestamp = request.data.get('timestamp', datetime.now().isoformat())
        tahun     = request.data.get('tahun', 2024)

        if not pdrb_data:
            return Response({"error": "Data PDRB tidak ditemukan"}, status=400)

        wb = Workbook()
        ws = wb.active
        ws.title = "PDRB"
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A4"

        next_row = _style_header(ws, 1, 3,
            "PDRB ATAS DASAR HARGA BERLAKU MENURUT PENGELUARAN",
            f"Sumber: BPS | Tahun {tahun} | Seluruh Provinsi Indonesia")

        headers    = ["No.", "Provinsi", "PDRB (Milyar Rupiah)"]
        col_widths = [6, 35, 28]
        next_row   = _style_col_headers(ws, next_row, headers, col_widths)

        data_rows = []
        for idx, (prov, data) in enumerate(sorted(pdrb_data.items()), start=1):
            val = data.get('nilai', data) if isinstance(data, dict) else data
            data_rows.append([idx, prov, val])

        _write_data_rows(ws, next_row, data_rows, number_cols={3})
        _add_source_footer(ws, next_row + len(data_rows), 3,
            f"BPS Web API - Variabel 534, Tahun {tahun}",
            timestamp[:19].replace('T', ' ') if timestamp else None)

        output = io.BytesIO()
        wb.save(output); output.seek(0)
        tanggal  = datetime.now().strftime("%Y-%m-%d")
        response = HttpResponse(output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Dataset_PDRB_BPS_{tahun}_{tanggal}.xlsx"'
        return response

    except Exception as e:
        import traceback; traceback.print_exc()
        return Response({"error": str(e)}, status=500)


# ─── DOWNLOAD XLSX: KEMISKINAN ───────────────────────────────────────────────
@api_view(['POST'])
def download_kemiskinan_xlsx(request):
    try:
        kemiskinan_data = request.data.get('kemiskinan_data')
        timestamp       = request.data.get('timestamp', datetime.now().isoformat())
        tahun           = request.data.get('tahun', 2024)

        if not kemiskinan_data:
            return Response({"error": "Data Kemiskinan tidak ditemukan"}, status=400)

        wb = Workbook()
        ws = wb.active
        ws.title = "Kemiskinan"
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A4"

        next_row = _style_header(ws, 1, 3,
            "PERSENTASE PENDUDUK MISKIN",
            f"Sumber: BPS Susenas | Tahun {tahun} | Seluruh Provinsi Indonesia")

        headers    = ["No.", "Provinsi", "Persentase Penduduk Miskin (%)"]
        col_widths = [6, 35, 30]
        next_row   = _style_col_headers(ws, next_row, headers, col_widths)

        data_rows = []
        for idx, (prov, data) in enumerate(sorted(kemiskinan_data.items()), start=1):
            val = data.get('nilai', data) if isinstance(data, dict) else data
            data_rows.append([idx, prov, val])

        _write_data_rows(ws, next_row, data_rows, number_cols={3})
        _add_source_footer(ws, next_row + len(data_rows), 3,
            f"BPS Web API - Susenas, Variabel 192, Tahun {tahun}",
            timestamp[:19].replace('T', ' ') if timestamp else None)

        output = io.BytesIO()
        wb.save(output); output.seek(0)
        tanggal  = datetime.now().strftime("%Y-%m-%d")
        response = HttpResponse(output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Dataset_Kemiskinan_BPS_{tahun}_{tanggal}.xlsx"'
        return response

    except Exception as e:
        import traceback; traceback.print_exc()
        return Response({"error": str(e)}, status=500)


# ─── DOWNLOAD XLSX: INVESTASI ────────────────────────────────────────────────
@api_view(['POST'])
def download_investasi_xlsx(request):
    try:
        investasi_data = request.data.get('investasi_data')
        timestamp      = request.data.get('timestamp', datetime.now().isoformat())
        tahun          = request.data.get('tahun', 2024)

        if not investasi_data:
            return Response({"error": "Data Investasi tidak ditemukan"}, status=400)

        wb = Workbook()
        ws = wb.active
        ws.title = "Investasi"
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A4"

        next_row = _style_header(ws, 1, 3,
            "REALISASI INVESTASI PMDN",
            f"Sumber: BPS | Tahun {tahun} | Seluruh Provinsi Indonesia")

        headers    = ["No.", "Provinsi", "Realisasi PMDN (Milyar Rupiah)"]
        col_widths = [6, 35, 30]
        next_row   = _style_col_headers(ws, next_row, headers, col_widths)

        data_rows = []
        for idx, (prov, data) in enumerate(sorted(investasi_data.items()), start=1):
            val = data.get('nilai', data) if isinstance(data, dict) else data
            data_rows.append([idx, prov, val])

        _write_data_rows(ws, next_row, data_rows, number_cols={3})
        _add_source_footer(ws, next_row + len(data_rows), 3,
            f"BPS Web API - Variabel 793, Tahun {tahun}",
            timestamp[:19].replace('T', ' ') if timestamp else None)

        output = io.BytesIO()
        wb.save(output); output.seek(0)
        tanggal  = datetime.now().strftime("%Y-%m-%d")
        response = HttpResponse(output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Dataset_Investasi_BPS_{tahun}_{tanggal}.xlsx"'
        return response

    except Exception as e:
        import traceback; traceback.print_exc()
        return Response({"error": str(e)}, status=500)


# ═══════════════════════════════════════════════════════════════════════════════
# HELPER: BANK KEBIJAKAN dari PostgreSQL
# ═══════════════════════════════════════════════════════════════════════════════

def get_bank_kebijakan2_by_kategori(kategori_list: list, limit_per_kategori: int = 8) -> list:
    results = []
    conn    = None
    try:
        conn = get_pg_connection()
        cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        for kategori in kategori_list:
            cur.execute("""
                SELECT id, kategori_utama, sub_sektor, prioritas, no_aksi,
                       nama_aksi, detail_aksi, timeline, budget_est,
                       sektor_terkait, indikator_dampak
                FROM bank_kebijakan2
                WHERE domain = 'ekonomi' AND kategori_utama = %s
                ORDER BY no_aksi ASC
                LIMIT %s
            """, (kategori, limit_per_kategori))

            docs = [dict(row) for row in cur.fetchall()]
            if docs:
                results.append({
                    "kategori":    kategori,
                    "prioritas":   docs[0].get("prioritas", "-"),
                    "jumlah_aksi": len(docs),
                    "aksi":        docs,
                })
        cur.close()
    except Exception as e:
        print(f"  ✗ Error get_bank_kebijakan2_by_kategori: {e}")
    finally:
        if conn:
            conn.close()

    return results


def build_kategori_list(kategori_iek: str, data_ekonomi: dict, indikator_config: dict) -> list:
    kategori_list = [kategori_iek]

    pdrb       = data_ekonomi.get("PDRB")
    kemiskinan = data_ekonomi.get("KEMISKINAN")
    investasi  = data_ekonomi.get("INVESTASI")

    cfg_pdrb = indikator_config.get("PDRB", {})
    cfg_kem  = indikator_config.get("KEMISKINAN", {})
    cfg_inv  = indikator_config.get("INVESTASI", {})

    if pdrb is not None and pdrb < cfg_pdrb.get("threshold_sedang", 50000):
        kategori_list.append("PDRB_RENDAH")
    if kemiskinan is not None and kemiskinan > cfg_kem.get("threshold_sedang", 12):
        kategori_list.append("KEMISKINAN_TINGGI")
    if investasi is not None and investasi < cfg_inv.get("threshold_sedang", 5000):
        kategori_list.append("INVESTASI_RENDAH")

    return list(dict.fromkeys(kategori_list))


# ═══════════════════════════════════════════════════════════════════════════════
# KELAS ANALITIK EKONOMI
# ═══════════════════════════════════════════════════════════════════════════════

class EkonomiAnalytics:

    COLORS = {
        "MAJU":       "#10b981",
        "BERKEMBANG": "#f59e0b",
        "TERTINGGAL": "#ef4444",
    }

    def __init__(self, tahun: int = 2024):
        self.tahun            = tahun
        self.th_code          = TAHUN_BPS_MAP.get(tahun, "124")
        self.indikator_config = get_indikator_config(tahun)
        self.timestamp_fetch  = None

    def fetch_all_data(self):
        return self.fetch_selected_data(list(self.indikator_config.keys()))

    def fetch_selected_data(self, keys: list) -> dict:
        all_data = {}
        self.timestamp_fetch = datetime.now().isoformat()
        for key in keys:
            config = self.indikator_config.get(key)
            if not config:
                continue
            try:
                url  = config["url_template"].format(key=BPS_API_KEY)
                print(f"Fetching {key} (th={self.th_code}, tahun={self.tahun}): {url}")
                resp = requests.get(url, timeout=30)
                if resp.status_code == 200:
                    raw = resp.json()
                    if raw and raw.get("datacontent"):
                        all_data[key] = raw
                        print(f"  ✓ {key}: Success")
                    else:
                        all_data[key] = None
                        print(f"  ✗ {key}: Kosong")
                else:
                    print(f"  ✗ {key}: HTTP {resp.status_code}")
                    all_data[key] = None
            except Exception as e:
                print(f"  ✗ {key}: Error - {e}")
                all_data[key] = None
        return all_data

    def parse_province_data(self, raw_data, indikator_key: str) -> dict:
        """
        Parse data BPS → { nama_provinsi: nilai_float }.
        Kembalikan juga raw_breakdown untuk download xlsx.
        """
        province_values = {}
        raw_breakdown   = {}

        if not raw_data:
            return province_values, raw_breakdown

        try:
            datacontent   = raw_data.get("datacontent", {})
            vervar_list   = raw_data.get("vervar", [])
            prov_code_map = {}
            for item in vervar_list:
                code  = str(item.get("val", ""))
                label = item.get("label", "")
                if code and label and code != "9999":
                    prov_code_map[code] = label

            for dc_key, value in datacontent.items():
                try:
                    prov_code = dc_key[:4]
                    if prov_code == "9999":
                        continue
                    prov_name = prov_code_map.get(prov_code)
                    if prov_name and value is not None:
                        prov_clean  = normalize_province_name(str(prov_name))
                        value_float = float(value)
                        # Ambil nilai terbaru jika ada duplikat (last-write)
                        province_values[prov_clean] = value_float
                        raw_breakdown[prov_clean]   = {"provinsi": prov_clean, "nilai": value_float}
                except (ValueError, TypeError, IndexError):
                    continue

            print(f"  ✅ {indikator_key}: Parsed {len(province_values)} provinces")

        except Exception as e:
            print(f"  ❌ Parse error {indikator_key}: {e}")
            import traceback; traceback.print_exc()

        return province_values, raw_breakdown

    def calculate_scores(self, data_ekonomi: dict, indikator: str = 'ALL') -> dict:
        """
        Hitung skor berdasarkan indikator yang aktif.
        ALL       : IEK = PDRB×0.40 + KEMISKINAN×0.40 + INVESTASI×0.20
        PDRB      : skor murni PDRB
        KEMISKINAN: skor murni Kemiskinan
        INVESTASI : skor murni Investasi
        """
        cfg = self.indikator_config

        def _skor_pdrb(val):
            if val is None: return 0
            c = cfg["PDRB"]
            return 3 if val > c["threshold_tinggi"] else 2 if val > c["threshold_sedang"] else 1

        def _skor_kemiskinan(val):
            if val is None: return 0
            c = cfg["KEMISKINAN"]
            return 3 if val < c["threshold_rendah"] else 2 if val < c["threshold_sedang"] else 1

        def _skor_investasi(val):
            if val is None: return 0
            c = cfg["INVESTASI"]
            return 3 if val > c["threshold_tinggi"] else 2 if val > c["threshold_sedang"] else 1

        s_pdrb  = _skor_pdrb(data_ekonomi.get("PDRB"))
        s_kem   = _skor_kemiskinan(data_ekonomi.get("KEMISKINAN"))
        s_inv   = _skor_investasi(data_ekonomi.get("INVESTASI"))

        if indikator == 'PDRB':
            skor_total = float(s_pdrb)
        elif indikator == 'KEMISKINAN':
            skor_total = float(s_kem)
        elif indikator == 'INVESTASI':
            skor_total = float(s_inv)
        else:
            # ALL dengan bobot
            total_score  = 0.0
            total_weight = 0.0
            pairs = [
                ("PDRB",       s_pdrb, 0.40),
                ("KEMISKINAN", s_kem,  0.40),
                ("INVESTASI",  s_inv,  0.20),
            ]
            for key, skor, bobot in pairs:
                if data_ekonomi.get(key) is not None:
                    total_score  += skor * bobot
                    total_weight += bobot
            skor_total = round(total_score / total_weight, 2) if total_weight > 0 else 0.0

        return {
            'skor_pdrb':       s_pdrb,
            'skor_kemiskinan': s_kem,
            'skor_investasi':  s_inv,
            'skor_total':      skor_total,
        }

    def categorize_province(self, skor_total: float):
        if skor_total >= 2.4:
            return "MAJU",       self.COLORS["MAJU"]
        elif skor_total >= 1.8:
            return "BERKEMBANG", self.COLORS["BERKEMBANG"]
        else:
            return "TERTINGGAL", self.COLORS["TERTINGGAL"]

    def generate_insights(self, provinsi, data_ekonomi, kategori, skor_total, indikator='ALL'):
        insights = [f"Provinsi {provinsi} berada pada kategori {kategori} dengan skor {skor_total}."]
        cfg      = self.indikator_config

        pdrb       = data_ekonomi.get("PDRB")
        kemiskinan = data_ekonomi.get("KEMISKINAN")
        investasi  = data_ekonomi.get("INVESTASI")

        if indikator in ('ALL', 'PDRB') and pdrb is not None:
            c = cfg["PDRB"]
            if pdrb > c["threshold_tinggi"]:
                insights.append(f"📈 PDRB: Rp{pdrb:.0f} milyar — TINGGI, kapasitas ekonomi kuat.")
            elif pdrb > c["threshold_sedang"]:
                insights.append(f"📊 PDRB: Rp{pdrb:.0f} milyar — SEDANG, perlu penguatan.")
            else:
                insights.append(f"📉 PDRB: Rp{pdrb:.0f} milyar — RENDAH, perhatian khusus diperlukan.")

        if indikator in ('ALL', 'KEMISKINAN') and kemiskinan is not None:
            c = cfg["KEMISKINAN"]
            if kemiskinan < c["threshold_rendah"]:
                insights.append(f"✅ Kemiskinan: {kemiskinan}% — RENDAH, kondisi baik.")
            elif kemiskinan < c["threshold_sedang"]:
                insights.append(f"⚠️ Kemiskinan: {kemiskinan}% — SEDANG, perlu intervensi.")
            else:
                insights.append(f"🚨 Kemiskinan: {kemiskinan}% — TINGGI, risiko sosial meningkat.")

        if indikator in ('ALL', 'INVESTASI') and investasi is not None:
            c = cfg["INVESTASI"]
            if investasi > c["threshold_tinggi"]:
                insights.append(f"💰 Investasi PMDN: Rp{investasi:.0f} milyar — TINGGI, iklim investasi kondusif.")
            elif investasi > c["threshold_sedang"]:
                insights.append(f"💵 Investasi PMDN: Rp{investasi:.0f} milyar — SEDANG.")
            else:
                insights.append(f"💸 Investasi PMDN: Rp{investasi:.0f} milyar — RENDAH, perlu stimulus.")

        return insights


# ─── NORMALIZE ───────────────────────────────────────────────────────────────
def normalize_province_name(name: str) -> str:
    if not isinstance(name, str):
        name = str(name)
    name = name.upper().strip()

    special = {
        "DKI JAKARTA":                   "JAKARTA",
        "DAERAH KHUSUS IBUKOTA JAKARTA": "JAKARTA",
        "DKI":                           "JAKARTA",
        "YOGYAKARTA":                    "DAERAH ISTIMEWA YOGYAKARTA",
        "DIY":                           "DAERAH ISTIMEWA YOGYAKARTA",
        "D.I. YOGYAKARTA":               "DAERAH ISTIMEWA YOGYAKARTA",
        "BANGKA BELITUNG":               "KEPULAUAN BANGKA BELITUNG",
        "KEP. BANGKA BELITUNG":          "KEPULAUAN BANGKA BELITUNG",
        "KEP. RIAU":                     "KEPULAUAN RIAU",
    }
    for k, v in special.items():
        if k in name:
            return v

    abbr = {
        "KEP.": "KEPULAUAN",
        "KEP ": "KEPULAUAN ",
        "NTB":  "NUSA TENGGARA BARAT",
        "NTT":  "NUSA TENGGARA TIMUR",
    }
    for a, f in abbr.items():
        name = name.replace(a, f)

    for prefix in ["PROVINSI ", "PROV. ", "PROV ", "DAERAH KHUSUS IBUKOTA "]:
        if name.startswith(prefix):
            name = name[len(prefix):]

    return name.strip()


# ═══════════════════════════════════════════════════════════════════════════════
# ENDPOINT: GET BANK KEBIJAKAN dari PostgreSQL
# ═══════════════════════════════════════════════════════════════════════════════

@api_view(["GET"])
def get_bank_kebijakan2(request):
    """
    GET /api/bank-kebijakan/
    Query params:
      - domain     : ekonomi | kesehatan | pendidikan (default: ekonomi)
      - kategori   : TERTINGGAL | BERKEMBANG | MAJU | PDRB_RENDAH | dst
      - limit      : jumlah aksi (default 50, max 300)
      - sub_sektor : filter opsional
    """
    domain     = request.GET.get("domain", "ekonomi").lower().strip()
    kategori   = request.GET.get("kategori", "").upper().strip()
    limit      = min(int(request.GET.get("limit", 50)), 300)
    sub_sektor = request.GET.get("sub_sektor", "").strip()

    conn = None
    try:
        conn = get_pg_connection()
        cur  = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        where_clauses = ["domain = %s"]
        params        = [domain]

        if kategori:
            where_clauses.append("kategori_utama = %s")
            params.append(kategori)
        if sub_sektor:
            where_clauses.append("sub_sektor ILIKE %s")
            params.append(f"%{sub_sektor}%")

        where_sql = "WHERE " + " AND ".join(where_clauses)
        params.append(limit)

        cur.execute(f"""
            SELECT id, domain, kategori_utama, sub_sektor, prioritas, no_aksi,
                   nama_aksi, detail_aksi, timeline, budget_est,
                   sektor_terkait, indikator_dampak
            FROM bank_kebijakan2
            {where_sql}
            ORDER BY no_aksi ASC
            LIMIT %s
        """, params)

        docs = [dict(row) for row in cur.fetchall()]

        cur.execute("""
            SELECT kategori_utama, COUNT(*) as jumlah
            FROM bank_kebijakan2
            WHERE domain = %s
            GROUP BY kategori_utama
            ORDER BY kategori_utama
        """, (domain,))
        distribusi = {row["kategori_utama"]: row["jumlah"] for row in cur.fetchall()}
        cur.close()

        return Response({
            "status":     "success",
            "domain":     domain,
            "total":      len(docs),
            "filter":     {"domain": domain, "kategori": kategori, "sub_sektor": sub_sektor},
            "distribusi": distribusi,
            "data":       docs,
        })

    except Exception as e:
        return Response({"error": str(e)}, status=500)
    finally:
        if conn:
            conn.close()


# ═══════════════════════════════════════════════════════════════════════════════
# ENDPOINT: ANALISIS EKONOMI BPS — pola seragam dengan health_views
# ═══════════════════════════════════════════════════════════════════════════════

@api_view(["POST"])
def analyze_ekonomi_bps(request):
    """
    Analisis data ekonomi menggunakan BPS Web API.
    Body JSON:
      - tahun     : 2020–2026 (default 2024)
      - indikator : ALL | PDRB | KEMISKINAN | INVESTASI (default ALL)
    """
    if not BPS_API_KEY:
        return Response({
            "error":   "BPS Web API Key belum dikonfigurasi",
            "message": "Tambahkan BPS_WEB_API_KEY di file .env",
        }, status=500)

    try:
        tahun     = request.data.get('tahun', 2024)
        indikator = request.data.get('indikator', 'ALL')

        try:
            tahun = int(tahun)
        except (ValueError, TypeError):
            tahun = 2024

        if tahun not in TAHUN_BPS_MAP:
            return Response({"error": f"Tahun {tahun} tidak didukung."}, status=400)

        if indikator not in INDIKATOR_DATASET_MAP:
            indikator = 'ALL'

        keys_aktif = INDIKATOR_DATASET_MAP[indikator]
        analytics  = EkonomiAnalytics(tahun=tahun)

        print(f"\n=== MULAI FETCH EKONOMI BPS | TAHUN={tahun} | INDIKATOR={indikator} ===")
        print(f"    Dataset: {keys_aktif}")

        raw_data = analytics.fetch_selected_data(keys_aktif)

        print("\n=== PARSE DATA ===")
        empty = ({}, {})

        pdrb_values,  pdrb_raw  = analytics.parse_province_data(raw_data.get('PDRB'),       'PDRB')       if 'PDRB'       in keys_aktif else empty
        kem_values,   kem_raw   = analytics.parse_province_data(raw_data.get('KEMISKINAN'), 'KEMISKINAN') if 'KEMISKINAN' in keys_aktif else empty
        inv_values,   inv_raw   = analytics.parse_province_data(raw_data.get('INVESTASI'),  'INVESTASI')  if 'INVESTASI'  in keys_aktif else empty

        print("\n=== LOAD BOUNDARY DATA ===")
        cursor            = mongo_db["batas_provinsi"].find({}, {'_id': 0})
        boundary_features = list(cursor)

        province_map = {}
        for feature in boundary_features:
            props = feature.get('properties', {})
            for field in ['NAMOBJ', 'name', 'WADMPR', 'Provinsi']:
                if field in props and props[field]:
                    official   = str(props[field]).upper().strip()
                    normalized = normalize_province_name(official)
                    province_map[normalized] = feature
                    province_map[official]   = feature

        all_provinces = set()
        if 'PDRB'       in keys_aktif: all_provinces.update(pdrb_values.keys())
        if 'KEMISKINAN' in keys_aktif: all_provinces.update(kem_values.keys())
        if 'INVESTASI'  in keys_aktif: all_provinces.update(inv_values.keys())

        print(f"\n=== PROCESSING {len(all_provinces)} PROVINCES | INDIKATOR={indikator} ===")

        matched_features = []
        analysis_summary = []
        kategori_counts  = {"MAJU": 0, "BERKEMBANG": 0, "TERTINGGAL": 0}

        for prov_name in sorted(all_provinces):
            data_ekonomi = {
                "PDRB":       pdrb_values.get(prov_name)  if 'PDRB'       in keys_aktif else None,
                "KEMISKINAN": kem_values.get(prov_name)   if 'KEMISKINAN' in keys_aktif else None,
                "INVESTASI":  inv_values.get(prov_name)   if 'INVESTASI'  in keys_aktif else None,
            }
            if not any(v is not None for v in data_ekonomi.values()):
                continue

            scores          = analytics.calculate_scores(data_ekonomi, indikator)
            kategori, warna = analytics.categorize_province(scores['skor_total'])
            insights        = analytics.generate_insights(prov_name, data_ekonomi, kategori, scores['skor_total'], indikator)

            kategori_list   = build_kategori_list(kategori, data_ekonomi, analytics.indikator_config)
            recommendations = get_bank_kebijakan2_by_kategori(kategori_list, limit_per_kategori=8)

            normalized_prov = normalize_province_name(prov_name)
            matched_feature = (province_map.get(normalized_prov) or province_map.get(prov_name))
            if not matched_feature:
                for map_name, feat in province_map.items():
                    if normalized_prov in map_name or map_name in normalized_prov:
                        matched_feature = feat
                        break
            if not matched_feature:
                print(f"  ✗ {prov_name}: no boundary match")
                continue

            kategori_counts[kategori] += 1
            feature_copy = matched_feature.copy()
            props        = feature_copy.get('properties', {})
            props['ekonomi_analysis'] = {
                'nama_provinsi':    prov_name,
                'indikator':        indikator,
                'kategori':         kategori,
                'warna':            warna,
                'skor_total':       scores['skor_total'],
                'skor_pdrb':        scores['skor_pdrb'],
                'skor_kemiskinan':  scores['skor_kemiskinan'],
                'skor_investasi':   scores['skor_investasi'],
                'insights':         insights,
                'rekomendasi':      recommendations,
                'kategori_applied': kategori_list,
                'data_ekonomi':     {
                    'PDRB':       data_ekonomi.get('PDRB'),
                    'KEMISKINAN': data_ekonomi.get('KEMISKINAN'),
                    'INVESTASI':  data_ekonomi.get('INVESTASI'),
                },
            }
            feature_copy['properties'] = props
            matched_features.append(feature_copy)

            analysis_summary.append({
                'provinsi':    prov_name,
                'indikator':   indikator,
                'kategori':    kategori,
                'warna':       warna,
                'skor_total':  scores['skor_total'],
                'pdrb':        data_ekonomi.get('PDRB'),
                'kemiskinan':  data_ekonomi.get('KEMISKINAN'),
                'investasi':   data_ekonomi.get('INVESTASI'),
            })
            print(f"  ✓ {prov_name}: {kategori} (Skor: {scores['skor_total']})")

        # Rangkuman nasional
        sorted_summary  = sorted(
            [s for s in analysis_summary if s['skor_total'] is not None],
            key=lambda x: x['skor_total'],
        )
        worst_provinces = sorted_summary[:5]
        best_provinces  = sorted_summary[-5:][::-1]

        national_recommendations = []
        if kategori_counts["TERTINGGAL"] > 0:
            national_recommendations.append({
                "priority": "Darurat",
                "title":    f"Fokus Pembangunan {kategori_counts['TERTINGGAL']} Provinsi Tertinggal",
                "content":  f"{kategori_counts['TERTINGGAL']} provinsi memerlukan percepatan pembangunan ekonomi.",
            })
        if kategori_counts["BERKEMBANG"] > 0:
            national_recommendations.append({
                "priority": "Tinggi",
                "title":    f"Penguatan {kategori_counts['BERKEMBANG']} Provinsi Berkembang",
                "content":  f"{kategori_counts['BERKEMBANG']} provinsi dalam perjalanan menuju status maju.",
            })
        if kategori_counts["MAJU"] > 0:
            national_recommendations.append({
                "priority": "Pemeliharaan",
                "title":    f"Sustain {kategori_counts['MAJU']} Provinsi Maju",
                "content":  f"{kategori_counts['MAJU']} provinsi dengan ekonomi yang kuat.",
            })

        print(f"\n=== ANALYSIS COMPLETE | {len(matched_features)} provinces ===")

        raw_datasets = {
            'timestamp':  analytics.timestamp_fetch,
            'tahun':      tahun,
            'indikator':  indikator,
            'PDRB':       pdrb_raw  if 'PDRB'       in keys_aktif else {},
            'KEMISKINAN': kem_raw   if 'KEMISKINAN' in keys_aktif else {},
            'INVESTASI':  inv_raw   if 'INVESTASI'  in keys_aktif else {},
        }

        return Response({
            'status':                   'success',
            'source':                   'BPS Web API + PostgreSQL Bank Kebijakan',
            'tahun':                    tahun,
            'indikator':                indikator,
            'dataset_aktif':            keys_aktif,
            'total_success':            len(matched_features),
            'kategori_distribusi':      kategori_counts,
            'timestamp':                analytics.timestamp_fetch,
            'matched_features': {
                'type':     'FeatureCollection',
                'features': matched_features,
            },
            'analysis_summary':         analysis_summary,
            'national_recommendations': national_recommendations,
            'worst_provinces':          worst_provinces,
            'best_provinces':           best_provinces,
            'colors':                   EkonomiAnalytics.COLORS,
            'indikator_info': {
                k: {
                    "nama":             v["nama"],
                    "satuan":           v["satuan"],
                    "penjelasan":       v["penjelasan"],
                    "bobot":            v["bobot"],
                }
                for k, v in analytics.indikator_config.items()
            },
            'raw_datasets': raw_datasets,
        })

    except Exception as e:
        import traceback; traceback.print_exc()
        return Response({'error': str(e), 'message': 'Gagal menganalisis data ekonomi dari BPS'}, status=500)


# ═══════════════════════════════════════════════════════════════════════════════
# ENDPOINT: DATA HISTORIS UNTUK GRAFIK TREN (dipertahankan)
# ═══════════════════════════════════════════════════════════════════════════════

@api_view(["GET"])
def get_historis_ekonomi(request):
    """
    GET /api/historis-ekonomi/
    Query params:
      - provinsi     : nama provinsi (opsional; kosong = agregat nasional)
      - tahun_mulai  : 2020 (default)
      - tahun_akhir  : 2024 (default)
    """
    if not BPS_API_KEY:
        return Response({"error": "BPS_WEB_API_KEY belum dikonfigurasi"}, status=500)

    provinsi_req = request.GET.get("provinsi", "").upper().strip()
    tahun_mulai  = int(request.GET.get("tahun_mulai", 2020))
    tahun_akhir  = int(request.GET.get("tahun_akhir", 2024))

    # Kumpulkan th_codes yang diperlukan
    th_codes = [
        th for tahun, th in TAHUN_BPS_MAP.items()
        if tahun_mulai <= tahun <= tahun_akhir
    ]
    th_codes.sort(key=lambda th: int(th))   # urut lama → baru

    try:
        tren = []
        for th in th_codes:
            tahun_kal  = TH_CODE_TO_TAHUN.get(th, int(th))
            analytics  = EkonomiAnalytics(tahun=tahun_kal)
            raw_data   = analytics.fetch_all_data()

            pdrb_vals, _ = analytics.parse_province_data(raw_data.get('PDRB'),       'PDRB')
            kem_vals,  _ = analytics.parse_province_data(raw_data.get('KEMISKINAN'), 'KEMISKINAN')
            inv_vals,  _ = analytics.parse_province_data(raw_data.get('INVESTASI'),  'INVESTASI')

            if provinsi_req:
                norm_req = normalize_province_name(provinsi_req)
                pdrb  = pdrb_vals.get(norm_req) or pdrb_vals.get(provinsi_req)
                kem   = kem_vals.get(norm_req)  or kem_vals.get(provinsi_req)
                inv   = inv_vals.get(norm_req)  or inv_vals.get(provinsi_req)
                tren.append({
                    "tahun":     tahun_kal,
                    "pdrb":      pdrb,
                    "kemiskinan":kem,
                    "investasi": inv,
                })
            else:
                pdrb_list = [v for v in pdrb_vals.values() if v is not None]
                kem_list  = [v for v in kem_vals.values()  if v is not None]
                inv_list  = [v for v in inv_vals.values()  if v is not None]
                tren.append({
                    "tahun":      tahun_kal,
                    "pdrb":       round(sum(pdrb_list) / len(pdrb_list), 2) if pdrb_list else None,
                    "kemiskinan": round(sum(kem_list)  / len(kem_list),  2) if kem_list  else None,
                    "investasi":  round(sum(inv_list)  / len(inv_list),  2) if inv_list  else None,
                })

        return Response({
            "status":   "success",
            "provinsi": provinsi_req or "NASIONAL",
            "tren":     tren,
        })

    except Exception as e:
        import traceback; traceback.print_exc()
        return Response({"error": str(e)}, status=500)


# ═══════════════════════════════════════════════════════════════════════════════
# CRUD: SIMPAN / LIST / DETAIL / HAPUS HASIL ANALISIS
# ═══════════════════════════════════════════════════════════════════════════════

@api_view(["POST"])
def save_ekonomi_analysis(request):
    try:
        analysis_name = request.data.get("name", "Analisis Ekonomi Tanpa Nama")
        analysis_data = request.data.get("analysis_data")
        if not analysis_data:
            return Response({"error": "Data analisis tidak ditemukan"}, status=400)
        analysis_id = str(uuid.uuid4())
        document = {
            "analysis_id": analysis_id,
            "name":        analysis_name,
            "type":        "ekonomi",
            "timestamp":   datetime.now().isoformat(),
            **analysis_data,
        }
        mongo_db["ekonomi_analysis"].insert_one(document)
        return Response({
            "status":      "success",
            "message":     f"Analisis '{analysis_name}' berhasil disimpan",
            "analysis_id": analysis_id,
            "saved_at":    document["timestamp"],
        })
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@api_view(["GET"])
def get_ekonomi_analysis_list(request):
    try:
        cursor = mongo_db["ekonomi_analysis"].find(
            {},
            {"_id": 0, "analysis_id": 1, "name": 1, "timestamp": 1,
             "total_success": 1, "kategori_distribusi": 1, "tahun": 1, "indikator": 1}
        ).sort("timestamp", -1)
        results = list(cursor)
        return Response({"status": "success", "count": len(results), "results": results})
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@api_view(["GET"])
def get_ekonomi_analysis_detail(request, analysis_id):
    try:
        result = mongo_db["ekonomi_analysis"].find_one(
            {"analysis_id": analysis_id}, {"_id": 0}
        )
        if not result:
            return Response({"error": "Analisis tidak ditemukan"}, status=404)
        return Response(result)
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@api_view(["DELETE"])
def delete_ekonomi_analysis(request, analysis_id):
    try:
        result = mongo_db["ekonomi_analysis"].delete_one({"analysis_id": analysis_id})
        if result.deleted_count == 0:
            return Response({"error": "Analisis tidak ditemukan"}, status=404)
        return Response({"status": "success", "message": "Analisis berhasil dihapus"})
    except Exception as e:
        return Response({"error": str(e)}, status=500)