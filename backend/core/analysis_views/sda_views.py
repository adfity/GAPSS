from rest_framework.decorators import api_view
from rest_framework.response import Response
from pymongo import MongoClient
import uuid, requests, os, json
import numpy as np
import pandas as pd
from dotenv import load_dotenv
from datetime import datetime
import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

MONGO_URI     = os.getenv("MONGO_URI")
DB_MONGO_NAME = os.getenv("DB_MONGO_NAME")
BPS_API_KEY   = os.getenv("BPS_WEB_API_KEY")

client   = MongoClient(MONGO_URI)
mongo_db = client[DB_MONGO_NAME]

# ── AI Model (lazy load) ──────────────────────────────────────────────────────
_AI_MODELS = _AI_ENCODERS = _AI_META = None
AI_MODEL_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ai_models", "sda")

def _load_ai_models():
    global _AI_MODELS, _AI_ENCODERS, _AI_META
    if _AI_MODELS is not None:
        return True
    try:
        import joblib
        meta_path = os.path.join(AI_MODEL_DIR, "metadata.json")
        if not os.path.exists(meta_path):
            return False
        with open(meta_path, "r", encoding="utf-8") as f:
            _AI_META = json.load(f)
        _AI_MODELS   = {t: joblib.load(os.path.join(AI_MODEL_DIR, f"model_{t}.pkl"))
                        for t in _AI_META["targets"]}
        _AI_ENCODERS = {
            "provinsi": joblib.load(os.path.join(AI_MODEL_DIR, "encoder_provinsi.pkl")),
            "pulau":    joblib.load(os.path.join(AI_MODEL_DIR, "encoder_pulau.pkl")),
        }
        return True
    except Exception as e:
        print(f"[AI-SDA] Gagal load model: {e}")
        _AI_MODELS = _AI_ENCODERS = _AI_META = None
        return False


# ── Konfigurasi BPS ───────────────────────────────────────────────────────────
TAHUN_SUPPORTED = list(range(2018, 2026))

# ── ID Tabel SIMDASI Perikanan ────────────────────────────────────────────────
# Format lama (2018-2022): kolom nilai ribu-Rp, unit_multiplier=3
SIMDASI_NILAI_IKAN_TABEL = {y: "NTdHM1BiQXJXUHAyUUNoMXNabkRwZz09" for y in range(2018, 2023)}
# Format baru (2023-2024): kolom nilai Rupiah (tanpa multiplier), volume kg
SIMDASI_VOLUME_IKAN_TABEL = {y: "Si8wS0pVcDhRUTJidGFnSzl5UDIxZz09" for y in range(2023, 2026)}

# ── URL Templates ─────────────────────────────────────────────────────────────
URL_IKAN_LAUT       = ("https://webapi.bps.go.id/v1/api/list/model/data/lang/ind"
                       "/domain/0000/var/1054/th/{th_val}/key/{key}/")
URL_PERKEBUNAN_LAMA = ("https://webapi.bps.go.id/v1/api/list/model/data/lang/ind"
                       "/domain/0000/var/132/th/{th_val}/key/{key}/")
URL_PERKEBUNAN_BARU = ("https://webapi.bps.go.id/v1/api/list/model/data/lang/ind"
                       "/domain/0000/var/2566/th/{th_val}/key/{key}/")
URL_NILAI_IKAN_SIMDASI = ("https://webapi.bps.go.id/v1/api/interoperabilitas/datasource/simdasi"
                           "/id/25/tahun/{tahun}/id_tabel/{id_tabel}/wilayah/0000000/key/{key}/")
URL_PENDUDUK_SDA    = ("https://webapi.bps.go.id/v1/api/list/model/data/lang/ind"
                       "/domain/7100/var/958/th/{th_val}/key/{key}/")
URL_PDRB_LAPANGAN   = ("https://webapi.bps.go.id/v1/api/list/model/data/lang/ind"
                       "/domain/0000/var/2268/th/{th_val}/key/{key}/")

# ── Kode komoditas perkebunan ─────────────────────────────────────────────────
KOMODITAS_LAMA = [252, 253, 254, 255, 256, 257, 258, 259]   # var=132,  8 komoditas
KOMODITAS_BARU = [2321, 2322, 2323, 2324, 2325, 2326, 2327] # var=2566, 7 komoditas

DATASET_LABELS = {
    'IKAN':       'Produksi Perikanan Laut per Provinsi',
    'PERKEBUNAN': 'Produksi Tanaman Perkebunan (7-8 Komoditas)',
    'NILAI_IKAN': 'Nilai Produksi Perikanan Tangkap per Provinsi',
    'PENDUDUK':   'Jumlah Penduduk per Provinsi',
    'PDRB':       'PDRB Lapangan Usaha (Sektor Pertanian & Perikanan)',
}
ALL_DATASETS = ['IKAN', 'PERKEBUNAN', 'NILAI_IKAN', 'PENDUDUK', 'PDRB']


def _tahun_to_th(tahun: int) -> int:
    return tahun - 1900


def get_sda_config(tahun: int) -> dict:
    th_val = _tahun_to_th(tahun)
    config = {}

    config["IKAN"] = {
        "url": URL_IKAN_LAUT.format(th_val=th_val, key=BPS_API_KEY),
        "nama": DATASET_LABELS["IKAN"], "jenis": "list_ikan",
    }

    if tahun <= 2023:
        config["PERKEBUNAN"] = {
            "url": URL_PERKEBUNAN_LAMA.format(th_val=th_val, key=BPS_API_KEY),
            "nama": DATASET_LABELS["PERKEBUNAN"], "jenis": "list_perkebunan_lama",
            "komoditas_codes": KOMODITAS_LAMA, "var_str": "132",
        }
    else:
        config["PERKEBUNAN"] = {
            "url": URL_PERKEBUNAN_BARU.format(th_val=th_val, key=BPS_API_KEY),
            "nama": DATASET_LABELS["PERKEBUNAN"], "jenis": "list_perkebunan_baru",
            "komoditas_codes": KOMODITAS_BARU, "var_str": "2566",
        }

    # Nilai Ikan — SIMDASI dua format
    if tahun <= 2022:
        id_tabel = SIMDASI_NILAI_IKAN_TABEL.get(tahun)
        if id_tabel:
            config["NILAI_IKAN"] = {
                "url": URL_NILAI_IKAN_SIMDASI.format(tahun=tahun, id_tabel=id_tabel, key=BPS_API_KEY),
                "nama": DATASET_LABELS["NILAI_IKAN"], "jenis": "simdasi_nilai_ikan_lama",
                # Kolom aktual dari JSON 2022: ribu-Rp, unit_multiplier=3 → ×1000
                "col_nilai_laut":  "uhzcmnxcue",
                "col_nilai_total": "s1dgw0lol3",
            }
    else:
        id_tabel = SIMDASI_VOLUME_IKAN_TABEL.get(tahun)
        if id_tabel:
            config["NILAI_IKAN"] = {
                "url": URL_NILAI_IKAN_SIMDASI.format(tahun=tahun, id_tabel=id_tabel, key=BPS_API_KEY),
                "nama": DATASET_LABELS["NILAI_IKAN"], "jenis": "simdasi_nilai_ikan_baru",
                # Kolom aktual dari JSON 2024: Rupiah, tidak ada unit_multiplier → ×1
                "col_nilai_laut":  "cij3qwjfji",
                "col_nilai_total": "un4872dm6h",
                "col_vol_laut":    "ph6llk7fuz",  # kg, bagi 1000 → ton
            }

    config["PENDUDUK"] = {
        "url": URL_PENDUDUK_SDA.format(th_val=th_val, key=BPS_API_KEY),
        "nama": DATASET_LABELS["PENDUDUK"], "jenis": "list_penduduk_sda",
    }
    config["PDRB"] = {
        "url": URL_PDRB_LAPANGAN.format(th_val=th_val, key=BPS_API_KEY),
        "nama": DATASET_LABELS["PDRB"], "jenis": "list_pdrb",
    }
    return config


def _is_sda_data_empty(data, dataset_key):
    if data is None:
        return True
    if dataset_key in ("IKAN", "PERKEBUNAN", "PENDUDUK", "PDRB"):
        return not bool(data.get("datacontent", {}))
    elif dataset_key == "NILAI_IKAN":
        dc = data.get("data", [])
        if not dc or len(dc) < 2:
            return True
        return not isinstance(dc[1], dict) or len(dc[1].get("data", [])) == 0
    return True


# ── Endpoint: Cek Ketersediaan Data ──────────────────────────────────────────
@api_view(['POST'])
def check_year_data_sda(request):
    if not BPS_API_KEY:
        return Response({"error": "BPS Web API Key belum dikonfigurasi"}, status=500)
    try:
        tahun = int(request.data.get('tahun', 2024))
    except (ValueError, TypeError):
        tahun = 2024
    if tahun not in TAHUN_SUPPORTED:
        return Response({"error": f"Tahun {tahun} tidak didukung."}, status=400)

    all_config     = get_sda_config(tahun)
    dataset_status = {}
    for key in ALL_DATASETS:
        cfg = all_config.get(key)
        if not cfg:
            dataset_status[key] = {"nama": DATASET_LABELS.get(key, key),
                                   "tersedia": False, "status": "Konfigurasi tidak ada"}
            continue
        try:
            resp = requests.get(cfg["url"], timeout=20)
            if resp.status_code == 200:
                kosong = _is_sda_data_empty(resp.json(), key)
                dataset_status[key] = {"nama": cfg["nama"], "tersedia": not kosong,
                                       "status": "Tersedia" if not kosong else "Kosong"}
            else:
                dataset_status[key] = {"nama": cfg["nama"], "tersedia": False,
                                       "status": f"HTTP {resp.status_code}"}
        except Exception as e:
            dataset_status[key] = {"nama": cfg.get("nama", key), "tersedia": False,
                                   "status": f"Gagal ({str(e)[:50]})"}

    tersedia_list   = [k for k, v in dataset_status.items() if v["tersedia"]]
    kosong_list     = [k for k, v in dataset_status.items() if not v["tersedia"]]
    semua_kosong    = len(tersedia_list) == 0
    ada_yang_kosong = len(kosong_list) > 0 and not semua_kosong

    ai_tersedia = ai_model_ready = ai_version = None
    if semua_kosong or ada_yang_kosong:
        ai_model_ready = _load_ai_models()
        ai_tersedia    = ai_model_ready
        ai_version     = _AI_META.get("version") if _AI_META else None

    return Response({
        "tahun": tahun, "is_ai_prediction": False,
        "dataset_status": dataset_status,
        "tersedia": tersedia_list, "kosong": kosong_list,
        "semua_kosong": semua_kosong, "ada_yang_kosong": ada_yang_kosong,
        "bisa_dieksekusi": len(tersedia_list) >= 3,
        "bps_kosong": semua_kosong,
        "ai_tersedia": ai_tersedia, "ai_model_ready": ai_model_ready,
        "ai_model_version": ai_version,
    })


# ── Helpers ───────────────────────────────────────────────────────────────────
def normalize_province_name(name):
    if not isinstance(name, str):
        name = str(name)
    name = name.upper().strip()
    for prefix in ['PROVINSI ', 'PROV. ', 'PROV ', 'DAERAH KHUSUS IBUKOTA ']:
        if name.startswith(prefix):
            name = name[len(prefix):]
    MAPPINGS = {
        'DKI JAKARTA': 'JAKARTA', 'DAERAH KHUSUS IBUKOTA JAKARTA': 'JAKARTA', 'DKI': 'JAKARTA',
        'YOGYAKARTA': 'DAERAH ISTIMEWA YOGYAKARTA', 'DIY': 'DAERAH ISTIMEWA YOGYAKARTA',
        'D.I. YOGYAKARTA': 'DAERAH ISTIMEWA YOGYAKARTA', 'DI YOGYAKARTA': 'DAERAH ISTIMEWA YOGYAKARTA',
        'BANGKA BELITUNG': 'KEPULAUAN BANGKA BELITUNG', 'KEP. BANGKA BELITUNG': 'KEPULAUAN BANGKA BELITUNG',
        'KEP. RIAU': 'KEPULAUAN RIAU', 'NTB': 'NUSA TENGGARA BARAT', 'NTT': 'NUSA TENGGARA TIMUR',
    }
    for k, v in MAPPINGS.items():
        if name == k or name.endswith(' ' + k):
            return v
    if 'KEP.' in name:
        name = name.replace('KEP.', 'KEPULAUAN')
    return name.strip()


def minmax_normalize(values_dict):
    valid = {k: v for k, v in values_dict.items() if v is not None and v > 0}
    if not valid:
        return {k: 0.0 for k in values_dict}
    v_min, v_max = min(valid.values()), max(valid.values())
    return {
        k: (1.0 if v_max == v_min else round((v - v_min) / (v_max - v_min), 4))
        if (v and v > 0) else 0.0
        for k, v in values_dict.items()
    }


def calculate_ipsda(rpp_ikan_n, rpp_perkebunan_n, npi_n, kps_n):
    return round((rpp_ikan_n + rpp_perkebunan_n + npi_n + kps_n) / 4, 4)


def classify_ipsda(ipsda):
    if ipsda >= 0.70: return "OPTIMAL", "#10b981"
    if ipsda >= 0.50: return "CUKUP",   "#f59e0b"
    if ipsda >= 0.30: return "KURANG",  "#f97316"
    return "RENDAH",                    "#ef4444"


def generate_sda_insights(provinsi, rpp_ikan, rpp_perkebunan, npi, kps, ipsda, status):
    out = [f"Provinsi {provinsi} memiliki status pemerataan SDA {status} dengan IPSDA {ipsda:.3f}."]
    if rpp_ikan is not None:
        if rpp_ikan > 0.05:   out.append(f"✅ Rasio produksi ikan tinggi ({rpp_ikan:.4f} ton/jiwa).")
        elif rpp_ikan > 0.01: out.append(f"📊 Rasio produksi ikan sedang ({rpp_ikan:.4f} ton/jiwa).")
        else:                 out.append(f"⚠️ Rasio produksi ikan rendah ({rpp_ikan:.4f} ton/jiwa).")
    if rpp_perkebunan is not None:
        if rpp_perkebunan > 0.5:   out.append(f"✅ Produksi perkebunan per kapita tinggi ({rpp_perkebunan:.4f} ton/jiwa).")
        elif rpp_perkebunan > 0.1: out.append(f"📊 Produksi perkebunan per kapita sedang ({rpp_perkebunan:.4f} ton/jiwa).")
        else:                      out.append(f"📉 Produksi perkebunan per kapita rendah ({rpp_perkebunan:.4f} ton/jiwa).")
    if npi is not None:
        if npi > 1_000_000:   out.append(f"✅ Nilai produksi ikan per kapita tinggi (Rp {npi:,.0f}/jiwa).")
        elif npi > 100_000:   out.append(f"📊 Nilai produksi ikan per kapita sedang (Rp {npi:,.0f}/jiwa).")
        else:                 out.append(f"⚠️ Nilai produksi ikan per kapita rendah (Rp {npi:,.0f}/jiwa).")
    if kps is not None:
        if kps >= 0.30:   out.append(f"✅ Kontribusi sektor pertanian & perikanan {kps*100:.1f}% terhadap PDRB.")
        elif kps >= 0.15: out.append(f"📊 Kontribusi pertanian & perikanan {kps*100:.1f}% — dapat ditingkatkan.")
        else:             out.append(f"📉 Kontribusi pertanian & perikanan hanya {kps*100:.1f}% terhadap PDRB.")
    return out


def generate_sda_recommendations(status, rpp_ikan, rpp_perkebunan, npi, kps):
    RECS = {
        "OPTIMAL": {"title": "Pertahankan & Kembangkan Inovasi SDA", "priority": "Rendah",
                    "actions": ["Ekspansi pasar ekspor produk perikanan & perkebunan",
                                "Implementasi smart farming & aquaculture teknologi tinggi",
                                "Kembangkan industri hilir pengolahan hasil SDA",
                                "Perkuat cadangan & buffer stock daerah"]},
        "CUKUP":   {"title": "Penguatan Pemanfaatan SDA", "priority": "Sedang",
                    "actions": ["Intensifikasi tangkap ikan dengan armada modern",
                                "Perluasan kebun dan diversifikasi komoditas",
                                "Pelatihan petani/nelayan & akses modal usaha",
                                "Optimasi kontribusi sektor primer ke PDRB"]},
        "KURANG":  {"title": "Percepatan Pemanfaatan SDA", "priority": "Tinggi",
                    "actions": ["Investasi infrastruktur pelabuhan & cold storage",
                                "Subsidi sarana produksi perikanan & perkebunan",
                                "Pemberdayaan koperasi nelayan & petani kebun",
                                "Kemitraan dengan off-taker industri"]},
        "RENDAH":  {"title": "Intervensi Mendesak Sektor SDA", "priority": "Sangat Tinggi",
                    "actions": ["Program khusus percepatan produktivitas perikanan",
                                "Redistribusi akses lahan & izin pengelolaan SDA",
                                "Bantuan langsung sarana produksi (jaring, bibit, pupuk)",
                                "Koneksi ke program nasional (KUR, BUMN pangan)"]},
    }
    recs = [RECS.get(status, RECS["KURANG"])]
    if npi is not None and npi < 500_000:
        recs.append({"priority": "Tinggi", "title": "Peningkatan Nilai Tambah Perikanan",
                     "actions": ["Pengembangan unit pengolahan ikan (UPI)",
                                 "Sertifikasi produk perikanan untuk pasar premium",
                                 "Digitalisasi pemasaran hasil tangkap"]})
    if kps is not None and kps < 0.15:
        recs.append({"priority": "Tinggi", "title": "Peningkatan Kontribusi Sektor Primer ke PDRB",
                     "actions": ["Insentif investasi agroindustri daerah",
                                 "Pengembangan kawasan ekonomi khusus berbasis SDA"]})
    return recs


# ── Analytics Class ───────────────────────────────────────────────────────────
class SdaAnalytics:
    def __init__(self, tahun=2024):
        self.tahun  = tahun
        self.config = get_sda_config(tahun)
        self.timestamp_fetch = None

    def fetch_all_data(self) -> dict:
        all_data = {}
        self.timestamp_fetch = datetime.now().isoformat()
        for key, cfg in self.config.items():
            try:
                resp = requests.get(cfg["url"], timeout=30)
                all_data[key] = resp.json() if resp.status_code == 200 else None
                print(f"{'✓' if resp.status_code == 200 else '✗'} SDA-{key}: {resp.status_code}")
            except Exception as e:
                print(f"✗ SDA-{key}: {e}"); all_data[key] = None
        return all_data

    def _clean_float(self, raw):
        if raw is None: return None
        s = str(raw).replace('\xa0', '').replace(' ', '').replace('.', '').replace(',', '.').strip()
        if s in ['-', '...', '', 'NA', '–']: return None
        try: return float(s)
        except ValueError: return None

    def _clean_prov(self, raw):
        if not raw: return None
        raw = raw.strip()
        if raw.upper() in ('INDONESIA', 'TOTAL', 'NASIONAL', 'JUMLAH'): return None
        if any(x in raw.lower() for x in ['kab.', 'kota ', 'kabupaten', 'kecamatan']): return None
        return normalize_province_name(raw)

    # ── 1. Produksi Ikan Laut (list model, var=1054) ─────────────────────────
    def parse_ikan_data(self, raw_data: dict) -> dict:
        """
        datacontent key dari JSON aktual: "{vervar_val}105401{th_val}0"
        Contoh: "9600105401240" = vervar 9600 + var 1054 + turvar 0 + th 124 + turtahun 0
        Vervar menggunakan kode wilayah BPS 4-digit (1100, 1200, ..., 9700, 9999)
        """
        result = {}
        if not raw_data or not raw_data.get("datacontent"):
            return result
        vervar_list = raw_data.get("vervar", [])
        datacontent = raw_data.get("datacontent", {})
        th_val      = raw_data.get("tahun", [{}])[0].get("val", _tahun_to_th(self.tahun))
        turvar_list = raw_data.get("turvar", [])
        tv_val      = str(turvar_list[0].get("val", 0)) if turvar_list else "0"

        for item in vervar_list:
            vv    = item.get("val")
            label = item.get("label", "").strip()
            if not label or label.upper() in ("INDONESIA", "TOTAL", "NASIONAL") or vv == 9999:
                continue
            prov = normalize_province_name(label)
            val  = None
            for key_try in [f"{vv}1054{tv_val}{th_val}0",
                             f"{vv}10540{th_val}0",
                             f"{vv}1054{th_val}0"]:
                if key_try in datacontent:
                    val = datacontent[key_try]; break
            if val is None:
                vv_str = str(vv)
                for k, v in datacontent.items():
                    if k.startswith(vv_str) and "1054" in k and vv_str != "9999":
                        val = v; break
            if val is None: continue
            cleaned = self._clean_float(val)
            if cleaned is not None and cleaned >= 0:
                result[prov] = float(cleaned)  # Ton
        return result

    # ── 2. Produksi Perkebunan (list model, var=132 atau 2566) ───────────────
    def parse_perkebunan_data(self, raw_data: dict, cfg: dict) -> dict:
        """
        Key: "{vervar}{var}{turvar}{th_val}0"
        var=132  contoh: "16001322521230" = 1600+132+252+123+0 = 4130.2 (ribu ton)
        var=2566 contoh: "1400256623211240" = 1400+2566+2321+124+0 = 9136.1 (ribu ton)
        Hasil: rata-rata semua komoditas × 1000 → Ton
        """
        result = {}
        if not raw_data or not raw_data.get("datacontent"):
            return result
        vervar_list    = raw_data.get("vervar", [])
        datacontent    = raw_data.get("datacontent", {})
        th_val         = raw_data.get("tahun", [{}])[0].get("val", _tahun_to_th(self.tahun))
        var_str        = cfg.get("var_str", "132")
        komoditas_vals = cfg.get("komoditas_codes", KOMODITAS_LAMA)

        for item in vervar_list:
            vv    = item.get("val")
            label = item.get("label", "").strip()
            if not label or label.upper() in ("INDONESIA", "TOTAL", "NASIONAL") or vv == 9999:
                continue
            prov    = normalize_province_name(label)
            total   = 0.0
            n_valid = 0
            for tv in komoditas_vals:
                key = f"{vv}{var_str}{tv}{th_val}0"
                val = datacontent.get(key)
                if val is not None:
                    cleaned = self._clean_float(val)
                    if cleaned is not None and cleaned >= 0:
                        total += cleaned; n_valid += 1
            if n_valid > 0:
                result[prov] = (total / n_valid) * 1000  # Ribu Ton → Ton
        return result

    # ── 3. Nilai Produksi Ikan (SIMDASI dua format) ───────────────────────────
    def parse_nilai_ikan_data(self, raw_data: dict, cfg: dict) -> dict:
        """
        Format SIMDASI: raw_data["data"][1]["data"] = list baris provinsi
        Setiap baris: {"label": ..., "kode_wilayah": ..., "variables": {"<col>": {"value_raw": ...}}}

        Format lama (2018-2022): nilai dalam ribu-Rp → unit_multiplier=3 → ×1000 → Rupiah
          col_nilai_laut="uhzcmnxcue", col_nilai_total="s1dgw0lol3"

        Format baru (2023-2024): nilai langsung Rupiah (tidak ada unit_multiplier)
          col_nilai_laut="cij3qwjfji", col_nilai_total="un4872dm6h"
          col_vol_laut="ph6llk7fuz" (kg)
        """
        result = {}
        if not raw_data: return result
        data_list = raw_data.get("data", [])
        if len(data_list) < 2: return result
        block = data_list[1]
        if not isinstance(block, dict): return result

        rows      = block.get("data", [])
        kolom_map = block.get("kolom", {})

        col_nilai_laut  = cfg.get("col_nilai_laut")
        col_nilai_total = cfg.get("col_nilai_total")

        def get_unit_mult(col_id):
            if col_id and col_id in kolom_map:
                um = kolom_map[col_id].get("unit_multiplier")
                if um is not None: return 10 ** int(um)
            return 1

        um_laut  = get_unit_mult(col_nilai_laut)
        um_total = get_unit_mult(col_nilai_total)

        for row in rows:
            if not isinstance(row, dict): continue
            label = (row.get("label") or row.get("label_raw") or "").strip()
            prov  = self._clean_prov(label)
            if not prov: continue
            kode = str(row.get("kode_wilayah", ""))
            if kode in ("0", "0000000", "00000000"): continue

            variables = row.get("variables", {})
            if not isinstance(variables, dict): continue

            def get_val(col_id, unit_mult):
                if not col_id: return None
                entry = variables.get(col_id, {})
                if not isinstance(entry, dict): return None
                raw = entry.get("value_raw") or entry.get("value")
                c   = self._clean_float(raw)
                return None if (c is None or c < 0) else c * unit_mult

            # Prioritaskan nilai total, fallback ke nilai laut
            nilai = get_val(col_nilai_total, um_total)
            if nilai is None:
                nilai = get_val(col_nilai_laut, um_laut)
            if nilai is None: continue
            result[prov] = float(nilai)  # Rupiah
        return result

    # ── 4. Jumlah Penduduk (list model, var=958, domain=7100) ────────────────
    def parse_penduduk_data(self, raw_data: dict) -> dict:
        """
        vervar val = 1..34 (urut provinsi; val=35 = Indonesia → skip)
        turvar: [{"val": "0", "label": "Tidak ada"}] → tv_val = "0"
        Key aktual: "{vv}958{tv_val}{th_val}0"
        Contoh: "195801240" = vv=1 (Aceh) + 958 + tv=0 + th=124 + turtahun=0
        Unit: Ribu Jiwa → ×1000 → Jiwa
        """
        result = {}
        if not raw_data or not raw_data.get("datacontent"):
            return result
        vervar_list = raw_data.get("vervar", [])
        datacontent = raw_data.get("datacontent", {})
        th_val      = raw_data.get("tahun", [{}])[0].get("val", _tahun_to_th(self.tahun))
        turvar_list = raw_data.get("turvar", [])
        tv_val      = str(turvar_list[0].get("val", 0)) if turvar_list else "0"

        for item in vervar_list:
            vv    = item.get("val")
            label = item.get("label", "").strip()
            if not label or label.upper() in ("INDONESIA", "TOTAL", "NASIONAL"):
                continue
            if vv == 35: continue  # val=35 = Indonesia
            prov = normalize_province_name(label)
            val  = None
            key_try = f"{vv}958{tv_val}{th_val}0"
            val = datacontent.get(key_try)
            if val is None:
                vv_str = str(vv)
                for k, v in datacontent.items():
                    if k.startswith(vv_str) and "958" in k:
                        val = v; break
            if val is None: continue
            cleaned = self._clean_float(val)
            if cleaned is not None and cleaned > 0:
                result[prov] = cleaned * 1000  # Ribu Jiwa → Jiwa
        return result

    # ── 5. PDRB Lapangan Usaha (list model, var=2268) ────────────────────────
    def parse_pdrb_data(self, raw_data: dict) -> tuple:
        """
        turvar 2005 = A Pertanian, Kehutanan dan Perikanan
        turvar 2022 = Total PDRB
        turtahun 35 = Tahunan (yang dipakai); fallback ke triwulan jika tahunan kosong
        vervar val = 11,12,...,97 (kode 2-digit BPS; 11=Aceh, 12=Sumut, ...)
        Key: "{vv}2268{tv}{th_val}{turtahun}"
        Contoh PDRB Pertanian Aceh Tahunan 2025: "112268200512535"
          = 11 + 2268 + 2005 + 125 + 35
        Satuan: Milyar Rupiah
        """
        pdrb_pertanian, pdrb_total = {}, {}
        if not raw_data or not raw_data.get("datacontent"):
            return pdrb_pertanian, pdrb_total

        vervar_list = raw_data.get("vervar", [])
        datacontent = raw_data.get("datacontent", {})
        th_val      = raw_data.get("tahun", [{}])[0].get("val", _tahun_to_th(self.tahun))
        TV_PERTANIAN = 2005
        TV_TOTAL     = 2022
        TURTAHUN_TAHUNAN = 35

        for item in vervar_list:
            vv    = item.get("val")
            label = item.get("label", "").strip()
            if not label or label.upper() in ("INDONESIA", "TOTAL", "NASIONAL"):
                continue
            prov = normalize_province_name(label)

            # Coba Tahunan dulu
            key_p = f"{vv}2268{TV_PERTANIAN}{th_val}{TURTAHUN_TAHUNAN}"
            key_t = f"{vv}2268{TV_TOTAL}{th_val}{TURTAHUN_TAHUNAN}"

            val_p = datacontent.get(key_p)
            val_t = datacontent.get(key_t)

            # Fallback ke triwulan I jika tahunan tidak ada
            if val_p is None:
                for turt in [31, 32, 33, 34]:
                    v = datacontent.get(f"{vv}2268{TV_PERTANIAN}{th_val}{turt}")
                    if v is not None: val_p = v; break
            if val_t is None:
                for turt in [31, 32, 33, 34]:
                    v = datacontent.get(f"{vv}2268{TV_TOTAL}{th_val}{turt}")
                    if v is not None: val_t = v; break

            if val_p is not None:
                c = self._clean_float(val_p)
                if c and c > 0: pdrb_pertanian[prov] = c
            if val_t is not None:
                c = self._clean_float(val_t)
                if c and c > 0: pdrb_total[prov] = c

        return pdrb_pertanian, pdrb_total


# ── XLSX Helpers ──────────────────────────────────────────────────────────────
_H_BLUE  = "0D47A1"
_H_SUB   = "1565C0"
_BORDER  = Border(*[Side(style="thin", color="FFFFFF")] * 4)
_BORDER2 = Border(*[Side(style="thin", color="CCCCCC")] * 4)


def _xlsx_response(ws_title, title, subtitle, headers, col_widths,
                   data_rows, num_cols, source_text, timestamp, filename):
    wb = Workbook(); ws = wb.active
    ws.title = ws_title; ws.sheet_view.showGridLines = False; ws.freeze_panes = "A4"
    n = len(headers)
    for row, (text, size, color, h) in enumerate([
        (title, 14, _H_BLUE, 30), (subtitle, 10, _H_SUB, 20)], start=1):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n)
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(name="Arial", bold=(row==1), italic=(row==2), color="FFFFFF", size=size)
        c.fill = PatternFill("solid", fgColor=color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = h
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=_H_SUB)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BORDER
    ws.row_dimensions[3].height = 35
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    for ri, row_data in enumerate(data_rows):
        rn = 4 + ri
        fill = PatternFill("solid", fgColor="E3F2FD" if ri % 2 == 0 else "FFFFFF")
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=rn, column=ci, value=val)
            c.fill = fill; c.border = _BORDER2; c.font = Font(name="Arial", size=10)
            if ci in num_cols:
                c.alignment = Alignment(horizontal="right", vertical="center")
                if isinstance(val, float): c.number_format = '#,##0.00'
                elif isinstance(val, int): c.number_format = '#,##0'
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[rn].height = 18
    fr = 4 + len(data_rows) + 1
    ws.merge_cells(start_row=fr, start_column=1, end_row=fr, end_column=n)
    ts = timestamp[:19].replace("T", " ") if timestamp else None
    c = ws.cell(row=fr, column=1,
                value=f"Sumber: {source_text}" + (f"  |  Waktu: {ts}" if ts else ""))
    c.font = Font(name="Arial", italic=True, color="595959", size=9)
    c.alignment = Alignment(horizontal="left", vertical="center")
    out = io.BytesIO(); wb.save(out); out.seek(0)
    resp = HttpResponse(out.read(),
                        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


# ── Download Endpoints ────────────────────────────────────────────────────────
@api_view(['POST'])
def download_ikan_xlsx(request):
    try:
        data = request.data.get("ikan_data")
        if not data: return Response({"error": "Data tidak ditemukan"}, status=400)
        tahun = request.data.get("tahun", 2024)
        rows = [[i, d.get("provinsi", p), d.get("produksi_ikan"), d.get("penduduk"), d.get("rpp_ikan")]
                for i, (p, d) in enumerate(sorted(data.items()), 1)]
        return _xlsx_response("Produksi Ikan",
            "PRODUKSI PERIKANAN LAUT MENURUT PROVINSI",
            f"Sumber: BPS — Kementerian Kelautan dan Perikanan | Tahun {tahun}",
            ["No.", "Provinsi", "Produksi Ikan (ton)", "Jumlah Penduduk (jiwa)", "RPP Ikan (ton/jiwa)"],
            [6, 35, 24, 26, 22], rows, {3, 4, 5},
            f"BPS Web API - Perikanan Tangkap, Tahun {tahun}",
            request.data.get("timestamp", datetime.now().isoformat()),
            f"Dataset_Produksi_Ikan_BPS_{tahun}_{datetime.now().strftime('%Y-%m-%d')}.xlsx")
    except Exception as e: return Response({"error": str(e)}, status=500)


@api_view(['POST'])
def download_perkebunan_xlsx(request):
    try:
        data = request.data.get("perkebunan_data")
        if not data: return Response({"error": "Data tidak ditemukan"}, status=400)
        tahun = request.data.get("tahun", 2024)
        rows = [[i, d.get("provinsi", p), d.get("produksi_perkebunan"), d.get("penduduk"), d.get("rpp_perkebunan")]
                for i, (p, d) in enumerate(sorted(data.items()), 1)]
        return _xlsx_response("Perkebunan",
            "PRODUKSI TANAMAN PERKEBUNAN (RATA-RATA 7-8 KOMODITAS) MENURUT PROVINSI",
            f"Sumber: BPS — Direktorat Jenderal Perkebunan | Tahun {tahun}",
            ["No.", "Provinsi", "Rata-rata Produksi (ton)", "Jumlah Penduduk (jiwa)", "RPP Perkebunan (ton/jiwa)"],
            [6, 35, 26, 26, 26], rows, {3, 4, 5},
            f"BPS Web API - Perkebunan, Tahun {tahun}",
            request.data.get("timestamp", datetime.now().isoformat()),
            f"Dataset_Perkebunan_BPS_{tahun}_{datetime.now().strftime('%Y-%m-%d')}.xlsx")
    except Exception as e: return Response({"error": str(e)}, status=500)


@api_view(['POST'])
def download_nilai_ikan_xlsx(request):
    try:
        data = request.data.get("nilai_ikan_data")
        if not data: return Response({"error": "Data tidak ditemukan"}, status=400)
        tahun = request.data.get("tahun", 2024)
        rows = [[i, d.get("provinsi", p), d.get("nilai_produksi"), d.get("penduduk"), d.get("npi")]
                for i, (p, d) in enumerate(sorted(data.items()), 1)]
        return _xlsx_response("Nilai Ikan",
            "NILAI PRODUKSI PERIKANAN TANGKAP PER KAPITA MENURUT PROVINSI",
            f"Sumber: BPS SIMDASI — Perikanan Tangkap | Tahun {tahun}",
            ["No.", "Provinsi", "Nilai Produksi (Rp)", "Jumlah Penduduk (jiwa)", "NPI (Rp/jiwa)"],
            [6, 35, 26, 26, 22], rows, {3, 4, 5},
            f"BPS Web API - Nilai Perikanan Tangkap, Tahun {tahun}",
            request.data.get("timestamp", datetime.now().isoformat()),
            f"Dataset_Nilai_Ikan_BPS_{tahun}_{datetime.now().strftime('%Y-%m-%d')}.xlsx")
    except Exception as e: return Response({"error": str(e)}, status=500)


@api_view(['POST'])
def download_pdrb_sda_xlsx(request):
    try:
        data = request.data.get("pdrb_data")
        if not data: return Response({"error": "Data tidak ditemukan"}, status=400)
        tahun = request.data.get("tahun", 2024)
        rows = [[i, d.get("provinsi", p), d.get("pdrb_pertanian"), d.get("pdrb_total"), d.get("kps")]
                for i, (p, d) in enumerate(sorted(data.items()), 1)]
        return _xlsx_response("PDRB SDA",
            "PDRB SEKTOR PERTANIAN, KEHUTANAN & PERIKANAN MENURUT PROVINSI",
            f"Sumber: BPS — [Seri 2010] PDRB Lapangan Usaha | Tahun {tahun}",
            ["No.", "Provinsi", "PDRB Pertanian (Milyar Rp)", "Total PDRB (Milyar Rp)", "KPS (rasio)"],
            [6, 35, 30, 28, 14], rows, {3, 4, 5},
            f"BPS Web API - PDRB Lapangan Usaha, Tahun {tahun}",
            request.data.get("timestamp", datetime.now().isoformat()),
            f"Dataset_PDRB_SDA_BPS_{tahun}_{datetime.now().strftime('%Y-%m-%d')}.xlsx")
    except Exception as e: return Response({"error": str(e)}, status=500)


@api_view(['POST'])
def download_ipsda_xlsx(request):
    try:
        data = request.data.get("ipsda_data")
        if not data: return Response({"error": "Data tidak ditemukan"}, status=400)
        tahun = request.data.get("tahun", 2024)
        rows = [[i, d.get("provinsi", p),
                 d.get("rpp_ikan_norm", "-"), d.get("rpp_perkebunan_norm", "-"),
                 d.get("npi_norm", "-"), d.get("kps_norm", "-"),
                 d.get("ipsda", "-"), d.get("status", "-"), d.get("warna", "-")]
                for i, (p, d) in enumerate(sorted(data.items()), 1)]
        return _xlsx_response("IPSDA",
            "INDEKS PEMERATAAN SUMBER DAYA ALAM (IPSDA) MENURUT PROVINSI",
            f"Sumber: BPS Web API | Tahun {tahun} | Seluruh Provinsi Indonesia",
            ["No.", "Provinsi", "RPP_Ikan_norm", "RPP_Kebun_norm",
             "NPI_norm", "KPS_norm", "IPSDA", "Status", "Warna"],
            [6, 35, 16, 16, 14, 14, 12, 14, 12], rows, {3, 4, 5, 6, 7},
            f"BPS Web API - IPSDA, Tahun {tahun}",
            request.data.get("timestamp", datetime.now().isoformat()),
            f"Dataset_IPSDA_BPS_{tahun}_{datetime.now().strftime('%Y-%m-%d')}.xlsx")
    except Exception as e: return Response({"error": str(e)}, status=500)


# ── Endpoint Utama ────────────────────────────────────────────────────────────
@api_view(['POST'])
def analyze_sda_bps(request):
    if not BPS_API_KEY:
        return Response({"error": "BPS Web API Key belum dikonfigurasi"}, status=500)
    try: tahun = int(request.data.get("tahun", 2024))
    except: tahun = 2024

    mode = request.data.get("mode", "bps")
    if mode == "ai":
        result = _run_ai_prediction_sda(tahun, request.data.get("historical_data", {}))
        return Response(result, status=500 if "error" in result else 200)

    if tahun not in TAHUN_SUPPORTED:
        return Response({"error": f"Tahun {tahun} tidak didukung."}, status=400)

    try:
        analytics = SdaAnalytics(tahun=tahun)
        raw_data  = analytics.fetch_all_data()
        cfg       = analytics.config

        ikan_values       = analytics.parse_ikan_data(raw_data.get("IKAN"))
        perkebunan_values = analytics.parse_perkebunan_data(raw_data.get("PERKEBUNAN"), cfg.get("PERKEBUNAN", {}))
        nilai_ikan_values = analytics.parse_nilai_ikan_data(raw_data.get("NILAI_IKAN"), cfg.get("NILAI_IKAN", {}))
        penduduk_values   = analytics.parse_penduduk_data(raw_data.get("PENDUDUK"))
        pdrb_pertanian, pdrb_total = analytics.parse_pdrb_data(raw_data.get("PDRB"))

        print(f"[SDA] ikan:{len(ikan_values)} kebun:{len(perkebunan_values)} "
              f"nilai_ikan:{len(nilai_ikan_values)} penduduk:{len(penduduk_values)} "
              f"pdrb_p:{len(pdrb_pertanian)} pdrb_t:{len(pdrb_total)}")

        # Dedup DKI Jakarta
        for alias, canon in [('DKI JAKARTA', 'JAKARTA'), ('DAERAH KHUSUS IBUKOTA JAKARTA', 'JAKARTA')]:
            for ds in [ikan_values, perkebunan_values, nilai_ikan_values,
                       penduduk_values, pdrb_pertanian, pdrb_total]:
                if alias in ds and canon in ds: del ds[alias]
                elif alias in ds: ds[canon] = ds.pop(alias)

        PROVINSI_CANONICAL_38 = [
            'ACEH','SUMATERA UTARA','SUMATERA BARAT','RIAU','JAMBI','SUMATERA SELATAN',
            'BENGKULU','LAMPUNG','KEPULAUAN BANGKA BELITUNG','KEPULAUAN RIAU','JAKARTA',
            'JAWA BARAT','JAWA TENGAH','DAERAH ISTIMEWA YOGYAKARTA','JAWA TIMUR','BANTEN',
            'BALI','NUSA TENGGARA BARAT','NUSA TENGGARA TIMUR','KALIMANTAN BARAT',
            'KALIMANTAN TENGAH','KALIMANTAN SELATAN','KALIMANTAN TIMUR','KALIMANTAN UTARA',
            'SULAWESI UTARA','SULAWESI TENGAH','SULAWESI SELATAN','SULAWESI TENGGARA',
            'GORONTALO','SULAWESI BARAT','MALUKU','MALUKU UTARA','PAPUA BARAT',
            'PAPUA BARAT DAYA','PAPUA','PAPUA SELATAN','PAPUA TENGAH','PAPUA PEGUNUNGAN',
        ]
        all_prov_set  = set(list(ikan_values)+list(perkebunan_values)+
                            list(nilai_ikan_values)+list(penduduk_values)+list(pdrb_pertanian))
        all_provinces = PROVINSI_CANONICAL_38 + sorted(all_prov_set - set(PROVINSI_CANONICAL_38))

        rpp_ikan_values = {}; rpp_perkebunan_values = {}
        npi_values      = {}; kps_values            = {}

        for prov in all_provinces:
            pddk = penduduk_values.get(prov)
            ikan = ikan_values.get(prov)
            if ikan and pddk and pddk > 0: rpp_ikan_values[prov] = ikan / pddk
            kebun = perkebunan_values.get(prov)
            if kebun and pddk and pddk > 0: rpp_perkebunan_values[prov] = kebun / pddk
            npi_raw = nilai_ikan_values.get(prov)
            if npi_raw and pddk and pddk > 0: npi_values[prov] = npi_raw / pddk
            pdrb_p = pdrb_pertanian.get(prov); pdrb_t = pdrb_total.get(prov)
            if pdrb_p and pdrb_t and pdrb_t > 0: kps_values[prov] = pdrb_p / pdrb_t

        rpp_ikan_norm_map       = minmax_normalize(rpp_ikan_values)
        rpp_perkebunan_norm_map = minmax_normalize(rpp_perkebunan_values)
        npi_norm_map            = minmax_normalize(npi_values)
        kps_norm_map            = minmax_normalize(kps_values)

        ikan_raw={}; perkebunan_raw={}; nilai_ikan_raw={}; pdrb_raw={}; ipsda_raw={}
        for prov in all_provinces:
            pddk = penduduk_values.get(prov)
            if ikan_values.get(prov) or pddk:
                ikan_raw[prov] = {"provinsi": prov, "produksi_ikan": ikan_values.get(prov),
                                  "penduduk": int(pddk) if pddk else None,
                                  "rpp_ikan": rpp_ikan_values.get(prov)}
            if perkebunan_values.get(prov) or pddk:
                perkebunan_raw[prov] = {"provinsi": prov, "produksi_perkebunan": perkebunan_values.get(prov),
                                        "penduduk": int(pddk) if pddk else None,
                                        "rpp_perkebunan": rpp_perkebunan_values.get(prov)}
            if nilai_ikan_values.get(prov) or pddk:
                nilai_ikan_raw[prov] = {"provinsi": prov, "nilai_produksi": nilai_ikan_values.get(prov),
                                        "penduduk": int(pddk) if pddk else None,
                                        "npi": npi_values.get(prov)}
            if pdrb_pertanian.get(prov) or pdrb_total.get(prov):
                pdrb_raw[prov] = {"provinsi": prov, "pdrb_pertanian": pdrb_pertanian.get(prov),
                                  "pdrb_total": pdrb_total.get(prov), "kps": kps_values.get(prov)}

        province_map = {}
        for feat in mongo_db["batas_provinsi"].find({}, {"_id": 0}):
            for field in ["NAMOBJ", "name", "WADMPR", "Provinsi"]:
                v = feat.get("properties", {}).get(field)
                if v:
                    n = normalize_province_name(str(v).upper().strip())
                    province_map[n] = province_map[str(v).upper().strip()] = feat

        matched_features=[]; analysis_summary=[]
        status_counts={"OPTIMAL":0,"CUKUP":0,"KURANG":0,"RENDAH":0}
        provinsi_data_kosong=[]

        for prov in all_provinces:
            has_ikan=prov in ikan_values; has_kebun=prov in perkebunan_values
            has_nilai=prov in nilai_ikan_values; has_pddk=prov in penduduk_values
            has_pdrb=prov in pdrb_pertanian
            has_any=has_ikan or has_kebun or has_nilai or has_pddk or has_pdrb

            kosong_dims=[]
            if not has_ikan:  kosong_dims.append("Produksi Ikan")
            if not has_kebun: kosong_dims.append("Produksi Perkebunan")
            if not has_nilai: kosong_dims.append("Nilai Produksi Ikan")
            if not has_pddk:  kosong_dims.append("Penduduk")
            if not has_pdrb:  kosong_dims.append("PDRB Pertanian")

            rn=rpp_ikan_norm_map.get(prov,0.0); rk=rpp_perkebunan_norm_map.get(prov,0.0)
            nn=npi_norm_map.get(prov,0.0);      kn=kps_norm_map.get(prov,0.0)
            rr=rpp_ikan_values.get(prov); kr=rpp_perkebunan_values.get(prov)
            nr=npi_values.get(prov);      ksr=kps_values.get(prov)

            ipsda=calculate_ipsda(rn,rk,nn,kn); status,warna=classify_ipsda(ipsda)
            if has_any: status_counts[status]+=1
            if kosong_dims: provinsi_data_kosong.append({"provinsi":prov,"dimensi_kosong":kosong_dims})

            analysis_summary.append({
                "provinsi":prov,"status":status if has_any else "-",
                "warna":warna if has_any else "#94a3b8",
                "ipsda":ipsda if has_any else None,
                "rpp_ikan":rr,"rpp_perkebunan":kr,"npi":nr,"kps":ksr,
                "has_complete_data":not kosong_dims,"dimensi_kosong":kosong_dims,
            })
            ipsda_raw[prov]={
                "provinsi":prov,
                "rpp_ikan_norm":rn if has_ikan else None,
                "rpp_perkebunan_norm":rk if has_kebun else None,
                "npi_norm":nn if has_nilai else None,
                "kps_norm":kn if has_pdrb else None,
                "ipsda":ipsda if has_any else None,
                "status":status if has_any else "-","warna":warna if has_any else "#94a3b8",
                "has_complete_data":not kosong_dims,"dimensi_kosong":kosong_dims,
            }
            feat=province_map.get(prov) or next(
                (f for n,f in province_map.items() if prov in n or n in prov),None)
            if not feat or not has_any: continue
            fc=feat.copy()
            fc["properties"]={**fc.get("properties",{}),"sda_analysis":{
                "nama_provinsi":prov,"status":status,"warna":warna,"ipsda":ipsda,
                "rpp_ikan_norm":rn,"rpp_perkebunan_norm":rk,"npi_norm":nn,"kps_norm":kn,
                "has_complete_data":not kosong_dims,"dimensi_kosong":kosong_dims,
                "insights":generate_sda_insights(prov,rr,kr,nr,ksr,ipsda,status),
                "rekomendasi":generate_sda_recommendations(status,rr,kr,nr,ksr),
                "data_sda":{"produksi_ikan":ikan_values.get(prov),
                            "produksi_perkebunan":perkebunan_values.get(prov),
                            "nilai_produksi_ikan":nilai_ikan_values.get(prov),
                            "pdrb_pertanian":pdrb_pertanian.get(prov),
                            "pdrb_total":pdrb_total.get(prov),
                            "jumlah_penduduk":int(penduduk_values[prov]) if penduduk_values.get(prov) else None,
                            "rpp_ikan":rr,"rpp_perkebunan":kr,"npi":nr,"kps":ksr},
            }}
            matched_features.append(fc)

        ada_kosong=len(provinsi_data_kosong)>0
        alert=(f"⚠️ Ada {len(provinsi_data_kosong)} dataset/provinsi data tidak lengkap."
               if ada_kosong else None)

        return Response({
            "status":"success","tahun":tahun,"source":"BPS Web API",
            "is_ai_prediction":False,"dataset_aktif":ALL_DATASETS,
            "total_provinsi":len(analysis_summary),"total_dipetakan":len(matched_features),
            "total_data_kosong":len(provinsi_data_kosong),"total_success":len(matched_features),
            "ada_data_kosong":ada_kosong,"alert_message":alert,
            "provinsi_data_kosong":provinsi_data_kosong,"status_distribusi":status_counts,
            "timestamp":analytics.timestamp_fetch,
            "matched_features":{"type":"FeatureCollection","features":matched_features},
            "analysis_summary":analysis_summary,
            "raw_datasets":{"timestamp":analytics.timestamp_fetch,"tahun":tahun,
                            "IKAN":ikan_raw,"PERKEBUNAN":perkebunan_raw,
                            "NILAI_IKAN":nilai_ikan_raw,"PDRB":pdrb_raw,"IPSDA":ipsda_raw},
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return Response({"error":str(e)},status=500)


# ── AI Prediction ─────────────────────────────────────────────────────────────
def _build_sda_feature_row(tahun_pred, prov_name, cache, meta):
    """Build satu feature row untuk prediksi SDA — identik strukturnya dengan IKP."""
    tahun_rel = tahun_pred - meta["tahun_rel_base"]
    targets   = meta["targets"]
    row = {
        "tahun_rel":  tahun_rel,
        "prov_enc":   cache["prov_enc"],
        "pulau_enc":  cache["pulau_enc"],
    }
    for col in targets:
        v_now  = cache.get(col, 0.0)
        v_prev = cache.get(f"{col}_prev", 0.0)
        row[f"{col}_lag1"]  = v_now
        row[f"{col}_lag2"]  = v_prev
        row[f"{col}_delta"] = v_now - v_prev
        row[f"{col}_roll2"] = (v_now + v_prev) / 2.0
    for col in ["rpp_ikan", "rpp_perkebunan", "npi", "kps"]:
        row[f"{col}_std2"] = abs(cache.get(col, 0.0) - cache.get(f"{col}_prev", 0.0)) * 0.5
    for dc in meta.get("pulau_dummy_cols", []):
        row[dc] = cache.get(dc, 0.0)
    return [row.get(f, 0.0) for f in meta["features"]]


def _run_ai_prediction_sda(tahun: int, historical_data: dict) -> dict:
    if not _load_ai_models():
        return {"error": "Model AI SDA tidak dapat dimuat. Pastikan .pkl tersedia di ai_models/sda/"}
    meta, models, encoders = _AI_META, _AI_MODELS, _AI_ENCODERS

    # Build hist_lookup dari historical_data atau MongoDB
    hist_lookup = {r["provinsi"]: {k: r.get(k) for k in ["rpp_ikan","rpp_perkebunan","npi","kps","ipsda"]}
                   for r in historical_data.get("analysis_summary", []) if r.get("provinsi")}
    if not hist_lookup:
        try:
            docs = list(mongo_db["sda_analysis"].find(
                {"type":"sda","is_ai_prediction":{"$ne":True}},
                {"_id":0,"analysis_summary":1}).sort("timestamp",-1).limit(1))
            if docs:
                hist_lookup = {r["provinsi"]: {k: r.get(k)
                               for k in ["rpp_ikan","rpp_perkebunan","npi","kps","ipsda"]}
                               for r in docs[0].get("analysis_summary",[]) if r.get("provinsi")}
        except Exception as e: print(f"[AI-SDA] MongoDB fallback gagal: {e}")

    timestamp_ai = datetime.now().isoformat()
    province_map = {}
    for feat in mongo_db["batas_provinsi"].find({},{"_id":0}):
        for field in ["NAMOBJ","name","WADMPR","Provinsi"]:
            v = feat.get("properties",{}).get(field)
            if v:
                n = normalize_province_name(str(v).upper().strip())
                province_map[n] = province_map[str(v).upper().strip()] = feat

    PROVINSI_LIST = meta.get("provinsi_list", [])
    targets       = meta.get("targets", ["rpp_ikan","rpp_perkebunan","npi","kps","ipsda"])
    TAHUN_BASE    = meta.get("tahun_rel_base", 2018)
    # Semua tahun dari base sampai tahun target (rolling)
    tahun_range   = list(range(2025, tahun + 1))
    CLIPS         = {col: (0.0, 1.0) for col in targets}  # semua komponen 0-1
    NEUTRAL       = {col: 0.3 for col in targets}

    # ── Kumpulkan prediksi dengan rolling per provinsi ────────────────────────
    raw_preds = []
    for prov_name in PROVINSI_LIST:
        hl = hist_lookup.get(prov_name, {})
        # Encode provinsi & pulau
        try:
            prov_enc = int(encoders["provinsi"].transform([prov_name])[0])
        except Exception:
            continue
        pulau_label = meta.get("pulau_map", {}).get(prov_name, "Sumatera")
        try:
            pulau_enc = int(encoders["pulau"].transform([pulau_label])[0])
        except Exception:
            pulau_enc = 0

        # Bangun cache nilai awal dari historical
        cache = {"prov_enc": prov_enc, "pulau_enc": pulau_enc}
        for col in targets:
            cache[col]           = float(hl.get(col) or NEUTRAL.get(col, 0.3))
            cache[f"{col}_prev"] = float(hl.get(col) or NEUTRAL.get(col, 0.3)) * 0.98
        # Dummy pulau
        for dc in meta.get("pulau_dummy_cols", []):
            cache[dc] = 1.0 if pulau_label == dc.replace("pulau_", "") else 0.0

        # Rolling prediction sampai tahun target
        last_pred = None
        for tahun_pred in tahun_range:
            X_row = _build_sda_feature_row(tahun_pred, prov_name, cache, meta)
            X_df  = pd.DataFrame([X_row], columns=meta["features"])
            pred  = {}
            for col in targets:
                if col not in models:
                    pred[col] = cache[col]
                    continue
                raw = float(models[col].predict(X_df)[0])
                pred[col] = round(float(np.clip(raw, *CLIPS.get(col, (0.0, None)))), 6)
            # Update cache (rolling)
            for col in targets:
                cache[f"{col}_prev"] = cache[col]
                cache[col]           = pred[col]
            last_pred = pred

        if last_pred is None:
            continue
        status, warna = classify_ipsda(last_pred.get("ipsda", 0.0))
        raw_preds.append({
            "provinsi":      prov_name,
            "tahun":         tahun,
            **last_pred,
            "status":        status,
            "warna":         warna,
            "is_prediction": True,
            "model_version": meta.get("version", "rf_v1.0"),
        })

    VAR_THRESHOLD=1e-4; degenerate_cols=set()
    for col in ["rpp_ikan","rpp_perkebunan","npi","kps"]:
        vals=[r[col] for r in raw_preds if r.get(col) is not None]
        if len(vals)>=2:
            mean=sum(vals)/len(vals)
            if sum((v-mean)**2 for v in vals)/len(vals)<VAR_THRESHOLD:
                degenerate_cols.add(col)
    if degenerate_cols:
        for r in raw_preds:
            for col in degenerate_cols: r[col]=NEUTRAL[col]
            r["ipsda"]=calculate_ipsda(r.get("rpp_ikan",0.3),r.get("rpp_perkebunan",0.3),
                                       r.get("npi",0.3),r.get("kps",0.3))
            r["status"],r["warna"]=classify_ipsda(r["ipsda"])

    matched_features=[]; analysis_summary=[]
    status_counts={"OPTIMAL":0,"CUKUP":0,"KURANG":0,"RENDAH":0}
    for r in raw_preds:
        pn=r["provinsi"]; status,warna=r["status"],r["warna"]
        status_counts[status]+=1
        analysis_summary.append({"provinsi":pn,"status":status,"warna":warna,"ipsda":r["ipsda"],
                                  "rpp_ikan":r.get("rpp_ikan"),"rpp_perkebunan":r.get("rpp_perkebunan"),
                                  "npi":r.get("npi"),"kps":r.get("kps"),
                                  "has_complete_data":True,"dimensi_kosong":[],
                                  "is_prediction":True,"degenerate_cols":list(degenerate_cols),
                                  "model_version":meta.get("version","rf_v1.0")})
        feat=province_map.get(pn) or next((f for n,f in province_map.items() if pn in n or n in pn),None)
        if not feat: continue
        fc=feat.copy()
        fc["properties"]={**fc.get("properties",{}),"sda_analysis":{
            "nama_provinsi":pn,"status":status,"warna":warna,"ipsda":r["ipsda"],
            "rpp_ikan_norm":r.get("rpp_ikan"),"rpp_perkebunan_norm":r.get("rpp_perkebunan"),
            "npi_norm":r.get("npi"),"kps_norm":r.get("kps"),
            "has_complete_data":True,"dimensi_kosong":[],"is_prediction":True,
            "degenerate_cols":list(degenerate_cols),"model_version":meta.get("version","rf_v1.0"),
            "insights":generate_sda_insights(pn,r.get("rpp_ikan"),r.get("rpp_perkebunan"),
                                             r.get("npi"),r.get("kps"),r["ipsda"],status),
            "rekomendasi":generate_sda_recommendations(status,r.get("rpp_ikan"),
                          r.get("rpp_perkebunan"),r.get("npi"),r.get("kps")),
            "data_sda":{k:r.get(k) for k in ["rpp_ikan","rpp_perkebunan","npi","kps","ipsda"]},
        }}
        matched_features.append(fc)

    return {"status":"success","tahun":tahun,"is_ai_prediction":True,
            "source":f"AI Prediction - Random Forest {meta.get('version','rf_v1.0')}",
            "model_version":meta.get("version","rf_v1.0"),"model_scores":meta.get("scores",{}),
            "dataset_aktif":ALL_DATASETS,"total_provinsi":len(analysis_summary),
            "total_dipetakan":len(matched_features),"total_data_kosong":0,
            "total_success":len(matched_features),"ada_data_kosong":False,
            "alert_message":None,"provinsi_data_kosong":[],"status_distribusi":status_counts,
            "timestamp":timestamp_ai,
            "matched_features":{"type":"FeatureCollection","features":matched_features},
            "analysis_summary":analysis_summary,"raw_datasets":{}}


@api_view(['POST'])
def analyze_sda_ai(request):
    try: tahun=int(request.data.get("tahun",2025))
    except: tahun=2025
    result=_run_ai_prediction_sda(tahun,request.data.get("historical_data",{}))
    return Response(result,status=500 if "error" in result else 200)


@api_view(['GET'])
def get_sda_ai_model_info(request):
    if not _load_ai_models():
        return Response({"loaded":False,"error":"Model tidak ditemukan.","ai_model_dir":AI_MODEL_DIR})
    return Response({"loaded":True,"version":_AI_META.get("version"),
                     "created_at":_AI_META.get("created_at"),
                     "tahun_historis":_AI_META.get("tahun_historis"),
                     "tahun_prediksi":_AI_META.get("tahun_prediksi"),
                     "provinsi_count":len(_AI_META.get("provinsi_list",[])),
                     "features_count":len(_AI_META.get("features",[])),
                     "targets":_AI_META.get("targets"),"scores":_AI_META.get("scores"),
                     "n_train_rows":_AI_META.get("n_train_rows"),"ai_model_dir":AI_MODEL_DIR})


# ── CRUD ──────────────────────────────────────────────────────────────────────
@api_view(['POST'])
def save_sda_analysis(request):
    try:
        name=request.data.get("name","Analisis SDA Tanpa Nama")
        data=request.data.get("analysis_data")
        if not data: return Response({"error":"Data analisis tidak ditemukan"},status=400)
        aid=str(uuid.uuid4())
        mongo_db["sda_analysis"].insert_one({"analysis_id":aid,"name":name,"type":"sda",
                                              "timestamp":datetime.now().isoformat(),**data})
        return Response({"status":"success","message":f"'{name}' berhasil disimpan","analysis_id":aid})
    except Exception as e: return Response({"error":str(e)},status=500)

@api_view(['GET'])
def get_sda_analysis_list(request):
    try:
        results=list(mongo_db["sda_analysis"].find({},{
            "_id":0,"analysis_id":1,"name":1,"timestamp":1,"total_success":1,
            "status_distribusi":1,"tahun":1,"is_ai_prediction":1,"source":1,
        }).sort("timestamp",-1))
        return Response({"status":"success","count":len(results),"results":results})
    except Exception as e: return Response({"error":str(e)},status=500)

@api_view(['GET'])
def get_sda_analysis_detail(request, analysis_id):
    try:
        result=mongo_db["sda_analysis"].find_one({"analysis_id":analysis_id},{"_id":0})
        if not result: return Response({"error":"Analisis tidak ditemukan"},status=404)
        return Response(result)
    except Exception as e: return Response({"error":str(e)},status=500)

@api_view(['DELETE'])
def delete_sda_analysis(request, analysis_id):
    try:
        r=mongo_db["sda_analysis"].delete_one({"analysis_id":analysis_id})
        if r.deleted_count==0: return Response({"error":"Analisis tidak ditemukan"},status=404)
        return Response({"status":"success","message":"Analisis berhasil dihapus"})
    except Exception as e: return Response({"error":str(e)},status=500)