from rest_framework.decorators import api_view
from rest_framework.response import Response
from pymongo import MongoClient
import requests
import uuid
from datetime import datetime
import os
from dotenv import load_dotenv

# download xlsx
import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

MONGO_URI     = os.getenv("MONGO_URI")
DB_MONGO_NAME = os.getenv("DB_MONGO_NAME")
BPS_API_KEY   = os.getenv("BPS_WEB_API_KEY")

client   = MongoClient(MONGO_URI)
mongo_db = client[DB_MONGO_NAME]

# ─── MAPPING TAHUN → KODE BPS ────────────────────────────────
TAHUN_BPS_MAP = {
    2020: 120,
    2021: 121,
    2022: 122,
    2023: 123,
    2024: 124,
    2025: 125,
    2026: 126,
}

# ─── MAPPING INDIKATOR → DATASET YANG PERLU DIFETCH ──────────
INDIKATOR_DATASET_MAP = {
    'ALL':      ['AHH', 'IMUNISASI', 'SANITASI'],
    'AHH':      ['AHH'],
    'IMUNISASI':['IMUNISASI'],
    'SANITASI': ['SANITASI'],
}

# ─── LABEL & BOBOT ───────────────────────────────────────────
INDIKATOR_LABELS = {
    'ALL':      'Semua Indikator',
    'AHH':      'Angka Harapan Hidup',
    'IMUNISASI':'Cakupan Imunisasi Dasar Lengkap',
    'SANITASI': 'Akses Sanitasi Layak',
}

DATASET_LABELS = {
    'AHH':      'Angka Harapan Hidup (AHH)',
    'IMUNISASI':'Cakupan Imunisasi Dasar Lengkap',
    'SANITASI': 'Akses Sanitasi Layak',
}


def get_indikator_config(tahun: int) -> dict:
    """Return konfigurasi indikator kesehatan dengan th BPS sesuai tahun user."""
    th = TAHUN_BPS_MAP.get(tahun, 124)

    return {
        "AHH": {
            "url_template": f"https://webapi.bps.go.id/v1/api/list/model/data/lang/ind/domain/0000/var/501/th/{th}/key/{{key}}/",
            "nama":     "Angka Harapan Hidup",
            "satuan":   "tahun",
            "threshold_baik":   72,
            "threshold_sedang": 68,
            "bobot":    0.40,
            "penjelasan": "Indikator utama kesehatan populasi yang mencerminkan kualitas layanan kesehatan, nutrisi, dan kondisi sanitasi",
        },
        "IMUNISASI": {
            "url_template": f"https://webapi.bps.go.id/v1/api/list/model/data/lang/ind/domain/0000/var/2280/th/{th}/key/{{key}}/",
            "nama":     "Cakupan Imunisasi Dasar Lengkap",
            "satuan":   "%",
            "threshold_baik":   90,
            "threshold_sedang": 80,
            "bobot":    0.35,
            "penjelasan": "Mencerminkan efektivitas program preventif kesehatan untuk melindungi bayi dan anak dari penyakit menular",
        },
        "SANITASI": {
            "url_template": f"https://webapi.bps.go.id/v1/api/list/model/data/lang/ind/domain/0000/var/847/th/{th}/key/{{key}}/",
            "nama":     "Akses Sanitasi Layak",
            "satuan":   "%",
            "threshold_baik":   85,
            "threshold_sedang": 70,
            "bobot":    0.25,
            "penjelasan": "Indikator infrastruktur dasar kesehatan lingkungan yang berdampak langsung pada pencegahan penyakit",
        },
    }


# ─── HELPER: CEK DATA KOSONG ─────────────────────────────────
def _is_data_empty(data, key):
    if data is None:
        return True
    datacontent = data.get("datacontent", {})
    if not datacontent:
        return True
    valid = [v for v in datacontent.values() if v is not None and v != 0]
    return len(valid) == 0


# ─── ENDPOINT: CEK DATA ──────────────────────────────────────
@api_view(['POST'])
def check_health_year_data(request):
    """
    Cek ketersediaan data BPS kesehatan untuk tahun dan indikator yang dipilih.
    Mirip check_year_data di education_views.
    """
    if not BPS_API_KEY:
        return Response({"error": "BPS Web API Key belum dikonfigurasi"}, status=500)

    tahun     = request.data.get('tahun', 2024)
    indikator = request.data.get('indikator', 'ALL')

    try:
        tahun = int(tahun)
    except (ValueError, TypeError):
        tahun = 2024

    if tahun not in TAHUN_BPS_MAP:
        return Response({"error": f"Tahun {tahun} tidak didukung. Pilih antara 2020–2026."}, status=400)

    keys_to_check = INDIKATOR_DATASET_MAP.get(indikator, INDIKATOR_DATASET_MAP['ALL'])
    all_config    = get_indikator_config(tahun)
    dataset_status = {}

    for key in keys_to_check:
        config = all_config.get(key)
        if not config:
            continue
        url = config["url_template"].format(key=BPS_API_KEY)
        try:
            resp   = requests.get(url, timeout=20)
            if resp.status_code == 200:
                data   = resp.json()
                kosong = _is_data_empty(data, key)
                dataset_status[key] = {
                    "nama":     config["nama"],
                    "tersedia": not kosong,
                    "status":   "Tersedia" if not kosong else "Kosong / Tidak Tersedia",
                }
            else:
                dataset_status[key] = {
                    "nama":     config["nama"],
                    "tersedia": False,
                    "status":   f"HTTP Error {resp.status_code}",
                }
        except Exception as e:
            dataset_status[key] = {
                "nama":     config["nama"],
                "tersedia": False,
                "status":   f"Gagal ({str(e)[:50]})",
            }

    tersedia_list   = [k for k, v in dataset_status.items() if v["tersedia"]]
    kosong_list     = [k for k, v in dataset_status.items() if not v["tersedia"]]
    semua_kosong    = len(tersedia_list) == 0
    ada_yang_kosong = len(kosong_list) > 0 and not semua_kosong

    return Response({
        "tahun":           tahun,
        "indikator":       indikator,
        "dataset_status":  dataset_status,
        "tersedia":        tersedia_list,
        "kosong":          kosong_list,
        "semua_kosong":    semua_kosong,
        "ada_yang_kosong": ada_yang_kosong,
        "bisa_dieksekusi": not semua_kosong and not ada_yang_kosong,
    })


# ─── HELPER STYLE EXCEL ──────────────────────────────────────
def _style_header(ws, row_num, col_count, title, subtitle=None):
    COLOR_HEADER    = "155724"
    COLOR_SUBHEADER = "198754"
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=col_count)
    cell           = ws.cell(row=row_num, column=1, value=title)
    cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=14)
    cell.fill      = PatternFill("solid", fgColor=COLOR_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row_num].height = 30
    if subtitle:
        row_num += 1
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=col_count)
        cell           = ws.cell(row=row_num, column=1, value=subtitle)
        cell.font      = Font(name="Arial", italic=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", fgColor=COLOR_SUBHEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_num].height = 20
        row_num += 1
    else:
        row_num += 1
    return row_num


def _style_col_headers(ws, row_num, headers, col_widths=None):
    COLOR_COL = "198754"
    thin   = Side(style="thin", color="FFFFFF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_idx, header in enumerate(headers, start=1):
        cell           = ws.cell(row=row_num, column=col_idx, value=header)
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", fgColor=COLOR_COL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
    ws.row_dimensions[row_num].height = 35
    if col_widths:
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    return row_num + 1


def _write_data_rows(ws, start_row, data_rows, number_cols=None):
    COLOR_EVEN = "D1ECF1"
    COLOR_ODD  = "FFFFFF"
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row_offset, row_data in enumerate(data_rows):
        row_num    = start_row + row_offset
        fill_color = COLOR_EVEN if row_offset % 2 == 0 else COLOR_ODD
        fill       = PatternFill("solid", fgColor=fill_color)
        for col_idx, value in enumerate(row_data, start=1):
            cell           = ws.cell(row=row_num, column=col_idx, value=value)
            cell.fill      = fill
            cell.border    = border
            cell.font      = Font(name="Arial", size=10)
            if number_cols and col_idx in number_cols:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if isinstance(value, float):
                    cell.number_format = '#,##0.00'
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row_num].height = 18
    return start_row + len(data_rows)


def _add_source_footer(ws, row_num, col_count, source_text, timestamp=None):
    row_num += 1
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=col_count)
    text           = f"Sumber: {source_text}"
    if timestamp:
        text       += f"  |  Waktu Pengambilan Data: {timestamp}"
    cell           = ws.cell(row=row_num, column=1, value=text)
    cell.font      = Font(name="Arial", italic=True, color="595959", size=9)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row_num].height = 16


# ─── DOWNLOAD XLSX: AHH ──────────────────────────────────────
@api_view(['POST'])
def download_ahh_xlsx(request):
    try:
        ahh_data  = request.data.get('ahh_data')
        timestamp = request.data.get('timestamp', datetime.now().isoformat())
        tahun     = request.data.get('tahun', 2024)

        if not ahh_data:
            return Response({"error": "Data AHH tidak ditemukan"}, status=400)

        wb = Workbook()
        ws = wb.active
        ws.title = "AHH"
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A4"

        next_row = _style_header(ws, 1, 5,
            "ANGKA HARAPAN HIDUP (AHH) MENURUT JENIS KELAMIN",
            f"Sumber: BPS Susenas | Tahun {tahun} | Seluruh Provinsi Indonesia")

        headers    = ["No.", "Provinsi", "AHH Laki-laki (tahun)", "AHH Perempuan (tahun)", "AHH Rata-rata (tahun)"]
        col_widths = [6, 35, 24, 24, 24]
        next_row   = _style_col_headers(ws, next_row, headers, col_widths)

        data_rows = []
        for idx, (prov, data) in enumerate(sorted(ahh_data.items()), start=1):
            data_rows.append([
                idx,
                data.get('provinsi', prov),
                data.get('ahh_laki_laki', '-'),
                data.get('ahh_perempuan', '-'),
                data.get('ahh_rata_rata', '-'),
            ])

        _write_data_rows(ws, next_row, data_rows, number_cols={3, 4, 5})
        _add_source_footer(ws, next_row + len(data_rows), 5,
            f"BPS Web API - Susenas, Variabel 501, Tahun {tahun}",
            timestamp[:19].replace('T', ' ') if timestamp else None)

        output = io.BytesIO()
        wb.save(output); output.seek(0)
        tanggal  = datetime.now().strftime("%Y-%m-%d")
        response = HttpResponse(output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Dataset_AHH_BPS_{tahun}_{tanggal}.xlsx"'
        return response

    except Exception as e:
        import traceback; traceback.print_exc()
        return Response({"error": str(e)}, status=500)


# ─── DOWNLOAD XLSX: IMUNISASI ─────────────────────────────────
@api_view(['POST'])
def download_imunisasi_xlsx(request):
    try:
        imunisasi_data = request.data.get('imunisasi_data')
        timestamp      = request.data.get('timestamp', datetime.now().isoformat())
        tahun          = request.data.get('tahun', 2024)

        if not imunisasi_data:
            return Response({"error": "Data Imunisasi tidak ditemukan"}, status=400)

        wb = Workbook()
        ws = wb.active
        ws.title = "Imunisasi"
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A4"

        next_row = _style_header(ws, 1, 3,
            "CAKUPAN IMUNISASI DASAR LENGKAP",
            f"Sumber: BPS Susenas | Tahun {tahun} | Seluruh Provinsi Indonesia")

        headers    = ["No.", "Provinsi", "Cakupan Imunisasi (%)"]
        col_widths = [6, 35, 24]
        next_row   = _style_col_headers(ws, next_row, headers, col_widths)

        data_rows = []
        for idx, (prov, data) in enumerate(sorted(imunisasi_data.items()), start=1):
            val = data.get('nilai', data) if isinstance(data, dict) else data
            data_rows.append([idx, prov, val])

        _write_data_rows(ws, next_row, data_rows, number_cols={3})
        _add_source_footer(ws, next_row + len(data_rows), 3,
            f"BPS Web API - Susenas, Variabel 2280, Tahun {tahun}",
            timestamp[:19].replace('T', ' ') if timestamp else None)

        output = io.BytesIO()
        wb.save(output); output.seek(0)
        tanggal  = datetime.now().strftime("%Y-%m-%d")
        response = HttpResponse(output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Dataset_Imunisasi_BPS_{tahun}_{tanggal}.xlsx"'
        return response

    except Exception as e:
        import traceback; traceback.print_exc()
        return Response({"error": str(e)}, status=500)


# ─── DOWNLOAD XLSX: SANITASI ──────────────────────────────────
@api_view(['POST'])
def download_sanitasi_xlsx(request):
    try:
        sanitasi_data = request.data.get('sanitasi_data')
        timestamp     = request.data.get('timestamp', datetime.now().isoformat())
        tahun         = request.data.get('tahun', 2024)

        if not sanitasi_data:
            return Response({"error": "Data Sanitasi tidak ditemukan"}, status=400)

        wb = Workbook()
        ws = wb.active
        ws.title = "Sanitasi"
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A4"

        next_row = _style_header(ws, 1, 3,
            "AKSES SANITASI LAYAK",
            f"Sumber: BPS | Tahun {tahun} | Seluruh Provinsi Indonesia")

        headers    = ["No.", "Provinsi", "Akses Sanitasi Layak (%)"]
        col_widths = [6, 35, 28]
        next_row   = _style_col_headers(ws, next_row, headers, col_widths)

        data_rows = []
        for idx, (prov, data) in enumerate(sorted(sanitasi_data.items()), start=1):
            val = data.get('nilai', data) if isinstance(data, dict) else data
            data_rows.append([idx, prov, val])

        _write_data_rows(ws, next_row, data_rows, number_cols={3})
        _add_source_footer(ws, next_row + len(data_rows), 3,
            f"BPS Web API - Variabel 847, Tahun {tahun}",
            timestamp[:19].replace('T', ' ') if timestamp else None)

        output = io.BytesIO()
        wb.save(output); output.seek(0)
        tanggal  = datetime.now().strftime("%Y-%m-%d")
        response = HttpResponse(output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Dataset_Sanitasi_BPS_{tahun}_{tanggal}.xlsx"'
        return response

    except Exception as e:
        import traceback; traceback.print_exc()
        return Response({"error": str(e)}, status=500)


# ─── ANALYTICS CLASS ─────────────────────────────────────────
class KesehatanAnalytics:
    def __init__(self, tahun=2024):
        self.tahun           = tahun
        self.indikator_config = get_indikator_config(tahun)
        self.colors = {
            "KRITIS":  "#ef4444",
            "WASPADA": "#f59e0b",
            "STABIL":  "#10b981",
        }
        self.timestamp_fetch = None

    def fetch_all_data(self):
        return self.fetch_selected_data(list(self.indikator_config.keys()))

    def fetch_selected_data(self, keys: list) -> dict:
        all_data = {}
        self.timestamp_fetch = datetime.now().isoformat()
        for key in keys:
            config = self.indikator_config.get(key)
            if not config:
                continue
            try:
                url  = config["url_template"].format(key=BPS_API_KEY)
                print(f"Fetching {key} ({self.tahun}): {url}")
                resp = requests.get(url, timeout=30)
                if resp.status_code == 200:
                    all_data[key] = resp.json()
                    print(f"✓ {key}: Success")
                else:
                    print(f"✗ {key}: HTTP {resp.status_code}")
                    all_data[key] = None
            except Exception as e:
                print(f"✗ {key}: Error - {e}")
                all_data[key] = None
        return all_data

    def parse_bps_data(self, raw_data, key):
        """
        Parse data BPS Susenas. AHH punya breakdown gender (turvar),
        IMUNISASI dan SANITASI langsung satu nilai per provinsi.
        """
        province_values  = {}
        province_details = {}
        raw_breakdown    = {}

        if not raw_data:
            return province_values, province_details, raw_breakdown

        try:
            datacontent = raw_data.get("datacontent", {})
            vervar_list = raw_data.get("vervar", [])
            turvar_list = raw_data.get("turvar", [])

            # Mapping kode provinsi
            province_code_map = {}
            for item in vervar_list:
                code  = str(item.get("val", ""))
                label = item.get("label", "")
                if code and label and code != "9999":
                    province_code_map[code] = label

            # Mapping turvar (gender untuk AHH)
            turvar_map = {}
            for item in turvar_list:
                code  = str(item.get("val", ""))
                label = item.get("label", "")
                turvar_map[code] = label

            temp_data = {}

            for dc_key, value in datacontent.items():
                try:
                    prov_code = dc_key[:4]
                    if prov_code == "9999":
                        continue
                    provinsi_name = province_code_map.get(prov_code)
                    if provinsi_name and value is not None:
                        provinsi_clean = normalize_province_name(str(provinsi_name))
                        value_float    = float(value)

                        if turvar_map and key == "AHH":
                            # AHH: ada breakdown laki-laki / perempuan
                            turvar_code  = dc_key[7:8]
                            gender_label = turvar_map.get(turvar_code, "Unknown")
                            if provinsi_clean not in temp_data:
                                temp_data[provinsi_clean] = {}
                            temp_data[provinsi_clean][gender_label] = value_float
                        else:
                            # IMUNISASI / SANITASI: satu nilai
                            province_values[provinsi_clean] = value_float

                except (ValueError, TypeError, IndexError):
                    continue

            # Proses AHH → rata-rata L+P
            if temp_data and key == "AHH":
                for prov, gender_data in temp_data.items():
                    province_details[prov] = gender_data
                    vals = list(gender_data.values())
                    if vals:
                        avg = round(sum(vals) / len(vals), 2)
                        province_values[prov] = avg

                        laki      = None
                        perempuan = None
                        for g_label, g_val in gender_data.items():
                            if "aki" in g_label.lower():
                                laki = g_val
                            elif "erempuan" in g_label.lower():
                                perempuan = g_val

                        raw_breakdown[prov] = {
                            "provinsi":       prov,
                            "ahh_laki_laki":  laki,
                            "ahh_perempuan":  perempuan,
                            "ahh_rata_rata":  avg,
                        }
            else:
                # IMUNISASI / SANITASI
                for prov, val in province_values.items():
                    raw_breakdown[prov] = {"provinsi": prov, "nilai": val}

            print(f"  ✅ {key}: Parsed {len(province_values)} provinces")

        except Exception as e:
            print(f"  ❌ Parse error {key}: {e}")
            import traceback; traceback.print_exc()

        return province_values, province_details, raw_breakdown

    def calculate_scores(self, data_kesehatan, indikator='ALL'):
        """
        Hitung skor berdasarkan indikator yang aktif.
        ALL  : IKK = AHH×0.40 + IMUNISASI×0.35 + SANITASI×0.25
        AHH  : skor murni AHH (1/2/3)
        IMUNISASI: skor murni Imunisasi
        SANITASI : skor murni Sanitasi
        """
        config = get_indikator_config(self.tahun)

        def _skor(key, val):
            if val is None:
                return 0
            c = config[key]
            if val >= c["threshold_baik"]:
                return 3
            elif val >= c["threshold_sedang"]:
                return 2
            else:
                return 1

        skor_ahh      = _skor("AHH",      data_kesehatan.get("AHH"))
        skor_imunisasi = _skor("IMUNISASI", data_kesehatan.get("IMUNISASI"))
        skor_sanitasi  = _skor("SANITASI",  data_kesehatan.get("SANITASI"))

        if indikator == 'AHH':
            skor_total = float(skor_ahh)
        elif indikator == 'IMUNISASI':
            skor_total = float(skor_imunisasi)
        elif indikator == 'SANITASI':
            skor_total = float(skor_sanitasi)
        else:
            # ALL: bobot standar
            skor_total = round(
                (0.40 * skor_ahh) + (0.35 * skor_imunisasi) + (0.25 * skor_sanitasi), 2
            )

        return {
            'skor_ahh':       skor_ahh,
            'skor_imunisasi': skor_imunisasi,
            'skor_sanitasi':  skor_sanitasi,
            'skor_total':     skor_total,
        }

    def categorize_province(self, skor_total):
        if skor_total >= 2.4:
            return "STABIL",  self.colors["STABIL"]
        elif skor_total >= 1.8:
            return "WASPADA", self.colors["WASPADA"]
        else:
            return "KRITIS",  self.colors["KRITIS"]

    def generate_insights(self, provinsi, data_kesehatan, kategori, skor_total, indikator='ALL'):
        insights = [f"Provinsi {provinsi} berada pada kategori {kategori} dengan skor {skor_total}."]
        config   = get_indikator_config(self.tahun)

        ahh       = data_kesehatan.get("AHH")
        imunisasi = data_kesehatan.get("IMUNISASI")
        sanitasi  = data_kesehatan.get("SANITASI")

        if indikator in ('ALL', 'AHH') and ahh is not None:
            c = config["AHH"]
            if ahh < c["threshold_sedang"]:
                insights.append(f"⚠️ Angka Harapan Hidup ({ahh} tahun) di bawah rata-rata nasional {c['threshold_sedang']} tahun.")
            elif ahh >= c["threshold_baik"]:
                insights.append(f"✅ Angka Harapan Hidup ({ahh} tahun) sudah melampaui target nasional {c['threshold_baik']} tahun.")
            else:
                insights.append(f"📘 Angka Harapan Hidup {ahh} tahun, mendekati target nasional {c['threshold_baik']} tahun.")

        if indikator in ('ALL', 'IMUNISASI') and imunisasi is not None:
            c = config["IMUNISASI"]
            if imunisasi < c["threshold_sedang"]:
                insights.append(f"🚨 Cakupan imunisasi ({imunisasi}%) masih sangat rendah, risiko KLB meningkat.")
            elif imunisasi >= c["threshold_baik"]:
                insights.append(f"✅ Cakupan imunisasi ({imunisasi}%) sangat baik, melindungi populasi.")
            else:
                insights.append(f"⚠️ Cakupan imunisasi {imunisasi}%, perlu ditingkatkan hingga ≥{c['threshold_baik']}%.")

        if indikator in ('ALL', 'SANITASI') and sanitasi is not None:
            c = config["SANITASI"]
            if sanitasi < c["threshold_sedang"]:
                insights.append(f"🚨 Akses sanitasi layak ({sanitasi}%) sangat rendah, risiko penyakit diare tinggi.")
            elif sanitasi >= c["threshold_baik"]:
                insights.append(f"✅ Akses sanitasi ({sanitasi}%) memadai.")
            else:
                insights.append(f"📋 Akses sanitasi {sanitasi}%, perlu perbaikan infrastruktur sanitasi.")

        return insights

    def generate_recommendations(self, kategori, data_kesehatan, indikator='ALL'):
        recommendations = []

        base_actions = {
            'KRITIS': {
                'ALL':       ['Alokasi dana darurat kesehatan', 'Penambahan tenaga kesehatan khusus', 'Program bantuan kesehatan gratis', 'Kampanye kesehatan ibu dan anak'],
                'AHH':       ['Program pencegahan PTM masif', 'Peningkatan akses layanan primer', 'GERMAS di seluruh kecamatan', 'Skrining kesehatan gratis'],
                'IMUNISASI': ['BIAN (Bulan Imunisasi Anak Nasional)', 'Distribusi vaksin ke daerah terpencil', 'Mobilisasi kader posyandu', 'Sistem tracking imunisasi digital'],
                'SANITASI':  ['Program STBM darurat', 'Pembangunan jamban sehat masif', 'Distribusi air bersih', 'Edukasi PHBS door-to-door'],
            },
            'WASPADA': {
                'ALL':       ['Optimalisasi BPJS / JKN-KIS', 'Peningkatan kualitas puskesmas', 'Pelatihan tenaga kesehatan', 'Pembangunan infrastruktur sanitasi'],
                'AHH':       ['Deteksi dini penyakit kronis', 'Kampanye gaya hidup sehat', 'Perkuat layanan rujukan', 'Program gizi seimbang'],
                'IMUNISASI': ['Sweeping imunisasi terjadwal', 'Sosialisasi manfaat imunisasi', 'Penambahan pos imunisasi', 'Kerja sama lintas sektor'],
                'SANITASI':  ['Rehabilitasi IPAL komunal', 'Subsidi sanitasi RT miskin', 'Penguatan Pokja AMPL', 'Monitoring berkala sanitasi'],
            },
            'STABIL': {
                'ALL':       ['Digitalisasi layanan kesehatan', 'Program preventif berbasis komunitas', 'Kemitraan rumah sakit swasta', 'Riset kesehatan lokal'],
                'AHH':       ['Pusat unggulan geriatri', 'Program paliatif komunitas', 'Inovasi layanan primer', 'Monitoring AHH per kelompok umur'],
                'IMUNISASI': ['Riset vaksin lokal', 'Sistem reminder otomatis', 'Sharing best practice antar daerah', 'Imunisasi dewasa & lansia'],
                'SANITASI':  ['Sanitasi berbasis teknologi', 'Program daur ulang air', 'Sertifikasi sanitasi sekolah', 'Open defecation free maintenance'],
            },
        }

        key_rekom = indikator if indikator in base_actions[kategori] else 'ALL'
        recommendations.append({
            'priority': 'Darurat' if kategori == 'KRITIS' else 'Tinggi' if kategori == 'WASPADA' else 'Pemeliharaan',
            'title':    f'{"Intervensi Segera" if kategori=="KRITIS" else "Penguatan" if kategori=="WASPADA" else "Inovasi"} — {INDIKATOR_LABELS.get(indikator, "Semua Indikator")}',
            'actions':  base_actions[kategori][key_rekom],
        })

        return recommendations


# ─── NORMALIZE ───────────────────────────────────────────────
def normalize_province_name(name):
    if not isinstance(name, str):
        name = str(name)
    name = name.upper().strip()
    special = {
        'DKI JAKARTA': 'JAKARTA', 'DAERAH KHUSUS IBUKOTA JAKARTA': 'JAKARTA', 'DKI': 'JAKARTA',
        'YOGYAKARTA': 'DAERAH ISTIMEWA YOGYAKARTA', 'DIY': 'DAERAH ISTIMEWA YOGYAKARTA',
        'D.I. YOGYAKARTA': 'DAERAH ISTIMEWA YOGYAKARTA',
        'BANGKA BELITUNG': 'KEPULAUAN BANGKA BELITUNG', 'KEP. BANGKA BELITUNG': 'KEPULAUAN BANGKA BELITUNG',
        'KEP. RIAU': 'KEPULAUAN RIAU',
    }
    for k, v in special.items():
        if k in name:
            return v
    abbr = {'KEP.': 'KEPULAUAN', 'KEP ': 'KEPULAUAN ', 'NTB': 'NUSA TENGGARA BARAT', 'NTT': 'NUSA TENGGARA TIMUR'}
    for a, f in abbr.items():
        if a in name:
            name = name.replace(a, f)
    for prefix in ['PROVINSI ', 'PROV. ', 'PROV ', 'DAERAH KHUSUS IBUKOTA ']:
        if name.startswith(prefix):
            name = name[len(prefix):]
    return name.strip()


# ─── ENDPOINT: ANALISIS ──────────────────────────────────────
@api_view(['POST'])
def analyze_health_bps(request):
    """
    Analisis kesehatan menggunakan BPS Web API.
    Mendukung pemilihan tahun (2020–2026) dan indikator (ALL/AHH/IMUNISASI/SANITASI).
    """
    if not BPS_API_KEY:
        return Response({"error": "BPS Web API Key belum dikonfigurasi"}, status=500)

    try:
        tahun     = request.data.get('tahun', 2024)
        indikator = request.data.get('indikator', 'ALL')

        try:
            tahun = int(tahun)
        except (ValueError, TypeError):
            tahun = 2024

        if tahun not in TAHUN_BPS_MAP:
            return Response({"error": f"Tahun {tahun} tidak didukung."}, status=400)

        if indikator not in INDIKATOR_DATASET_MAP:
            indikator = 'ALL'

        keys_aktif = INDIKATOR_DATASET_MAP[indikator]
        analytics  = KesehatanAnalytics(tahun=tahun)

        print(f"\n=== MULAI FETCH KESEHATAN BPS | TAHUN={tahun} | INDIKATOR={indikator} ===")
        print(f"    Dataset: {keys_aktif}")

        raw_data = analytics.fetch_selected_data(keys_aktif)

        print("\n=== PARSE DATA ===")
        empty = ({}, {}, {})

        ahh_values,  ahh_details,  ahh_raw   = analytics.parse_bps_data(raw_data.get('AHH'),       'AHH')       if 'AHH'       in keys_aktif else empty
        imun_values, _,            imun_raw   = analytics.parse_bps_data(raw_data.get('IMUNISASI'), 'IMUNISASI') if 'IMUNISASI' in keys_aktif else empty
        sanit_values, _,           sanit_raw  = analytics.parse_bps_data(raw_data.get('SANITASI'),  'SANITASI')  if 'SANITASI'  in keys_aktif else empty

        print("\n=== LOAD BOUNDARY DATA ===")
        cursor            = mongo_db["batas_provinsi"].find({}, {'_id': 0})
        boundary_features = list(cursor)

        province_map = {}
        for feature in boundary_features:
            props = feature.get('properties', {})
            for field in ['NAMOBJ', 'name', 'WADMPR', 'Provinsi']:
                if field in props and props[field]:
                    official   = str(props[field]).upper().strip()
                    normalized = normalize_province_name(official)
                    province_map[normalized] = feature
                    province_map[official]   = feature

        all_provinces = set()
        if 'AHH'       in keys_aktif: all_provinces.update(ahh_values.keys())
        if 'IMUNISASI' in keys_aktif: all_provinces.update(imun_values.keys())
        if 'SANITASI'  in keys_aktif: all_provinces.update(sanit_values.keys())

        print(f"\n=== PROCESSING {len(all_provinces)} PROVINCES | INDIKATOR={indikator} ===")

        matched_features = []
        analysis_summary = []
        kategori_counts  = {"KRITIS": 0, "WASPADA": 0, "STABIL": 0}

        for prov_name in sorted(all_provinces):
            data_kesehatan = {
                "AHH":       ahh_values.get(prov_name)   if 'AHH'       in keys_aktif else None,
                "IMUNISASI": imun_values.get(prov_name)  if 'IMUNISASI' in keys_aktif else None,
                "SANITASI":  sanit_values.get(prov_name) if 'SANITASI'  in keys_aktif else None,
            }
            if not any(v is not None for v in data_kesehatan.values()):
                continue

            scores            = analytics.calculate_scores(data_kesehatan, indikator)
            kategori, warna   = analytics.categorize_province(scores['skor_total'])
            insights          = analytics.generate_insights(prov_name, data_kesehatan, kategori, scores['skor_total'], indikator)
            recommendations   = analytics.generate_recommendations(kategori, data_kesehatan, indikator)

            normalized_prov = normalize_province_name(prov_name)
            matched_feature = (province_map.get(normalized_prov) or province_map.get(prov_name))
            if not matched_feature:
                for map_name, feat in province_map.items():
                    if normalized_prov in map_name or map_name in normalized_prov:
                        matched_feature = feat
                        break
            if not matched_feature:
                continue

            kategori_counts[kategori] += 1
            feature_copy = matched_feature.copy()
            props        = feature_copy.get('properties', {})
            props['health_analysis'] = {
                'nama_provinsi':   prov_name,
                'indikator':       indikator,
                'kategori':        kategori,
                'warna':           warna,
                'skor_total':      scores['skor_total'],
                'skor_ahh':        scores['skor_ahh'],
                'skor_imunisasi':  scores['skor_imunisasi'],
                'skor_sanitasi':   scores['skor_sanitasi'],
                'insights':        insights,
                'rekomendasi':     recommendations,
                'data_kesehatan':  {
                    'AHH':           data_kesehatan.get('AHH'),
                    'AHH_DETAIL':    ahh_details.get(prov_name, {}),
                    'IMUNISASI':     data_kesehatan.get('IMUNISASI'),
                    'SANITASI':      data_kesehatan.get('SANITASI'),
                },
            }
            feature_copy['properties'] = props
            matched_features.append(feature_copy)

            analysis_summary.append({
                'provinsi':   prov_name,
                'indikator':  indikator,
                'kategori':   kategori,
                'warna':      warna,
                'skor_total': scores['skor_total'],
                'ahh':        data_kesehatan.get('AHH'),
                'imunisasi':  data_kesehatan.get('IMUNISASI'),
                'sanitasi':   data_kesehatan.get('SANITASI'),
            })
            print(f"  ✓ {prov_name}: {kategori} (Skor: {scores['skor_total']})")

        print(f"\n=== ANALYSIS COMPLETE | {len(matched_features)} provinces ===")

        raw_datasets = {
            'timestamp':  analytics.timestamp_fetch,
            'tahun':      tahun,
            'indikator':  indikator,
            'AHH':        ahh_raw       if 'AHH'       in keys_aktif else {},
            'IMUNISASI':  imun_raw      if 'IMUNISASI' in keys_aktif else {},
            'SANITASI':   sanit_raw     if 'SANITASI'  in keys_aktif else {},
        }

        return Response({
            'status':              'success',
            'source':              'BPS Web API',
            'tahun':               tahun,
            'indikator':           indikator,
            'dataset_aktif':       keys_aktif,
            'total_success':       len(matched_features),
            'kategori_distribusi': kategori_counts,
            'timestamp':           analytics.timestamp_fetch,
            'matched_features': {
                'type':     'FeatureCollection',
                'features': matched_features,
            },
            'analysis_summary': analysis_summary,
            'raw_datasets':     raw_datasets,
        })

    except Exception as e:
        import traceback; traceback.print_exc()
        return Response({'error': str(e), 'message': 'Gagal menganalisis data kesehatan dari BPS'}, status=500)


# ─── CRUD ─────────────────────────────────────────────────────
@api_view(['POST'])
def save_health_analysis(request):
    try:
        analysis_name = request.data.get('name', 'Analisis Kesehatan Tanpa Nama')
        analysis_data = request.data.get('analysis_data')
        if not analysis_data:
            return Response({"error": "Data analisis tidak ditemukan"}, status=400)
        analysis_id = str(uuid.uuid4())
        document    = {
            "analysis_id": analysis_id,
            "name":        analysis_name,
            "type":        "health",
            "timestamp":   datetime.now().isoformat(),
            **analysis_data,
        }
        mongo_db["health_analysis"].insert_one(document)
        return Response({
            "status":      "success",
            "message":     f"Analisis kesehatan '{analysis_name}' berhasil disimpan",
            "analysis_id": analysis_id,
            "saved_at":    document["timestamp"],
        })
    except Exception as e:
        return Response({"error": str(e), "message": "Gagal menyimpan analisis"}, status=500)


@api_view(['GET'])
def get_health_analysis_list(request):
    try:
        cursor  = mongo_db["health_analysis"].find({}, {
            '_id': 0, 'analysis_id': 1, 'name': 1, 'timestamp': 1,
            'total_success': 1, 'kategori_distribusi': 1, 'tahun': 1, 'indikator': 1,
        }).sort('timestamp', -1)
        results = list(cursor)
        return Response({"status": "success", "count": len(results), "results": results})
    except Exception as e:
        return Response({"error": str(e), "message": "Gagal mengambil daftar analisis"}, status=500)


@api_view(['GET'])
def get_health_analysis_detail(request, analysis_id):
    try:
        result = mongo_db["health_analysis"].find_one({"analysis_id": analysis_id}, {'_id': 0})
        if not result:
            return Response({"error": "Analisis tidak ditemukan"}, status=404)
        return Response(result)
    except Exception as e:
        return Response({"error": str(e), "message": "Gagal mengambil detail analisis"}, status=500)


@api_view(['DELETE'])
def delete_health_analysis(request, analysis_id):
    try:
        result = mongo_db["health_analysis"].delete_one({"analysis_id": analysis_id})
        if result.deleted_count == 0:
            return Response({"error": "Analisis tidak ditemukan"}, status=404)
        return Response({"status": "success", "message": "Analisis berhasil dihapus"})
    except Exception as e:
        return Response({"error": str(e), "message": "Gagal menghapus analisis"}, status=500)