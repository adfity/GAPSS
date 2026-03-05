from rest_framework.decorators import api_view
from rest_framework.response import Response
from pymongo import MongoClient
import uuid
import requests
import os
from dotenv import load_dotenv
from datetime import datetime

# download xlsx
import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

MONGO_URI = os.getenv("MONGO_URI")
DB_MONGO_NAME = os.getenv("DB_MONGO_NAME")
BPS_API_KEY = os.getenv("BPS_WEB_API_KEY")

# Koneksi MongoDB
client = MongoClient(MONGO_URI)
mongo_db = client[DB_MONGO_NAME]

# Mapping kode tahun BPS (th parameter)
TAHUN_BPS_MAP = {
    2020: 120,
    2021: 121,
    2022: 122,
    2023: 123,
    2024: 124,
    2025: 125,
    2026: 126,
}

# Mapping indikator pilihan
INDIKATOR_DATASET_MAP = {
    'ALL':   ['RLS', 'APS', 'SD', 'SMP', 'SMA', 'SMK'],
    'RLS':   ['RLS'],
    'APS':   ['APS'],
    'RASIO': ['SD', 'SMP', 'SMA', 'SMK'],
}

def get_indikator_config(tahun: int) -> dict:
    """Return INDIKATOR_PENDIDIKAN config dengan tahun yang dipilih user"""
    th = TAHUN_BPS_MAP.get(tahun, 124)  # default 2024

    return {
        "RLS": {
            "url_template": f"https://webapi.bps.go.id/v1/api/list/model/data/lang/ind/domain/0000/var/459/th/{th}/key/{{key}}/",
            "nama": "Rata-rata Lama Sekolah",
            "satuan": "tahun",
            "bobot": 0.30
        },
        "APS": {
            "url_template": f"https://webapi.bps.go.id/v1/api/list/model/data/lang/ind/domain/0000/var/2211/th/{th}/key/{{key}}/",
            "nama": "Angka Partisipasi Sekolah",
            "satuan": "%",
            "bobot": 0.50
        },
        "SD": {
            "url_template": f"https://webapi.bps.go.id/v1/api/interoperabilitas/datasource/simdasi/id/25/tahun/{tahun}/id_tabel/UkJNaEl6ZHRVYXNaMzZhZG9BbS9ZZz09/wilayah/0000000/key/{{key}}/",
            "nama": "Data SD",
            "jenis": "sekolah"
        },
        "SMP": {
            "url_template": f"https://webapi.bps.go.id/v1/api/interoperabilitas/datasource/simdasi/id/25/tahun/{tahun}/id_tabel/dzdoVmp3YWdGNU0yWEgraVIwbmRqZz09/wilayah/0000000/key/{{key}}/",
            "nama": "Data SMP",
            "jenis": "sekolah"
        },
        "SMA": {
            "url_template": f"https://webapi.bps.go.id/v1/api/interoperabilitas/datasource/simdasi/id/25/tahun/{tahun}/id_tabel/a1lFcnlHNXNYMFlueG8xL0ZOZnU0Zz09/wilayah/0000000/key/{{key}}/",
            "nama": "Data SMA",
            "jenis": "sekolah"
        },
        "SMK": {
            "url_template": f"https://webapi.bps.go.id/v1/api/interoperabilitas/datasource/simdasi/id/25/tahun/{tahun}/id_tabel/MU90V01YZ0RxenhmbFdsU21iUHh2Zz09/wilayah/0000000/key/{{key}}/",
            "nama": "Data SMK",
            "jenis": "sekolah"
        }
    }


def _is_data_empty(data, indikator_key):
    """
    Cek apakah data dari BPS benar-benar kosong / tidak tersedia.
    Returns True jika data kosong/tidak valid.
    """
    if data is None:
        return True

    jenis = "sekolah" if indikator_key in ("SD", "SMP", "SMA", "SMK") else "bps"

    if jenis == "bps":
        # Untuk RLS & APS: cek datacontent
        datacontent = data.get("datacontent", {})
        if not datacontent:
            return True
        # Cek apakah semua value None / 0
        valid_values = [v for v in datacontent.values() if v is not None and v != 0]
        return len(valid_values) == 0

    else:
        # Untuk SD/SMP/SMA/SMK: cek data rows
        data_container = data.get("data", [])
        if not data_container or len(data_container) < 2:
            return True
        table_data = data_container[1]
        if not isinstance(table_data, dict):
            return True
        data_rows = table_data.get("data", [])
        return len(data_rows) == 0


@api_view(['POST'])
def check_year_data(request):
    """
    Cek ketersediaan data BPS untuk tahun dan indikator yang dipilih.
    Hanya cek dataset yang relevan dengan indikator pilihan.
    Returns status tiap dataset dan apakah bisa dilanjutkan.
    """
    if not BPS_API_KEY:
        return Response({"error": "BPS Web API Key belum dikonfigurasi"}, status=500)

    tahun = request.data.get('tahun', 2024)
    indikator = request.data.get('indikator', 'ALL')  # ALL | RLS | APS | RASIO

    try:
        tahun = int(tahun)
    except (ValueError, TypeError):
        tahun = 2024

    if tahun not in TAHUN_BPS_MAP:
        return Response({"error": f"Tahun {tahun} tidak didukung. Pilih antara 2020–2026."}, status=400)

    # Tentukan dataset mana yang perlu dicek sesuai indikator
    keys_to_check = INDIKATOR_DATASET_MAP.get(indikator, INDIKATOR_DATASET_MAP['ALL'])
    all_config    = get_indikator_config(tahun)
    dataset_status = {}

    for key in keys_to_check:
        config = all_config.get(key)
        if not config:
            continue
        url = config["url_template"].format(key=BPS_API_KEY)
        try:
            resp = requests.get(url, timeout=20)
            if resp.status_code == 200:
                data   = resp.json()
                kosong = _is_data_empty(data, key)
                dataset_status[key] = {
                    "nama":     config["nama"],
                    "tersedia": not kosong,
                    "status":   "Tersedia" if not kosong else "Kosong / Tidak Tersedia"
                }
            else:
                dataset_status[key] = {
                    "nama":     config["nama"],
                    "tersedia": False,
                    "status":   f"HTTP Error {resp.status_code}"
                }
        except Exception as e:
            dataset_status[key] = {
                "nama":     config["nama"],
                "tersedia": False,
                "status":   f"Gagal ({str(e)[:50]})"
            }

    tersedia_list   = [k for k, v in dataset_status.items() if v["tersedia"]]
    kosong_list     = [k for k, v in dataset_status.items() if not v["tersedia"]]
    semua_kosong    = len(tersedia_list) == 0
    ada_yang_kosong = len(kosong_list) > 0 and not semua_kosong

    return Response({
        "tahun":            tahun,
        "indikator":        indikator,
        "dataset_status":   dataset_status,
        "tersedia":         tersedia_list,
        "kosong":           kosong_list,
        "semua_kosong":     semua_kosong,
        "ada_yang_kosong":  ada_yang_kosong,
        "bisa_dieksekusi":  not semua_kosong and not ada_yang_kosong
    })


# ✅ HELPER: STYLE OPENPYXL
def _style_header(ws, row_num, col_count, title, subtitle=None):
    """Tambah header biru dengan judul dan subtitle opsional"""
    COLOR_HEADER = "1F4E79"
    COLOR_SUBHEADER = "2E75B6"

    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=col_count)
    cell = ws.cell(row=row_num, column=1, value=title)
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=14)
    cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row_num].height = 30

    if subtitle:
        row_num += 1
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=col_count)
        cell = ws.cell(row=row_num, column=1, value=subtitle)
        cell.font = Font(name="Arial", italic=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=COLOR_SUBHEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_num].height = 20
        row_num += 1
    else:
        row_num += 1

    return row_num


def _style_col_headers(ws, row_num, headers, col_widths=None):
    COLOR_COL_HEADER = "2E75B6"
    thin = Side(style="thin", color="FFFFFF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=header)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=COLOR_COL_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ws.row_dimensions[row_num].height = 35

    if col_widths:
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    return row_num + 1


def _write_data_rows(ws, start_row, data_rows, number_cols=None, highlight_last=True):
    COLOR_EVEN = "DEEAF1"
    COLOR_ODD = "FFFFFF"
    COLOR_TOTAL = "FFF2CC"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_offset, row_data in enumerate(data_rows):
        row_num = start_row + row_offset
        is_last = (row_offset == len(data_rows) - 1) and highlight_last

        if is_last:
            fill_color = COLOR_TOTAL
            font_bold = True
        elif row_offset % 2 == 0:
            fill_color = COLOR_EVEN
            font_bold = False
        else:
            fill_color = COLOR_ODD
            font_bold = False

        fill = PatternFill("solid", fgColor=fill_color)

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.fill = fill
            cell.border = border
            cell.font = Font(name="Arial", bold=font_bold, size=10)

            if number_cols and col_idx in number_cols:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if isinstance(value, float):
                    cell.number_format = '#,##0.00'
                elif isinstance(value, int):
                    cell.number_format = '#,##0'
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        ws.row_dimensions[row_num].height = 18

    return start_row + len(data_rows)


def _add_source_footer(ws, row_num, col_count, source_text, timestamp=None):
    row_num += 1
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=col_count)
    text = f"Sumber: {source_text}"
    if timestamp:
        text += f"  |  Waktu Pengambilan Data: {timestamp}"
    cell = ws.cell(row=row_num, column=1, value=text)
    cell.font = Font(name="Arial", italic=True, color="595959", size=9)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row_num].height = 16


# ENDPOINT: DOWNLOAD XLSX RLS
@api_view(['POST'])
def download_rls_xlsx(request):
    try:
        rls_data  = request.data.get('rls_data')
        timestamp = request.data.get('timestamp', datetime.now().isoformat())
        tahun     = request.data.get('tahun', 2024)

        if not rls_data:
            return Response({"error": "Data RLS tidak ditemukan"}, status=400)

        wb = Workbook()
        ws = wb.active
        ws.title = "RLS"
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A4"

        next_row = _style_header(
            ws, 1, 4,
            "RATA-RATA LAMA SEKOLAH (RLS) MENURUT JENIS KELAMIN",
            f"Sumber: BPS Susenas | Tahun {tahun} | Seluruh Provinsi Indonesia"
        )

        headers = ["No.", "Provinsi", "Laki-laki (Tahun)", "Perempuan (Tahun)", "Rata-rata (Tahun)"]
        col_widths = [6, 35, 20, 20, 20]
        next_row = _style_col_headers(ws, next_row, headers, col_widths)

        sorted_provinces = sorted(rls_data.items(), key=lambda x: x[0])
        data_rows = []
        for idx, (prov, data) in enumerate(sorted_provinces, start=1):
            data_rows.append([
                idx,
                data.get('provinsi', prov),
                data.get('rls_laki_laki', '-'),
                data.get('rls_perempuan', '-'),
                data.get('rls_rata_rata', '-'),
            ])

        _write_data_rows(ws, next_row, data_rows, number_cols={3, 4, 5}, highlight_last=False)
        _add_source_footer(ws, next_row + len(data_rows), 5,
                           f"BPS Web API - Susenas, Variabel 459, Tahun {tahun}",
                           timestamp[:19].replace('T', ' ') if timestamp else None)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        tanggal = datetime.now().strftime("%Y-%m-%d")
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="Dataset_RLS_BPS_{tahun}_{tanggal}.xlsx"'
        return response

    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({"error": str(e)}, status=500)


# ENDPOINT: DOWNLOAD XLSX APS
@api_view(['POST'])
def download_aps_xlsx(request):
    try:
        aps_data  = request.data.get('aps_data')
        timestamp = request.data.get('timestamp', datetime.now().isoformat())
        tahun     = request.data.get('tahun', 2024)

        if not aps_data:
            return Response({"error": "Data APS tidak ditemukan"}, status=400)

        wb = Workbook()
        ws = wb.active
        ws.title = "APS"
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A4"

        next_row = _style_header(
            ws, 1, 5,
            "ANGKA PARTISIPASI SEKOLAH (APS) MENURUT KELOMPOK UMUR",
            f"Sumber: BPS Susenas | Tahun {tahun} | Seluruh Provinsi Indonesia"
        )

        headers = ["No.", "Provinsi", "APS 7-12 Tahun (%)", "APS 13-15 Tahun (%)", "APS 16-18 Tahun (%)", "APS 19-23 Tahun (%)"]
        col_widths = [6, 35, 20, 20, 20, 20]
        next_row = _style_col_headers(ws, next_row, headers, col_widths)

        sorted_provinces = sorted(aps_data.items(), key=lambda x: x[0])
        data_rows = []
        for idx, (prov, data) in enumerate(sorted_provinces, start=1):
            data_rows.append([
                idx,
                data.get('provinsi', prov),
                data.get('aps_7_12_tahun', '-'),
                data.get('aps_13_15_tahun', '-'),
                data.get('aps_16_18_tahun', '-'),
                data.get('aps_19_23_tahun', '-'),
            ])

        _write_data_rows(ws, next_row, data_rows, number_cols={3, 4, 5, 6}, highlight_last=False)
        _add_source_footer(ws, next_row + len(data_rows), 6,
                           f"BPS Web API - Susenas, Variabel 2211, Tahun {tahun}",
                           timestamp[:19].replace('T', ' ') if timestamp else None)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        tanggal = datetime.now().strftime("%Y-%m-%d")
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="Dataset_APS_BPS_{tahun}_{tanggal}.xlsx"'
        return response

    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({"error": str(e)}, status=500)


# ENDPOINT: DOWNLOAD XLSX RASIO
@api_view(['POST'])
def download_rasio_xlsx(request):
    try:
        rasio_sd  = request.data.get('rasio_sd', {})
        rasio_smp = request.data.get('rasio_smp', {})
        rasio_sma = request.data.get('rasio_sma', {})
        rasio_smk = request.data.get('rasio_smk', {})
        timestamp = request.data.get('timestamp', datetime.now().isoformat())
        tahun     = request.data.get('tahun', 2024)

        JENJANG_CONFIG = [
            ("SD",  rasio_sd,  "Sekolah Dasar (SD)"),
            ("SMP", rasio_smp, "Sekolah Menengah Pertama (SMP)"),
            ("SMA", rasio_sma, "Sekolah Menengah Atas (SMA)"),
            ("SMK", rasio_smk, "Sekolah Menengah Kejuruan (SMK)"),
        ]

        wb = Workbook()
        wb.remove(wb.active)

        for sheet_name, data_dict, label_panjang in JENJANG_CONFIG:
            ws = wb.create_sheet(title=sheet_name)
            ws.sheet_view.showGridLines = False
            ws.freeze_panes = "A4"

            next_row = _style_header(
                ws, 1, 4,
                f"RASIO MURID-GURU - {label_panjang.upper()}",
                f"Sumber: Kemdikbudristek DAPODIK | Semester Ganjil {tahun}/{tahun+1} | Seluruh Provinsi Indonesia"
            )

            headers = ["No.", "Provinsi", f"Jumlah Guru {sheet_name}", f"Jumlah Murid {sheet_name}", "Rasio Murid per Guru"]
            col_widths = [6, 35, 22, 22, 22]
            next_row = _style_col_headers(ws, next_row, headers, col_widths)

            sorted_provinces = sorted(data_dict.items(), key=lambda x: x[0])
            data_rows = []
            for idx, (prov, data) in enumerate(sorted_provinces, start=1):
                data_rows.append([
                    idx,
                    data.get('provinsi', prov),
                    data.get('jumlah_guru', '-'),
                    data.get('jumlah_murid', '-'),
                    data.get('rasio_murid_per_guru', '-'),
                ])

            _write_data_rows(ws, next_row, data_rows, number_cols={3, 4, 5}, highlight_last=False)
            _add_source_footer(ws, next_row + len(data_rows), 5,
                               f"BPS Web API - SIMDASI Kemdikbudristek, {sheet_name}, Tahun {tahun}",
                               timestamp[:19].replace('T', ' ') if timestamp else None)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        tanggal = datetime.now().strftime("%Y-%m-%d")
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="Dataset_Rasio_Murid_Guru_BPS_{tahun}_{tanggal}.xlsx"'
        return response

    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({"error": str(e)}, status=500)


# CLASS & FUNGSI ANALISIS
class PendidikanAnalytics:
    """Model analisis pendidikan dengan data BPS"""

    def __init__(self, tahun=2024):
        self.tahun = tahun
        self.indikator_config = get_indikator_config(tahun)
        self.colors = {
            "BAIK": "#10b981",
            "SEDANG": "#f59e0b",
            "KRITIS": "#ef4444"
        }
        self.timestamp_fetch = None

    def fetch_all_data(self):
        return self.fetch_selected_data(list(self.indikator_config.keys()))

    def fetch_selected_data(self, keys: list) -> dict:
        """Fetch hanya dataset dengan key yang ada di list keys"""
        all_data = {}
        self.timestamp_fetch = datetime.now().isoformat()

        for indikator_key in keys:
            config = self.indikator_config.get(indikator_key)
            if not config:
                continue
            try:
                url = config["url_template"].format(key=BPS_API_KEY)
                print(f"Fetching {indikator_key} ({self.tahun}): {url}")
                response = requests.get(url, timeout=30)
                if response.status_code == 200:
                    all_data[indikator_key] = response.json()
                    print(f"✓ {indikator_key}: Success")
                else:
                    print(f"✗ {indikator_key}: HTTP {response.status_code}")
                    all_data[indikator_key] = None
            except Exception as e:
                print(f"✗ {indikator_key}: Error - {e}")
                all_data[indikator_key] = None

        return all_data

    def parse_rls_data(self, raw_data):
        rls_values = {}
        rls_details = {}
        rls_raw_breakdown = {}

        if not raw_data:
            print("  ❌ RLS: No raw data")
            return rls_values, rls_details, rls_raw_breakdown

        try:
            datacontent = raw_data.get("datacontent", {})

            if not datacontent:
                print("  ❌ RLS: datacontent is empty")
                return rls_values, rls_details, rls_raw_breakdown

            print(f"  📊 RLS: datacontent has {len(datacontent)} keys")

            province_names = {
                '1100': 'ACEH', '1200': 'SUMATERA UTARA', '1300': 'SUMATERA BARAT',
                '1400': 'RIAU', '1500': 'JAMBI', '1600': 'SUMATERA SELATAN',
                '1700': 'BENGKULU', '1800': 'LAMPUNG', '1900': 'KEPULAUAN BANGKA BELITUNG',
                '2100': 'KEPULAUAN RIAU', '3100': 'JAKARTA', '3200': 'JAWA BARAT',
                '3300': 'JAWA TENGAH', '3400': 'DAERAH ISTIMEWA YOGYAKARTA', '3500': 'JAWA TIMUR',
                '3600': 'BANTEN', '5100': 'BALI', '5200': 'NUSA TENGGARA BARAT',
                '5300': 'NUSA TENGGARA TIMUR', '6100': 'KALIMANTAN BARAT',
                '6200': 'KALIMANTAN TENGAH', '6300': 'KALIMANTAN SELATAN',
                '6400': 'KALIMANTAN TIMUR', '6500': 'KALIMANTAN UTARA',
                '7100': 'SULAWESI UTARA', '7200': 'SULAWESI TENGAH',
                '7300': 'SULAWESI SELATAN', '7400': 'SULAWESI TENGGARA',
                '7500': 'GORONTALO', '7600': 'SULAWESI BARAT',
                '8100': 'MALUKU', '8200': 'MALUKU UTARA',
                '9100': 'PAPUA BARAT', '9200': 'PAPUA BARAT DAYA',
                '9400': 'PAPUA', '9500': 'PAPUA SELATAN',
                '9600': 'PAPUA TENGAH', '9700': 'PAPUA PEGUNUNGAN'
            }

            th = TAHUN_BPS_MAP.get(self.tahun, 124)

            for prov_code, prov_name in province_names.items():
                try:
                    key_male   = f"{prov_code}4592111{th}0"
                    key_female = f"{prov_code}4592121{th}0"

                    rls_male   = None
                    rls_female = None

                    if key_male in datacontent and key_female in datacontent:
                        rls_male   = datacontent.get(key_male)
                        rls_female = datacontent.get(key_female)
                    else:
                        matching_keys = [k for k in datacontent.keys() if k.startswith(prov_code + '459')]
                        if len(matching_keys) >= 2:
                            rls_male   = datacontent.get(matching_keys[0])
                            rls_female = datacontent.get(matching_keys[1]) if len(matching_keys) > 1 else rls_male

                    if rls_male is not None and rls_female is not None:
                        try:
                            rls_male_val   = float(rls_male)
                            rls_female_val = float(rls_female)
                            rls_avg = round((rls_male_val + rls_female_val) / 2, 2)

                            prov_normalized = normalize_province_name(prov_name)
                            rls_values[prov_normalized] = rls_avg

                            rls_details[prov_normalized] = {
                                "laki_laki": round(rls_male_val, 2),
                                "perempuan": round(rls_female_val, 2),
                                "rata_rata": rls_avg
                            }

                            rls_raw_breakdown[prov_normalized] = {
                                "provinsi": prov_normalized,
                                "rls_laki_laki": round(rls_male_val, 2),
                                "rls_perempuan": round(rls_female_val, 2),
                                "rls_rata_rata": rls_avg
                            }

                        except (ValueError, TypeError) as e:
                            print(f"    ✗ {prov_name}: Error converting values - {e}")
                            continue

                except Exception as e:
                    print(f"    ✗ {prov_name}: Error - {e}")
                    continue

            print(f"  ✅ RLS: Parsed {len(rls_values)} provinces")

        except Exception as e:
            print(f"  ❌ Parse error for RLS: {e}")
            import traceback
            traceback.print_exc()

        return rls_values, rls_details, rls_raw_breakdown

    def parse_aps_data(self, raw_data):
        aps_values = {}
        aps_raw_breakdown = {}

        if not raw_data:
            return aps_values, aps_raw_breakdown

        try:
            datacontent  = raw_data.get("datacontent", {})
            vervar_list  = raw_data.get("vervar", [])
            turvar_list  = raw_data.get("turvar", [])

            province_code_map = {}
            for item in vervar_list:
                code  = str(item.get("val", ""))
                label = item.get("label", "")
                if code and label and code != "9999" and code.endswith("00"):
                    province_code_map[code] = label

            turvar_map = {}
            for item in turvar_list:
                code  = str(item.get("val", ""))
                label = item.get("label", "")
                turvar_map[code] = label

            temp_data = {}

            for key, value in datacontent.items():
                try:
                    if len(key) != 16:
                        continue

                    prov_code   = key[0:4]
                    var_code    = key[4:8]
                    turvar_code = key[8:12]

                    if prov_code == "9999" or var_code != "2211":
                        continue

                    provinsi_name = province_code_map.get(prov_code)

                    if provinsi_name and value is not None:
                        provinsi_clean = normalize_province_name(str(provinsi_name))
                        value_float    = float(value)
                        umur_label     = turvar_map.get(turvar_code)

                        if umur_label:
                            if provinsi_clean not in temp_data:
                                temp_data[provinsi_clean] = {}
                            temp_data[provinsi_clean][umur_label] = value_float

                except (ValueError, TypeError, IndexError):
                    continue

            for prov, umur_data in temp_data.items():
                formatted_data    = {}
                raw_breakdown_data = {"provinsi": prov}

                for key, val in umur_data.items():
                    if key == '7-12':
                        formatted_data['7_12']             = val
                        raw_breakdown_data['aps_7_12_tahun'] = val
                    elif key == '13-15':
                        formatted_data['13_15']              = val
                        raw_breakdown_data['aps_13_15_tahun'] = val
                    elif key == '16-18':
                        formatted_data['16_18']              = val
                        raw_breakdown_data['aps_16_18_tahun'] = val
                    elif key == '19-23':
                        formatted_data['19_23']              = val
                        raw_breakdown_data['aps_19_23_tahun'] = val
                    else:
                        formatted_data[key.replace('-', '_')]        = val
                        raw_breakdown_data[f'aps_{key}'] = val

                aps_values[prov]        = formatted_data
                aps_raw_breakdown[prov] = raw_breakdown_data

            print(f"  ✅ Parsed {len(aps_values)} provinces for APS")

        except Exception as e:
            print(f"  ❌ Parse error for APS: {e}")

        return aps_values, aps_raw_breakdown

    def parse_school_data(self, raw_data, jenjang):
        rasio_values = {}
        raw_breakdown = {}

        if not raw_data:
            return rasio_values, raw_breakdown

        try:
            data_container = raw_data.get('data', [])

            if not data_container or len(data_container) < 2:
                return rasio_values, raw_breakdown

            table_data = data_container[1]

            if not isinstance(table_data, dict):
                return rasio_values, raw_breakdown

            data_rows = table_data.get('data', [])

            if not data_rows:
                return rasio_values, raw_breakdown

            for row_data in data_rows:
                try:
                    if not isinstance(row_data, dict):
                        continue

                    prov_name_raw = row_data.get('label', '').strip()

                    if not prov_name_raw or prov_name_raw.upper() == 'INDONESIA':
                        continue

                    prov_lower = prov_name_raw.lower()
                    if any(x in prov_lower for x in ['kab', 'kota', 'kabupaten', 'kecamatan']):
                        continue

                    variables = row_data.get('variables', {})
                    if not variables:
                        continue

                    guru_key  = None
                    murid_key = None
                    kolom_info = table_data.get('kolom', {})

                    for key, value in variables.items():
                        var_name = kolom_info.get(key, {}).get('nama_variabel', '')
                        if 'Guru' in var_name and '(Negeri+Swasta)' in var_name:
                            guru_key = key
                        elif 'Murid' in var_name and '(Negeri+Swasta)' in var_name:
                            murid_key = key

                    if not guru_key or not murid_key:
                        continue

                    guru_data  = variables.get(guru_key, {})
                    murid_data = variables.get(murid_key, {})

                    guru_raw  = guru_data.get('value', '').strip()
                    murid_raw = murid_data.get('value', '').strip()

                    guru_clean  = guru_raw.replace('.', '').replace(',', '').replace(' ', '').strip()
                    murid_clean = murid_raw.replace('.', '').replace(',', '').replace(' ', '').strip()

                    if not guru_clean or not murid_clean or guru_clean in ['-', '...'] or murid_clean in ['-', '...']:
                        continue

                    try:
                        guru  = float(guru_clean)
                        murid = float(murid_clean)
                    except ValueError:
                        continue

                    if guru > 0 and murid > 0:
                        prov_name = normalize_province_name(prov_name_raw)
                        rasio     = round(murid / guru, 2)
                        rasio_values[prov_name] = rasio

                        raw_breakdown[prov_name] = {
                            "provinsi": prov_name,
                            "jumlah_guru": int(guru),
                            "jumlah_murid": int(murid),
                            "rasio_murid_per_guru": rasio
                        }

                except (ValueError, TypeError, KeyError):
                    continue

            print(f"  ✅ {jenjang}: Successfully parsed {len(rasio_values)} provinces")

        except Exception as e:
            print(f"  ❌ {jenjang}: Parse error - {str(e)}")

        return rasio_values, raw_breakdown


def normalize_province_name(name):
    if not isinstance(name, str):
        name = str(name)

    name = name.upper().strip()

    special_mappings = {
        'DKI JAKARTA': 'JAKARTA',
        'DAERAH KHUSUS IBUKOTA JAKARTA': 'JAKARTA',
        'DKI': 'JAKARTA',
        'YOGYAKARTA': 'DAERAH ISTIMEWA YOGYAKARTA',
        'DIY': 'DAERAH ISTIMEWA YOGYAKARTA',
        'D.I. YOGYAKARTA': 'DAERAH ISTIMEWA YOGYAKARTA',
        'BANGKA BELITUNG': 'KEPULAUAN BANGKA BELITUNG',
        'KEP. BANGKA BELITUNG': 'KEPULAUAN BANGKA BELITUNG',
        'KEPULAUAN RIAU': 'KEPULAUAN RIAU',
        'KEP. RIAU': 'KEPULAUAN RIAU',
    }

    for key, value in special_mappings.items():
        if key in name:
            return value

    abbreviations = {
        'KEP.': 'KEPULAUAN',
        'KEP ': 'KEPULAUAN ',
        'NTB': 'NUSA TENGGARA BARAT',
        'NTT': 'NUSA TENGGARA TIMUR',
    }

    for abbr, full in abbreviations.items():
        if abbr in name:
            name = name.replace(abbr, full)

    prefixes = ['PROVINSI ', 'PROV. ', 'PROV ', 'DAERAH KHUSUS IBUKOTA ']
    for prefix in prefixes:
        if name.startswith(prefix):
            name = name[len(prefix):]

    return name.strip()


def calculate_scores(rls_value, aps_data, rasio_data, indikator='ALL'):
    """
    Hitung skor berdasarkan indikator yang aktif.
    - ALL:   bobot RLS 0.30 + APS 0.50 + Rasio 0.20
    - RLS:   skor murni dari RLS (skala 1-3 → dinormalisasi ke 1-3)
    - APS:   skor murni dari APS gabungan
    - RASIO: skor murni dari Rasio murid-guru
    """
    # --- Skor RLS ---
    skor_rls = 0
    if rls_value is not None:
        if rls_value > 9.5:   skor_rls = 3
        elif rls_value >= 8.0: skor_rls = 2
        else:                  skor_rls = 1

    # --- Skor APS ---
    skor_aps_list = []
    aps_details   = {}
    if isinstance(aps_data, dict):
        for key in ['7_12', '13_15', '16_18', '19_23']:
            value            = aps_data.get(key)
            aps_details[key] = value
            if value is not None:
                if value > 80:    skor_aps_list.append(3)
                elif value >= 70: skor_aps_list.append(2)
                else:             skor_aps_list.append(1)
    skor_aps = round(sum(skor_aps_list) / len(skor_aps_list), 2) if skor_aps_list else 0

    # --- Skor Rasio ---
    rasio_values  = []
    rasio_details = {}
    for key in ['SD', 'SMP', 'SMA', 'SMK']:
        value            = rasio_data.get(key)
        rasio_details[key] = value
        if value is not None and value > 0:
            rasio_values.append(value)
    rasio_rata = round(sum(rasio_values) / len(rasio_values), 2) if rasio_values else 0

    skor_rasio = 0
    if rasio_rata > 0:
        if rasio_rata < 12:   skor_rasio = 3
        elif rasio_rata <= 16: skor_rasio = 2
        else:                  skor_rasio = 1

    # --- Skor Total: bobot berbeda sesuai indikator aktif ---
    if indikator == 'RLS':
        skor_total = float(skor_rls)          # 1, 2, atau 3
    elif indikator == 'APS':
        skor_total = skor_aps                  # rata-rata 1-3
    elif indikator == 'RASIO':
        skor_total = float(skor_rasio)         # 1, 2, atau 3
    else:
        # ALL: bobot standar
        skor_total = round((0.30 * skor_rls) + (0.50 * skor_aps) + (0.20 * skor_rasio), 2)

    return {
        'skor_rls':    skor_rls,
        'skor_aps':    skor_aps,
        'skor_rasio':  skor_rasio,
        'skor_total':  skor_total,
        'rasio_rata':  rasio_rata,
        'aps_details': aps_details,
        'rasio_details': rasio_details,
        'aps_count':   len(skor_aps_list)
    }


def categorize_province(skor_total):
    """
    Kategorikan berdasarkan skor total.
    ALL → threshold 2.4 / 1.8 (range 1-3, bobot campur)
    RLS/APS/RASIO → skor 1/2/3 langsung
    """
    colors = {"BAIK": "#10b981", "SEDANG": "#f59e0b", "KRITIS": "#ef4444"}

    if skor_total >= 2.4:
        return "BAIK",   colors["BAIK"]
    elif skor_total >= 1.8:
        return "SEDANG", colors["SEDANG"]
    else:
        return "KRITIS", colors["KRITIS"]


def generate_insights(provinsi, rls_value, aps_data, rasio_rata, kategori, skor_total, aps_details=None, rasio_details=None, indikator='ALL'):
    """Generate insights berdasarkan indikator yang aktif"""
    insights = []
    insights.append(f"Provinsi {provinsi} berada pada kategori {kategori} dengan skor {skor_total}.")

    if indikator in ('ALL', 'RLS') and rls_value is not None:
        if rls_value < 8.0:
            insights.append(f"⚠️ Rata-rata lama sekolah ({rls_value} tahun) masih di bawah target nasional 9 tahun.")
        elif rls_value >= 9.5:
            insights.append(f"✅ Rata-rata lama sekolah ({rls_value} tahun) sudah melampaui target nasional.")
        else:
            insights.append(f"📘 Rata-rata lama sekolah {rls_value} tahun, mendekati target 9 tahun.")

    if indikator in ('ALL', 'APS') and aps_details:
        aps_vals = [v for v in aps_details.values() if v is not None]
        if aps_vals:
            min_aps = min(aps_vals)
            avg_aps = round(sum(aps_vals) / len(aps_vals), 1)
            if min_aps < 70:
                insights.append(f"📉 Ada kelompok umur dengan APS di bawah 70% (terendah {min_aps}%), perlu perhatian.")
            elif avg_aps >= 85:
                insights.append(f"✅ Rata-rata APS {avg_aps}% - partisipasi sekolah sangat tinggi di semua kelompok umur.")
            else:
                insights.append(f"📊 Rata-rata APS {avg_aps}% - masih ada ruang peningkatan partisipasi.")

    if indikator in ('ALL', 'RASIO') and rasio_rata > 0:
        if rasio_rata > 16:
            insights.append(f"👥 Rasio murid-guru {rasio_rata} - kekurangan guru, beban mengajar tinggi.")
        elif rasio_rata < 12:
            insights.append(f"✅ Rasio murid-guru {rasio_rata} - dalam kondisi ideal.")
        else:
            insights.append(f"📋 Rasio murid-guru {rasio_rata} - beban mengajar dalam batas wajar.")

    return insights


def generate_recommendations(kategori, rls_value, aps_data, rasio_rata, indikator='ALL'):
    """Generate rekomendasi berbasis indikator yang aktif"""
    recommendations = []

    # Rekomendasi umum berdasarkan kategori
    if kategori == "KRITIS":
        base_actions = {
            'ALL':   ['Alokasi dana darurat pendidikan', 'Program beasiswa masif', 'Rekrutmen guru prioritas', 'Kampanye wajib belajar 12 tahun'],
            'RLS':   ['Program percepatan kejar paket A/B/C', 'Insentif siswa tetap bersekolah', 'Penguatan pendidikan non-formal', 'Kerja sama daerah-pusat untuk target RLS'],
            'APS':   ['Subsidi transport dan seragam sekolah', 'Pendirian sekolah di daerah terpencil', 'Program beasiswa penuh untuk keluarga miskin', 'Digitalisasi pembelajaran jarak jauh'],
            'RASIO': ['Rekrutmen mendesak guru ASN/PPPK', 'Redistribusi guru dari daerah surplus', 'Program guru tamu dan relawan', 'Percepatan sertifikasi guru honorer'],
        }
    elif kategori == "SEDANG":
        base_actions = {
            'ALL':   ['Pelatihan guru berkelanjutan', 'Perbaikan sarana prasarana', 'Pengurangan angka putus sekolah', 'Teknologi dalam pembelajaran'],
            'RLS':   ['Beasiswa SMP-SMA untuk daerah tertinggal', 'Penguatan program PKBM', 'Monitoring kehadiran siswa', 'Insentif guru di pedalaman'],
            'APS':   ['Peningkatan kualitas PAUD & SD', 'Kampanye pentingnya pendidikan tinggi', 'Penguatan BOS afirmasi', 'Beasiswa perguruan tinggi vokasi'],
            'RASIO': ['Pemetaan kebutuhan guru per sekolah', 'Penambahan formasi PPPK guru', 'Optimasi jadwal mengajar', 'Pelatihan guru multi-mata-pelajaran'],
        }
    else:
        base_actions = {
            'ALL':   ['Pengembangan pusat keunggulan', 'Program pertukaran guru-siswa', 'Inovasi kurikulum abad 21', 'Kolaborasi internasional'],
            'RLS':   ['Program S1/S2 akselerasi', 'Pusat studi unggulan daerah', 'Kemitraan dengan universitas', 'Beasiswa riset lokal'],
            'APS':   ['Perluasan akses PT berkualitas', 'Penguatan pendidikan vokasi', 'Inkubator talenta muda', 'Program dual degree internasional'],
            'RASIO': ['Rasio ideal dipertahankan', 'Inovasi model team teaching', 'Pengembangan platform e-learning', 'Pusat pelatihan guru nasional'],
        }

    key_rekom = indikator if indikator in base_actions.get('ALL', {}) else 'ALL'
    actions   = base_actions.get(key_rekom, base_actions['ALL'])

    recommendations.append({
        'priority': 'Tinggi' if kategori == 'KRITIS' else 'Sedang' if kategori == 'SEDANG' else 'Rendah',
        'title':    f'{"Percepatan" if kategori=="KRITIS" else "Optimalisasi" if kategori=="SEDANG" else "Inovasi"} Pendidikan - Fokus {"Semua Indikator" if indikator=="ALL" else ("RLS" if indikator=="RLS" else "APS" if indikator=="APS" else "Rasio Murid-Guru")}',
        'actions':  actions
    })

    # Rekomendasi tambahan khusus Rasio jika di luar batas (hanya jika indikator ALL atau RASIO)
    if indikator in ('ALL', 'RASIO') and rasio_rata > 16:
        recommendations.append({
            'priority': 'Tinggi',
            'title':    'Penanganan Darurat Kekurangan Guru',
            'actions':  ['Rekrutmen guru honorer → PNS/PPPK', 'Penempatan guru daerah terpencil', 'Guru multi-kompetensi', 'Kelas virtual untuk daerah terpencil']
        })

    return recommendations


@api_view(['POST'])
def analyze_education_bps(request):
    """
    Analisis pendidikan menggunakan BPS Web API.
    Mendukung pemilihan tahun (2020-2026) dan indikator (ALL/RLS/APS/RASIO).
    Hanya fetch & proses dataset yang relevan dengan indikator yang dipilih.
    """
    if not BPS_API_KEY:
        return Response({
            "error": "BPS Web API Key belum dikonfigurasi",
            "message": "Silakan tambahkan BPS_WEB_API_KEY di file .env"
        }, status=500)

    try:
        tahun     = request.data.get('tahun', 2024)
        indikator = request.data.get('indikator', 'ALL')   # ALL | RLS | APS | RASIO

        try:
            tahun = int(tahun)
        except (ValueError, TypeError):
            tahun = 2024

        if tahun not in TAHUN_BPS_MAP:
            return Response({"error": f"Tahun {tahun} tidak didukung."}, status=400)

        if indikator not in INDIKATOR_DATASET_MAP:
            indikator = 'ALL'

        # Keys dataset yang perlu di-fetch untuk indikator ini
        keys_aktif = INDIKATOR_DATASET_MAP[indikator]

        analytics = PendidikanAnalytics(tahun=tahun)

        print(f"\n=== MULAI FETCH DATA BPS | TAHUN={tahun} | INDIKATOR={indikator} ===")
        print(f"    Dataset yang akan di-fetch: {keys_aktif}")

        # Fetch hanya dataset yang dibutuhkan
        raw_data = analytics.fetch_selected_data(keys_aktif)

        print("\n=== PARSE DATA PER PROVINSI ===")

        # Parse - gunakan dict kosong jika tidak di-fetch
        empty_rls   = ({}, {}, {})
        empty_aps   = ({}, {})
        empty_rasio = ({}, {})

        rls_values, rls_details, rls_raw_breakdown = (
            analytics.parse_rls_data(raw_data.get('RLS')) if 'RLS' in keys_aktif else empty_rls
        )
        aps_values, aps_raw_breakdown = (
            analytics.parse_aps_data(raw_data.get('APS')) if 'APS' in keys_aktif else empty_aps
        )
        rasio_sd,  rasio_sd_raw  = (analytics.parse_school_data(raw_data.get('SD'),  'SD')  if 'SD'  in keys_aktif else empty_rasio)
        rasio_smp, rasio_smp_raw = (analytics.parse_school_data(raw_data.get('SMP'), 'SMP') if 'SMP' in keys_aktif else empty_rasio)
        rasio_sma, rasio_sma_raw = (analytics.parse_school_data(raw_data.get('SMA'), 'SMA') if 'SMA' in keys_aktif else empty_rasio)
        rasio_smk, rasio_smk_raw = (analytics.parse_school_data(raw_data.get('SMK'), 'SMK') if 'SMK' in keys_aktif else empty_rasio)

        print("\n=== LOAD BOUNDARY DATA ===")
        cursor            = mongo_db["batas_provinsi"].find({}, {'_id': 0})
        boundary_features = list(cursor)

        province_map = {}
        for feature in boundary_features:
            props = feature.get('properties', {})
            for field in ['NAMOBJ', 'name', 'WADMPR', 'Provinsi']:
                if field in props and props[field]:
                    official_name = str(props[field]).upper().strip()
                    normalized    = normalize_province_name(official_name)
                    province_map[normalized]    = feature
                    province_map[official_name] = feature

        print(f"Loaded {len(province_map)} province boundaries")

        # Kumpulkan semua provinsi dari dataset yang aktif
        all_provinces = set()
        if 'RLS'  in keys_aktif: all_provinces.update(rls_values.keys())
        if 'APS'  in keys_aktif: all_provinces.update(aps_values.keys())
        if 'SD'   in keys_aktif: all_provinces.update(rasio_sd.keys())
        if 'SMP'  in keys_aktif: all_provinces.update(rasio_smp.keys())
        if 'SMA'  in keys_aktif: all_provinces.update(rasio_sma.keys())
        if 'SMK'  in keys_aktif: all_provinces.update(rasio_smk.keys())

        print(f"\n=== PROCESSING {len(all_provinces)} PROVINCES | INDIKATOR={indikator} ===")

        matched_features = []
        analysis_summary = []
        kategori_counts  = {"BAIK": 0, "SEDANG": 0, "KRITIS": 0}

        for prov_name in sorted(all_provinces):
            rls_value  = rls_values.get(prov_name)  if 'RLS' in keys_aktif else None
            aps_prov   = aps_values.get(prov_name, {}) if 'APS' in keys_aktif else {}
            rasio_prov = {
                'SD':  rasio_sd.get(prov_name)  if 'SD'  in keys_aktif else None,
                'SMP': rasio_smp.get(prov_name) if 'SMP' in keys_aktif else None,
                'SMA': rasio_sma.get(prov_name) if 'SMA' in keys_aktif else None,
                'SMK': rasio_smk.get(prov_name) if 'SMK' in keys_aktif else None,
            }

            # Hitung skor hanya untuk indikator yang aktif
            scores    = calculate_scores(rls_value, aps_prov, rasio_prov, indikator)
            kategori, warna = categorize_province(scores['skor_total'])

            normalized_prov = normalize_province_name(prov_name)
            matched_feature = None

            if normalized_prov in province_map:
                matched_feature = province_map[normalized_prov]
            elif prov_name in province_map:
                matched_feature = province_map[prov_name]
            else:
                for map_name, feature in province_map.items():
                    if normalized_prov in map_name or map_name in normalized_prov:
                        matched_feature = feature
                        break

            if not matched_feature:
                continue

            insights        = generate_insights(
                prov_name, rls_value, aps_prov,
                scores['rasio_rata'], kategori, scores['skor_total'],
                scores.get('aps_details'), scores.get('rasio_details'),
                indikator
            )
            recommendations = generate_recommendations(
                kategori, rls_value, aps_prov, scores['rasio_rata'], indikator
            )

            kategori_counts[kategori] += 1

            feature_copy = matched_feature.copy()
            props        = feature_copy.get('properties', {})

            props['education_analysis'] = {
                'nama_provinsi': prov_name,
                'indikator':     indikator,
                'kategori':      kategori,
                'warna':         warna,
                'skor_total':    scores['skor_total'],
                'skor_rls':      scores['skor_rls'],
                'skor_aps':      scores['skor_aps'],
                'skor_rasio':    scores['skor_rasio'],
                'insights':      insights,
                'rekomendasi':   recommendations,
                'data_pendidikan': {
                    'RLS':        rls_value,
                    'RLS_DETAIL': rls_details.get(prov_name, {}),
                    'APS_7_12':   aps_prov.get('7_12'),
                    'APS_13_15':  aps_prov.get('13_15'),
                    'APS_16_18':  aps_prov.get('16_18'),
                    'APS_19_23':  aps_prov.get('19_23'),
                    'SKOR_APS':   scores['skor_aps'],
                    'RASIO_SD':   rasio_prov['SD'],
                    'RASIO_SMP':  rasio_prov['SMP'],
                    'RASIO_SMA':  rasio_prov['SMA'],
                    'RASIO_SMK':  rasio_prov['SMK'],
                    'RASIO_RATA': scores['rasio_rata']
                }
            }

            feature_copy['properties'] = props
            matched_features.append(feature_copy)

            analysis_summary.append({
                'provinsi':   prov_name,
                'indikator':  indikator,
                'kategori':   kategori,
                'warna':      warna,
                'skor_total': scores['skor_total'],
                'rls':        rls_value,
                'skor_aps':   scores['skor_aps'],
                'rasio_rata': scores['rasio_rata']
            })

            print(f"  ✓ {prov_name}: {kategori} (Skor: {scores['skor_total']})")

        print(f"\n=== ANALYSIS COMPLETE | {len(matched_features)} provinces ===")

        raw_datasets = {
            'timestamp':  analytics.timestamp_fetch,
            'tahun':      tahun,
            'indikator':  indikator,
            'RLS':        rls_raw_breakdown      if 'RLS' in keys_aktif else {},
            'APS':        aps_raw_breakdown      if 'APS' in keys_aktif else {},
            'RASIO_SD':   rasio_sd_raw           if 'SD'  in keys_aktif else {},
            'RASIO_SMP':  rasio_smp_raw          if 'SMP' in keys_aktif else {},
            'RASIO_SMA':  rasio_sma_raw          if 'SMA' in keys_aktif else {},
            'RASIO_SMK':  rasio_smk_raw          if 'SMK' in keys_aktif else {},
        }

        return Response({
            'status':              'success',
            'source':              'BPS Web API',
            'tahun':               tahun,
            'indikator':           indikator,
            'dataset_aktif':       keys_aktif,
            'total_success':       len(matched_features),
            'kategori_distribusi': kategori_counts,
            'timestamp':           analytics.timestamp_fetch,
            'matched_features': {
                'type':     'FeatureCollection',
                'features': matched_features
            },
            'analysis_summary': analysis_summary,
            'raw_datasets':     raw_datasets
        })

    except Exception as e:
        print(f"ERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response({
            'error':   str(e),
            'message': 'Gagal menganalisis data pendidikan dari BPS'
        }, status=500)


@api_view(['POST'])
def save_education_analysis(request):
    try:
        data          = request.data
        analysis_name = data.get('name', 'Analisis Pendidikan Tanpa Nama')
        analysis_data = data.get('analysis_data')

        if not analysis_data:
            return Response({"error": "Data analisis tidak ditemukan"}, status=400)

        analysis_id = str(uuid.uuid4())

        document = {
            "analysis_id": analysis_id,
            "name": analysis_name,
            "type": "education",
            "timestamp": datetime.now().isoformat(),
            **analysis_data
        }

        mongo_db["education_analysis"].insert_one(document)

        return Response({
            "status": "success",
            "message": f"Analisis pendidikan '{analysis_name}' berhasil disimpan",
            "analysis_id": analysis_id,
            "saved_at": document["timestamp"]
        })

    except Exception as e:
        return Response({
            "error": str(e),
            "message": "Gagal menyimpan analisis"
        }, status=500)


@api_view(['GET'])
def get_education_analysis_list(request):
    try:
        cursor = mongo_db["education_analysis"].find(
            {},
            {
                '_id': 0,
                'analysis_id': 1,
                'name': 1,
                'timestamp': 1,
                'total_success': 1,
                'kategori_distribusi': 1,
                'tahun': 1,
                'indikator': 1,
            }
        ).sort('timestamp', -1)

        results = list(cursor)

        return Response({
            "status": "success",
            "count": len(results),
            "results": results
        })

    except Exception as e:
        return Response({
            "error": str(e),
            "message": "Gagal mengambil daftar analisis"
        }, status=500)


@api_view(['GET'])
def get_education_analysis_detail(request, analysis_id):
    try:
        result = mongo_db["education_analysis"].find_one(
            {"analysis_id": analysis_id},
            {'_id': 0}
        )

        if not result:
            return Response({"error": "Analisis tidak ditemukan"}, status=404)

        return Response(result)

    except Exception as e:
        return Response({
            "error": str(e),
            "message": "Gagal mengambil detail analisis"
        }, status=500)


@api_view(['DELETE'])
def delete_education_analysis(request, analysis_id):
    try:
        result = mongo_db["education_analysis"].delete_one(
            {"analysis_id": analysis_id}
        )

        if result.deleted_count == 0:
            return Response({"error": "Analisis tidak ditemukan"}, status=404)

        return Response({
            "status": "success",
            "message": "Analisis berhasil dihapus"
        })

    except Exception as e:
        return Response({
            "error": str(e),
            "message": "Gagal menghapus analisis"
        }, status=500)