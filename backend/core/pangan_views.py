from rest_framework.decorators import api_view
from rest_framework.response import Response
from pymongo import MongoClient
import uuid, requests, os, json
import numpy as np
import pandas as pd
from dotenv import load_dotenv
from datetime import datetime
import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

MONGO_URI     = os.getenv("MONGO_URI")
DB_MONGO_NAME = os.getenv("DB_MONGO_NAME")
BPS_API_KEY   = os.getenv("BPS_WEB_API_KEY")

client   = MongoClient(MONGO_URI)
mongo_db = client[DB_MONGO_NAME]

# ── AI Model (lazy load) ──────────────────────────────────────────────────────
_AI_MODELS = _AI_ENCODERS = _AI_META = None
AI_MODEL_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ai_models", "pangan")

def _load_ai_models():
    global _AI_MODELS, _AI_ENCODERS, _AI_META
    if _AI_MODELS is not None:
        return True
    try:
        import joblib
        meta_path = os.path.join(AI_MODEL_DIR, "metadata.json")
        if not os.path.exists(meta_path):
            return False
        with open(meta_path, "r", encoding="utf-8") as f:
            _AI_META = json.load(f)
        _AI_MODELS = {t: joblib.load(os.path.join(AI_MODEL_DIR, f"model_{t}.pkl")) for t in _AI_META["targets"]}
        _AI_ENCODERS = {
            "provinsi": joblib.load(os.path.join(AI_MODEL_DIR, "encoder_provinsi.pkl")),
            "pulau":    joblib.load(os.path.join(AI_MODEL_DIR, "encoder_pulau.pkl")),
        }
        return True
    except Exception as e:
        print(f"[AI] Gagal load model: {e}")
        _AI_MODELS = _AI_ENCODERS = _AI_META = None
        return False


# ── Konfigurasi BPS ───────────────────────────────────────────────────────────
TAHUN_SUPPORTED  = list(range(2020, 2031))
SIMDASI_PADI_TABEL     = {y: "ZjZ6MXlacGJNR0JaaHBPRSs0TzNUdz09" for y in range(2018, 2027)}
SIMDASI_PENDUDUK_TABEL = {y: "WVRlTTcySlZDa3lUcFp6czNwbHl4QT09" for y in range(2018, 2027)}
URL_KEMISKINAN_LIST = (
    "https://webapi.bps.go.id/v1/api/list/model/data/lang/ind"
    "/domain/0000/var/192/th/{th_val}/key/{key}/"
)
URL_KONSUMSI_STATIC = (
    "https://webapi.bps.go.id/v1/api/view/domain/0000"
    "/model/statictable/lang/ind/id/951/key/{key}/"
)
INDIKATOR_DATASET_MAP = {
    'ALL':          ['PADI', 'KONSUMSI', 'KEMISKINAN', 'PENDUDUK'],
    'KETERSEDIAAN': ['PADI', 'PENDUDUK'],
    'KONSUMSI':     ['KONSUMSI'],
    'AKSES':        ['KEMISKINAN'],
}
DATASET_LABELS = {
    'PADI':       'Produksi, Luas Panen & Produktivitas Padi',
    'KONSUMSI':   'Konsumsi Kalori & Protein per Kapita',
    'KEMISKINAN': 'Persentase Penduduk Miskin',
    'PENDUDUK':   'Jumlah Penduduk',
}
COL_PENDUDUK_JUMLAH = "nzudy5elv7"


def get_pangan_config(tahun: int) -> dict:
    def _nearest(tbl, yr):
        v = tbl.get(yr)
        if not v:
            for d in [-1, -2, 1, 2]:
                v = tbl.get(yr + d)
                if v: break
        return v

    config = {}
    id_padi = _nearest(SIMDASI_PADI_TABEL, tahun)
    if id_padi:
        config["PADI"] = {
            "url_template": (
                f"https://webapi.bps.go.id/v1/api/interoperabilitas/datasource/simdasi"
                f"/id/25/tahun/{tahun}/id_tabel/{id_padi}/wilayah/0000000/key/{{key}}/"
            ),
            "nama": "Produksi, Luas Panen & Produktivitas Padi", "jenis": "simdasi_padi",
        }
    config["KONSUMSI"] = {
        "url_template": URL_KONSUMSI_STATIC,
        "nama": "Konsumsi Kalori & Protein per Kapita", "jenis": "statictable_konsumsi",
    }
    th_val = tahun - 1900
    config["KEMISKINAN"] = {
        "url_template": URL_KEMISKINAN_LIST.replace("{th_val}", str(th_val)),
        "nama": "Persentase Penduduk Miskin", "jenis": "list_kemiskinan", "th_val": th_val,
    }
    id_pddk = _nearest(SIMDASI_PENDUDUK_TABEL, tahun)
    if id_pddk:
        config["PENDUDUK"] = {
            "url_template": (
                f"https://webapi.bps.go.id/v1/api/interoperabilitas/datasource/simdasi"
                f"/id/25/tahun/{tahun}/id_tabel/{id_pddk}/wilayah/0000000/key/{{key}}/"
            ),
            "nama": "Jumlah Penduduk", "jenis": "simdasi_penduduk",
        }
    return config


def _is_pangan_data_empty(data, dataset_key, tahun=None):
    if data is None:
        return True
    if dataset_key == "KONSUMSI":
        table_html = data.get("data", {}).get("table", "")
        if not table_html or len(table_html) < 100:
            return True
        # Tabel konsumsi BPS adalah statis - data tersedia selama tabel tidak kosong.
        # Tidak strict-check tahun karena BPS update tabel ini dengan delay 1-2 tahun.
        return False
    elif dataset_key == "KEMISKINAN":
        return not data.get("datacontent", {})
    else:
        data_container = data.get("data", [])
        if not data_container or len(data_container) < 2:
            return True
        table_data = data_container[1]
        return not isinstance(table_data, dict) or len(table_data.get("data", [])) == 0


# ── Endpoint: Cek Ketersediaan Data ──────────────────────────────────────────
@api_view(['POST'])
def check_year_data_pangan(request):
    if not BPS_API_KEY:
        return Response({"error": "BPS Web API Key belum dikonfigurasi"}, status=500)
    try:
        tahun = int(request.data.get('tahun', 2025))
    except (ValueError, TypeError):
        tahun = 2025
    if tahun not in TAHUN_SUPPORTED:
        return Response({"error": f"Tahun {tahun} tidak didukung."}, status=400)

    all_config = get_pangan_config(tahun)
    dataset_status = {}
    for key in INDIKATOR_DATASET_MAP['ALL']:
        config = all_config.get(key)
        if not config:
            dataset_status[key] = {"nama": DATASET_LABELS.get(key, key), "tersedia": False, "status": "Konfigurasi tidak ada"}
            continue
        try:
            resp = requests.get(config["url_template"].format(key=BPS_API_KEY), timeout=20)
            if resp.status_code == 200:
                kosong = _is_pangan_data_empty(resp.json(), key, tahun=tahun)
                dataset_status[key] = {
                    "nama": config["nama"], "tersedia": not kosong,
                    "status": "Tersedia" if not kosong else "Kosong / Tidak Tersedia",
                }
            else:
                dataset_status[key] = {"nama": config["nama"], "tersedia": False, "status": f"HTTP {resp.status_code}"}
        except Exception as e:
            dataset_status[key] = {"nama": config.get("nama", key), "tersedia": False, "status": f"Gagal ({str(e)[:50]})"}

    tersedia_list = [k for k, v in dataset_status.items() if v["tersedia"]]
    kosong_list   = [k for k, v in dataset_status.items() if not v["tersedia"]]
    semua_kosong  = len(tersedia_list) == 0
    ada_yang_kosong = len(kosong_list) > 0 and not semua_kosong

    ai_model_ready = ai_version = None
    ai_tersedia = False
    if semua_kosong or ada_yang_kosong:
        ai_model_ready = _load_ai_models()
        ai_tersedia    = ai_model_ready
        ai_version     = _AI_META.get("version") if _AI_META else None

    return Response({
        "tahun": tahun, "is_ai_prediction": False, "indikator": 'ALL',
        "dataset_status": dataset_status,
        "tersedia": tersedia_list, "kosong": kosong_list,
        "semua_kosong": semua_kosong, "ada_yang_kosong": ada_yang_kosong,
        "bisa_dieksekusi": len(tersedia_list) >= 2,  # bisa jalan kalau >= 2 dataset tersedia
        "bps_kosong": semua_kosong,
        "ai_tersedia": ai_tersedia, "ai_model_ready": ai_model_ready, "ai_model_version": ai_version,
    })


# ── XLSX Helpers ──────────────────────────────────────────────────────────────
_H_GREEN = "1B5E20"
_H_SUB   = "388E3C"
_BORDER  = Border(*[Side(style="thin", color="FFFFFF")] * 4)
_BORDER2 = Border(*[Side(style="thin", color="CCCCCC")] * 4)

def _xlsx_response(ws_title, title, subtitle, headers, col_widths, data_rows, num_cols, source_text, timestamp, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = ws_title
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    n = len(headers)

    # Title rows
    for row, (text, size, color, h) in enumerate([
        (title,    14, _H_GREEN, 30),
        (subtitle, 10, _H_SUB,   20),
    ], start=1):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n)
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(name="Arial", bold=(row == 1), italic=(row == 2), color="FFFFFF", size=size)
        c.fill = PatternFill("solid", fgColor=color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = h

    # Column headers
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=_H_SUB)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BORDER
    ws.row_dimensions[3].height = 35
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Data rows
    for ri, row_data in enumerate(data_rows):
        rn = 4 + ri
        fill = PatternFill("solid", fgColor="E8F5E9" if ri % 2 == 0 else "FFFFFF")
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=rn, column=ci, value=val)
            c.fill = fill; c.border = _BORDER2
            c.font = Font(name="Arial", size=10)
            if ci in num_cols:
                c.alignment = Alignment(horizontal="right", vertical="center")
                if isinstance(val, float): c.number_format = '#,##0.00'
                elif isinstance(val, int): c.number_format = '#,##0'
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[rn].height = 18

    # Footer
    fr = 4 + len(data_rows) + 1
    ws.merge_cells(start_row=fr, start_column=1, end_row=fr, end_column=n)
    ts = timestamp[:19].replace('T', ' ') if timestamp else None
    c = ws.cell(row=fr, column=1, value=f"Sumber: {source_text}" + (f"  |  Waktu: {ts}" if ts else ""))
    c.font = Font(name="Arial", italic=True, color="595959", size=9)
    c.alignment = Alignment(horizontal="left", vertical="center")

    out = io.BytesIO()
    wb.save(out); out.seek(0)
    resp = HttpResponse(out.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


# ── Download Endpoints ────────────────────────────────────────────────────────
@api_view(['POST'])
def download_padi_xlsx(request):
    try:
        padi_data = request.data.get('padi_data')
        if not padi_data:
            return Response({"error": "Data tidak ditemukan"}, status=400)
        tahun = request.data.get('tahun', 2024)
        rows  = [
            [i, d.get('provinsi', p), d.get('luas_panen', '-'), d.get('produktivitas', '-'),
             d.get('produksi', '-'), d.get('rpp', '-')]
            for i, (p, d) in enumerate(sorted(padi_data.items()), 1)
        ]
        return _xlsx_response(
            "Produksi Padi",
            "PRODUKSI, LUAS PANEN & PRODUKTIVITAS PADI MENURUT PROVINSI",
            f"Sumber: BPS SIMDASI | Tahun {tahun} | Seluruh Provinsi Indonesia",
            ["No.", "Provinsi", "Luas Panen (ha)", "Produktivitas (ku/ha)", "Produksi (ton)", "RPP (ton/jiwa)"],
            [6, 35, 20, 22, 20, 18], rows, {3, 4, 5, 6},
            f"BPS Web API - SIMDASI, Produksi Padi, Tahun {tahun}",
            request.data.get('timestamp', datetime.now().isoformat()),
            f"Dataset_Produksi_Padi_BPS_{tahun}_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
        )
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@api_view(['POST'])
def download_konsumsi_xlsx(request):
    try:
        konsumsi_data = request.data.get('konsumsi_data')
        if not konsumsi_data:
            return Response({"error": "Data tidak ditemukan"}, status=400)
        tahun = request.data.get('tahun', 2024)
        rows  = [
            [i, d.get('provinsi', p), d.get('kalori', '-'), d.get('protein', '-'), d.get('ik', '-')]
            for i, (p, d) in enumerate(sorted(konsumsi_data.items()), 1)
        ]
        return _xlsx_response(
            "Konsumsi",
            "RATA-RATA KONSUMSI KALORI & PROTEIN PER KAPITA PER HARI MENURUT PROVINSI",
            f"Sumber: BPS Susenas | Tahun {tahun} | Seluruh Provinsi Indonesia",
            ["No.", "Provinsi", "Kalori (kkal/kap/hari)", "Protein (gr/kap/hari)", "Indeks Konsumsi (IK)"],
            [6, 35, 24, 24, 22], rows, {3, 4, 5},
            f"BPS Web API - Susenas Static Table 951, Tahun {tahun}",
            request.data.get('timestamp', datetime.now().isoformat()),
            f"Dataset_Konsumsi_Pangan_BPS_{tahun}_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
        )
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@api_view(['POST'])
def download_penduduk_xlsx(request):
    try:
        penduduk_data = request.data.get('penduduk_data')
        if not penduduk_data:
            return Response({"error": "Data tidak ditemukan"}, status=400)
        tahun = request.data.get('tahun', 2024)
        rows  = []
        for i, (p, d) in enumerate(sorted(penduduk_data.items()), 1):
            pct = d.get('pct_kemiskinan')
            ia  = round(1 - pct / 100, 4) if pct is not None else '-'
            ket = ('Akses Baik' if isinstance(ia, float) and ia >= 0.90
                   else 'Akses Sedang' if isinstance(ia, float) and ia >= 0.75
                   else 'Akses Lemah' if isinstance(ia, float) else '-')
            rows.append([i, d.get('provinsi', p), d.get('jumlah_penduduk', '-'),
                         pct if pct is not None else '-', ia, ket])
        return _xlsx_response(
            "Penduduk",
            "JUMLAH PENDUDUK DAN PERSENTASE PENDUDUK MISKIN MENURUT PROVINSI",
            f"Sumber: BPS SIMDASI | Tahun {tahun} | Seluruh Provinsi Indonesia",
            ["No.", "Provinsi", "Jumlah Penduduk (jiwa)", "% Penduduk Miskin", "Indeks Akses (IA)", "Keterangan"],
            [6, 35, 26, 22, 18, 20], rows, {3, 4, 5},
            f"BPS Web API - SIMDASI Penduduk & Kemiskinan, Tahun {tahun}",
            request.data.get('timestamp', datetime.now().isoformat()),
            f"Dataset_Penduduk_BPS_{tahun}_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
        )
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@api_view(['POST'])
def download_ikp_xlsx(request):
    try:
        ikp_data = request.data.get('ikp_data')
        if not ikp_data:
            return Response({"error": "Data tidak ditemukan"}, status=400)
        tahun = request.data.get('tahun', 2024)
        rows  = [
            [i, d.get('provinsi', p), d.get('rpp_norm', '-'), d.get('pl_norm', '-'),
             d.get('ik_norm', '-'), d.get('ia_norm', '-'), d.get('ikp', '-'),
             d.get('status', '-'), d.get('warna', '-')]
            for i, (p, d) in enumerate(sorted(ikp_data.items()), 1)
        ]
        return _xlsx_response(
            "IKP",
            "INDEKS KETAHANAN PANGAN (IKP) MENURUT PROVINSI",
            f"Sumber: BPS Web API | Tahun {tahun} | Seluruh Provinsi Indonesia",
            ["No.", "Provinsi", "RPP_norm", "PL_norm", "IK_norm", "IA_norm", "IKP", "Status", "Warna"],
            [6, 35, 12, 12, 12, 12, 12, 16, 12], rows, {3, 4, 5, 6, 7},
            f"BPS Web API - Indeks Ketahanan Pangan, Tahun {tahun}",
            request.data.get('timestamp', datetime.now().isoformat()),
            f"Dataset_IKP_BPS_{tahun}_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
        )
    except Exception as e:
        return Response({"error": str(e)}, status=500)


# ── Analytics Class ───────────────────────────────────────────────────────────
class PanganAnalytics:
    def __init__(self, tahun=2024):
        self.tahun = tahun
        self.config = get_pangan_config(tahun)
        self.timestamp_fetch = None

    def fetch_selected_data(self, keys: list) -> dict:
        all_data = {}
        self.timestamp_fetch = datetime.now().isoformat()
        for key in keys:
            config = self.config.get(key)
            if not config:
                all_data[key] = None; continue
            try:
                url  = config["url_template"].format(key=BPS_API_KEY)
                resp = requests.get(url, timeout=30)
                all_data[key] = resp.json() if resp.status_code == 200 else None
                print(f"{'✓' if resp.status_code == 200 else '✗'} {key}: {resp.status_code}")
            except Exception as e:
                print(f"✗ {key}: {e}")
                all_data[key] = None
        return all_data

    def _parse_simdasi_table(self, raw_data):
        """Ambil data_container dan kolom dari respons SIMDASI."""
        if not raw_data:
            return None, None
        for item in raw_data.get('data', []):
            if isinstance(item, dict) and 'data' in item and isinstance(item['data'], list):
                return item.get('data', []), item.get('kolom', {})
        return None, None

    def _clean_float(self, raw):
        s = str(raw).replace(' ', '').replace('.', '').replace(',', '.').strip()
        return None if s in ['-', '...', ''] else (float(s) if s else None)

    def _clean_prov(self, raw):
        raw = raw.strip()
        if raw.upper() in ('INDONESIA', 'TOTAL', 'NASIONAL', 'JUMLAH'):
            return None
        if any(x in raw.lower() for x in ['kab.', 'kota ', 'kabupaten', 'kecamatan']):
            return None
        return normalize_province_name(raw)

    def parse_padi_data(self, raw_data):
        padi_values, padi_raw = {}, {}
        data_rows, kolom_info = self._parse_simdasi_table(raw_data)
        if not data_rows:
            return padi_values, padi_raw
        try:
            col_produksi = col_luas_panen = col_produktivitas = None
            for k, v in kolom_info.items():
                nama = v.get('nama_variabel', '').lower()
                if 'produksi' in nama and 'padi' in nama:  col_produksi = k
                elif 'luas panen' in nama:                  col_luas_panen = k
                elif 'produktivitas' in nama:               col_produktivitas = k
            if not all([col_produksi, col_luas_panen, col_produktivitas]):
                ks = list(kolom_info.keys())
                if len(ks) >= 3:
                    col_luas_panen    = col_luas_panen    or ks[0]
                    col_produktivitas = col_produktivitas or ks[1]
                    col_produksi      = col_produksi      or ks[2]

            for row in data_rows:
                if not isinstance(row, dict): continue
                prov_raw = (row.get('label') or row.get('nama_wilayah') or row.get('nama') or '').strip()
                prov = self._clean_prov(prov_raw)
                if not prov: continue
                vars_ = row.get('variables', {})
                def pv(col):
                    if not col or col not in vars_: return None
                    return self._clean_float(vars_[col].get('value', ''))
                produksi = pv(col_produksi)
                if produksi is None: continue
                padi_values[prov] = {'produksi': produksi, 'luas_panen': pv(col_luas_panen), 'produktivitas': pv(col_produktivitas)}
                padi_raw[prov]    = {'provinsi': prov, **padi_values[prov], 'rpp': None}
        except Exception as e:
            print(f"❌ Parse PADI: {e}")
        return padi_values, padi_raw

    def parse_konsumsi_data(self, raw_data: dict, target_year: int) -> dict:
        import html as html_module, re
        from bs4 import BeautifulSoup
        result = {}
        try:
            table_html = html_module.unescape(raw_data.get('data', {}).get('table', ''))
            if not table_html:
                return result, {}
            rows = BeautifulSoup(table_html, 'html.parser').find('table').find_all('tr')
            if len(rows) < 4:
                return result, {}

            yr_row_idx = next((i for i, r in enumerate(rows)
                               if sum(1 for c in r.find_all(['td','th']) if re.fullmatch(r'\d{4}', c.get_text(strip=True))) >= 10), None)
            if yr_row_idx is None:
                return result, {}

            yrs = [c.get_text(strip=True) for c in rows[yr_row_idx].find_all(['td','th'])]
            tys = str(target_year)
            k_idx = p_idx = None
            for idx, yr in enumerate(yrs):
                if yr == tys:
                    if k_idx is None: k_idx = idx
                    else: p_idx = idx; break

            if k_idx is None:
                avail = sorted({int(y) for y in yrs if re.fullmatch(r'\d{4}', y)})
                if not avail: return result, {}
                tys = str(min(avail, key=lambda y: abs(y - target_year)))
                for idx, yr in enumerate(yrs):
                    if yr == tys:
                        if k_idx is None: k_idx = idx
                        else: p_idx = idx; break

            def sf(cells, idx):
                if idx is None or idx >= len(cells): return None
                raw = re.sub(r'[\s\u00a0]', '', cells[idx].get_text(strip=True)).replace(',', '.')
                return None if raw in ('', '-', 'n/a') else (float(raw) if raw else None)

            SKIP = {'indonesia', 'diolah', 'sumber', 'catatan', '1)', '-'}
            for row in rows[yr_row_idx + 1:]:
                cells = row.find_all(['td','th'])
                if not cells: continue
                prov = re.sub(r'\s*\d+\)?\s*$', '', ' '.join(cells[0].get_text().split())).strip()
                if not prov or len(prov) < 3 or prov.lower() in SKIP: continue
                k, p = sf(cells, k_idx + 1 if k_idx else None), sf(cells, p_idx + 1 if p_idx else None)
                if k is None and p is None: continue
                result[normalize_province_name(prov)] = {'kalori': k, 'protein': p}
        except Exception as e:
            print(f"❌ Parse KONSUMSI: {e}")

        konsumsi_raw = {prov: {'provinsi': prov, **d, 'ik': round(d['protein'] / 57.0, 4) if d.get('protein') else None}
                        for prov, d in result.items()}
        return result, konsumsi_raw

    def parse_kemiskinan_data(self, raw_data):
        miskin_values, miskin_raw = {}, {}
        if not raw_data:
            return miskin_values, miskin_raw
        try:
            vervar_list = raw_data.get('vervar', [])
            datacontent = raw_data.get('datacontent', {})
            th_val_list = raw_data.get('tahun', [])
            if not vervar_list or not datacontent:
                return miskin_values, miskin_raw
            th_val = th_val_list[0]['val'] if th_val_list else (self.tahun - 1900)

            for item in vervar_list:
                vv, label = item.get('val'), item.get('label', '').strip()
                if not label or label.upper() in ('INDONESIA', 'TOTAL', 'NASIONAL') or vv == 9999:
                    continue
                pct = (datacontent.get(f"{vv:04d}192434{th_val}{61:02d}") or
                       datacontent.get(f"{vv:04d}192434{th_val}{62:02d}"))
                if pct is None:
                    for tv in [433, 432]:
                        pct = datacontent.get(f"{vv:04d}192{tv}{th_val}{61:02d}")
                        if pct is not None: break
                if pct is None: continue
                prov = normalize_province_name(label)
                miskin_values[prov] = float(pct)
                miskin_raw[prov]    = {'provinsi': prov, 'pct_kemiskinan': float(pct), 'ia': round(1 - float(pct)/100, 4)}
        except Exception as e:
            print(f"❌ Parse KEMISKINAN: {e}")
        return miskin_values, miskin_raw

    def parse_penduduk_data(self, raw_data):
        penduduk_values, penduduk_raw = {}, {}
        data_rows, kolom_info = self._parse_simdasi_table(raw_data)
        if not data_rows:
            return penduduk_values, penduduk_raw
        try:
            col_pddk = COL_PENDUDUK_JUMLAH
            unit_mult = 1000
            if col_pddk in kolom_info:
                um = kolom_info[col_pddk].get('unit_multiplier')
                if um is not None: unit_mult = 10 ** int(um)
            else:
                col_pddk = None
                for k, v in kolom_info.items():
                    nama = v.get('nama_variabel', '').lower()
                    if 'jumlah penduduk' in nama or ('population' in nama and 'growth' not in nama):
                        col_pddk = k
                        um = v.get('unit_multiplier')
                        unit_mult = 10 ** int(um) if um is not None else 1000
                        break
                if not col_pddk:
                    col_pddk = list(kolom_info.keys())[0] if kolom_info else None
            if not col_pddk:
                return penduduk_values, penduduk_raw

            for row in data_rows:
                if not isinstance(row, dict): continue
                prov_raw = (row.get('label') or row.get('nama_wilayah') or row.get('nama') or '').strip()
                prov = self._clean_prov(prov_raw)
                if not prov: continue
                if str(row.get('kode_wilayah', '')).replace(' ', '') in ('0', '0000000', '00000000'):
                    continue
                val_entry = row.get('variables', {}).get(col_pddk, {})
                if not val_entry: continue
                val_raw = self._clean_float(val_entry.get('value', ''))
                if val_raw is None or val_raw <= 0: continue
                jiwa = val_raw * unit_mult
                penduduk_values[prov] = jiwa
                penduduk_raw[prov]    = {'provinsi': prov, 'jumlah_penduduk': int(jiwa), 'pct_kemiskinan': None}
        except Exception as e:
            print(f"❌ Parse PENDUDUK: {e}")
        return penduduk_values, penduduk_raw


# ── Helpers Normalisasi & IKP ─────────────────────────────────────────────────
def normalize_province_name(name):
    if not isinstance(name, str): name = str(name)
    name = name.upper().strip()
    for prefix in ['PROVINSI ', 'PROV. ', 'PROV ', 'DAERAH KHUSUS IBUKOTA ']:
        if name.startswith(prefix):
            name = name[len(prefix):]
    MAPPINGS = {
        'DKI JAKARTA': 'JAKARTA', 'DAERAH KHUSUS IBUKOTA JAKARTA': 'JAKARTA', 'DKI': 'JAKARTA',
        'YOGYAKARTA': 'DAERAH ISTIMEWA YOGYAKARTA', 'DIY': 'DAERAH ISTIMEWA YOGYAKARTA',
        'D.I. YOGYAKARTA': 'DAERAH ISTIMEWA YOGYAKARTA', 'DI YOGYAKARTA': 'DAERAH ISTIMEWA YOGYAKARTA',
        'BANGKA BELITUNG': 'KEPULAUAN BANGKA BELITUNG', 'KEP. BANGKA BELITUNG': 'KEPULAUAN BANGKA BELITUNG',
        'KEP. RIAU': 'KEPULAUAN RIAU', 'NTB': 'NUSA TENGGARA BARAT', 'NTT': 'NUSA TENGGARA TIMUR',
    }
    for k, v in MAPPINGS.items():
        if name == k or name.endswith(' ' + k): return v
    if 'KEP.' in name: name = name.replace('KEP.', 'KEPULAUAN')
    return name.strip()


def minmax_normalize(values_dict):
    valid = {k: v for k, v in values_dict.items() if v is not None and v > 0}
    if not valid: return {k: 0.0 for k in values_dict}
    v_min, v_max = min(valid.values()), max(valid.values())
    return {
        k: (1.0 if v_max == v_min else round((v - v_min) / (v_max - v_min), 4)) if v and v > 0 else 0.0
        for k, v in values_dict.items()
    }

def calculate_ikp(rpp_n, pl_n, ik_n, ia_n):
    return round(0.4 * ((rpp_n + pl_n) / 2) + 0.3 * ik_n + 0.3 * ia_n, 4)

def classify_ikp(ikp):
    if ikp >= 0.70: return "TINGGI", "#10b981"
    if ikp >= 0.50: return "SEDANG", "#f59e0b"
    return "RENDAH", "#ef4444"


def generate_pangan_insights(provinsi, rpp, pl, ik, ia, ikp, status):
    out = [f"Provinsi {provinsi} memiliki status ketahanan pangan {status} dengan IKP {ikp:.3f}."]
    if rpp is not None:
        if rpp > 0.5:   out.append(f"✅ Produksi padi cukup tinggi ({rpp:.4f} ton/jiwa), ketersediaan pangan baik.")
        elif rpp > 0.2: out.append(f"📊 Rasio produksi padi {rpp:.4f} ton/jiwa - masih dalam batas wajar.")
        else:           out.append(f"⚠️ Rasio produksi padi rendah ({rpp:.4f} ton/jiwa), rentan kekurangan stok.")
    if ik is not None:
        if ik >= 1.0:   out.append(f"✅ Indeks konsumsi protein {ik:.2f} ≥ 1.0 - kecukupan gizi terpenuhi.")
        elif ik >= 0.85:out.append(f"📊 Indeks konsumsi protein {ik:.2f} - mendekati standar kecukupan.")
        else:           out.append(f"📉 Indeks konsumsi protein {ik:.2f} < 0.85 - konsumsi protein di bawah standar.")
    if ia is not None:
        pct = round((1 - ia) * 100, 2)
        if ia >= 0.90:  out.append(f"✅ Akses pangan baik - tingkat kemiskinan hanya {pct}%.")
        elif ia >= 0.75:out.append(f"📊 Akses pangan sedang - kemiskinan {pct}%, perlu peningkatan daya beli.")
        else:           out.append(f"⚠️ Akses pangan lemah - kemiskinan {pct}%, risiko rawan pangan tinggi.")
    return out


def generate_pangan_recommendations(status, rpp, pl, ik, ia):
    RECS = {
        "RENDAH": {
            "title": "Percepatan Ketahanan Pangan", "priority": "Tinggi",
            "actions": ['Program GERNAS peningkatan produksi padi', 'Distribusi subsidi pupuk dan benih',
                        'Pemberdayaan BULOG untuk stabilisasi stok', 'Bantuan sosial pangan (BPNT/PKH)',
                        'Kampanye diversifikasi pangan lokal'],
        },
        "SEDANG": {
            "title": "Penguatan Ketahanan Pangan", "priority": "Sedang",
            "actions": ['Perluasan lahan & optimasi irigasi', 'Pelatihan petani & mekanisasi',
                        'Pengembangan pasar tani lokal', 'Program gizi & fortifikasi pangan',
                        'Penguatan rantai pasok pangan daerah'],
        },
        "TINGGI": {
            "title": "Inovasi Ketahanan Pangan", "priority": "Rendah",
            "actions": ['Smart farming & agrikultur presisi', 'Ekspansi pasar ekspor pertanian',
                        'Industri pengolahan pangan lokal', 'Ketahanan pangan berbasis komunitas',
                        'Pengembangan cadangan pangan strategis'],
        },
    }
    recs = [RECS.get(status, RECS["SEDANG"])]
    if ik is not None and ik < 0.85:
        recs.append({'priority': 'Tinggi', 'title': 'Peningkatan Kecukupan Protein',
            'actions': ['Diversifikasi protein (ikan, telur, kedelai)', 'Fortifikasi produk pangan',
                        'Edukasi gizi keluarga via Posyandu', 'Subsidi protein hewani untuk keluarga rentan']})
    if ia is not None and ia < 0.80:
        recs.append({'priority': 'Tinggi', 'title': 'Peningkatan Akses Ekonomi Pangan',
            'actions': ['Perluas cakupan BPNT & PKH', 'Operasi pasar murah berkala',
                        'Pengembangan lumbung pangan desa', 'Subsidi harga beras untuk rumah tangga miskin']})
    return recs


# ── AI Prediction Helpers ─────────────────────────────────────────────────────
def _build_feature_row(tahun_pred, prov_name, cache, meta):
    tahun_rel = tahun_pred - meta["tahun_rel_base"]
    row = {"tahun_rel": tahun_rel, "prov_enc": cache["prov_enc"], "pulau_enc": cache["pulau_enc"]}
    for col in meta["targets"]:
        v, vp = cache.get(col, 0.0), cache.get(f"{col}_prev", 0.0)
        row[f"{col}_lag1"] = v; row[f"{col}_lag2"] = vp
        row[f"{col}_delta"] = v - vp; row[f"{col}_roll2"] = (v + vp) / 2.0
    for col in ["rpp", "pl", "ik", "ia"]:
        row[f"{col}_std2"] = abs(cache.get(col, 0.0) - cache.get(f"{col}_prev", 0.0)) * 0.5
    for dc in meta["pulau_dummy_cols"]:
        row[dc] = cache.get(dc, 0.0)
    return [row.get(f, 0.0) for f in meta["features"]]


def _predict_one_province(prov_name, tahun_list, hist_last, hist_prev, meta, models, encoders):
    try:
        prov_enc = int(encoders["provinsi"].transform([prov_name])[0])
    except Exception:
        return []
    pulau_label = meta["pulau_map"].get(prov_name, "Sumatera")
    try: pulau_enc = int(encoders["pulau"].transform([pulau_label])[0])
    except: pulau_enc = 0

    cache = {"prov_enc": prov_enc, "pulau_enc": pulau_enc}
    for col in meta["targets"]:
        cache[col]           = float(hist_last.get(col) or 0.0)
        cache[f"{col}_prev"] = float(hist_prev.get(col) or 0.0)
    for dc in meta["pulau_dummy_cols"]:
        cache[dc] = 1.0 if pulau_label == dc.replace("pulau_", "") else 0.0

    results = []
    CLIPS = {"ikp": (0.0, 1.0), "ia": (0.0, 1.0), "ik": (0.0, 3.0)}
    for tahun_pred in tahun_list:
        X = pd.DataFrame([_build_feature_row(tahun_pred, prov_name, cache, meta)], columns=meta["features"])
        pred = {}
        for col in meta["targets"]:
            raw = float(models[col].predict(X)[0])
            pred[col] = round(float(np.clip(raw, *CLIPS.get(col, (0.0, None)))), 6)
        status, warna = classify_ikp(pred["ikp"])
        results.append({
            "provinsi": prov_name, "tahun": tahun_pred, **pred,
            "status": status, "warna": warna,
            "pct_miskin_est": round((1 - pred["ia"]) * 100, 2),
            "is_prediction": True, "model_version": meta.get("version", "rf_v1.0"),
        })
        for col in meta["targets"]:
            cache[f"{col}_prev"] = cache[col]; cache[col] = pred[col]
    return results


def _run_ai_prediction(tahun: int, historical_data: dict) -> dict:
    if not _load_ai_models():
        return {"error": "Model AI tidak dapat dimuat."}

    # Build hist_lookup dari historical_data atau MongoDB
    hist_lookup = {r["provinsi"]: {k: r.get(k) for k in ["rpp","pl","ik","ia","ikp"]}
                   for r in historical_data.get("analysis_summary", []) if r.get("provinsi")}
    if not hist_lookup:
        try:
            docs = list(mongo_db["pangan_analysis"].find(
                {"type": "pangan", "is_ai_prediction": {"$ne": True}},
                {"_id": 0, "analysis_summary": 1}
            ).sort("timestamp", -1).limit(1))
            if docs:
                hist_lookup = {r["provinsi"]: {k: r.get(k) for k in ["rpp","pl","ik","ia","ikp"]}
                               for r in docs[0].get("analysis_summary", []) if r.get("provinsi")}
        except Exception as e:
            print(f"[AI] MongoDB fallback gagal: {e}")

    meta, models, encoders = _AI_META, _AI_MODELS, _AI_ENCODERS
    tahun_range = list(range(2025, tahun + 1))
    timestamp_ai = datetime.now().isoformat()

    boundary_features = list(mongo_db["batas_provinsi"].find({}, {"_id": 0}))
    province_map = {}
    for feat in boundary_features:
        for field in ["NAMOBJ", "name", "WADMPR", "Provinsi"]:
            v = feat.get("properties", {}).get(field)
            if v:
                n = normalize_province_name(str(v).upper().strip())
                province_map[n] = province_map[str(v).upper().strip()] = feat

    # ── Kumpulkan semua prediksi mentah dulu ─────────────────────────────────
    raw_preds = []
    for prov_name in meta["provinsi_list"]:
        hl = hist_lookup.get(prov_name, {})
        hp = {c: (hl.get(c) or 0.0) * 0.98 for c in ["rpp","pl","ik","ia","ikp"]} if hl else {}
        results = _predict_one_province(prov_name, tahun_range, hl, hp, meta, models, encoders)
        if results:
            raw_preds.append(results[-1])

    # ── Deteksi & perbaiki kolom degenerate (variance ≈ 0 → model stuck) ─────
    # Terjadi saat semua lag-features = 0 (tidak ada data historis) →
    # model RF konvergen ke nilai konstan, biasanya 1.0 untuk 'pl'.
    NEUTRAL   = {"rpp": 0.3, "pl": 0.5, "ik": 0.5, "ia": 0.5}
    VAR_THRESHOLD = 1e-4
    degenerate_cols = set()
    for col in ["rpp", "pl", "ik", "ia"]:
        vals = [r[col] for r in raw_preds if r.get(col) is not None]
        if len(vals) >= 2:
            mean = sum(vals) / len(vals)
            variance = sum((v - mean) ** 2 for v in vals) / len(vals)
            if variance < VAR_THRESHOLD:
                degenerate_cols.add(col)
                print(f"[AI] ⚠ '{col}' degenerate (var={variance:.2e}, mean={mean:.4f}) "
                      f"→ fallback netral {NEUTRAL[col]}")

    if degenerate_cols:
        for r in raw_preds:
            for col in degenerate_cols:
                r[col] = NEUTRAL[col]
            # Recalculate IKP pakai nilai yang sudah disanitasi
            r["ikp"] = calculate_ikp(
                r.get("rpp") or 0.0, r.get("pl") or 0.0,
                r.get("ik") or 0.0,  r.get("ia") or 0.0,
            )
            r["status"], r["warna"] = classify_ikp(r["ikp"])
            r["pct_miskin_est"] = round((1 - (r.get("ia") or 0.0)) * 100, 2)

    # ── Build output ──────────────────────────────────────────────────────────
    matched_features, analysis_summary = [], []
    status_counts = {"TINGGI": 0, "SEDANG": 0, "RENDAH": 0}

    for r in raw_preds:
        prov_name = r["provinsi"]
        status, warna = r["status"], r["warna"]
        status_counts[status] += 1

        analysis_summary.append({
            "provinsi": prov_name, "indikator": "ALL", "status": status, "warna": warna,
            "ikp": r["ikp"], "rpp": r["rpp"], "pl": r["pl"], "ik": r["ik"], "ia": r["ia"],
            "has_complete_data": True, "dimensi_kosong": [],
            "is_prediction": True, "degenerate_cols": list(degenerate_cols),
            "model_version": meta.get("version", "rf_v1.0"),
        })

        feat = province_map.get(prov_name) or next(
            (f for n, f in province_map.items() if prov_name in n or n in prov_name), None)
        if not feat: continue

        fc = feat.copy()
        fc["properties"] = {**fc.get("properties", {}), "pangan_analysis": {
            "nama_provinsi": prov_name, "indikator": "ALL", "status": status, "warna": warna,
            "ikp": r["ikp"], "rpp_norm": r["rpp"], "pl_norm": r["pl"], "ik_norm": r["ik"], "ia_norm": r["ia"],
            "has_complete_data": True, "dimensi_kosong": [],
            "is_prediction": True, "degenerate_cols": list(degenerate_cols),
            "model_version": meta.get("version", "rf_v1.0"),
            "insights": generate_pangan_insights(prov_name, r["rpp"], r["pl"], r["ik"], r["ia"], r["ikp"], status),
            "rekomendasi": generate_pangan_recommendations(status, r["rpp"], r["pl"], r["ik"], r["ia"]),
            "data_pangan": {k: r[k] for k in ["rpp","pl","ik","ia","ikp","pct_miskin_est"]},
        }}
        matched_features.append(fc)

    return {
        "status": "success", "tahun": tahun, "is_ai_prediction": True,
        "source": f"AI Prediction - Random Forest {meta.get('version','rf_v1.0')}",
        "model_version": meta.get("version", "rf_v1.0"),
        "model_scores": meta.get("scores", {}), "indikator": "ALL",
        "dataset_aktif": ["PADI", "KONSUMSI", "KEMISKINAN", "PENDUDUK"],
        "total_provinsi": len(analysis_summary), "total_dipetakan": len(matched_features),
        "total_data_kosong": 0, "total_success": len(matched_features),
        "ada_data_kosong": False, "alert_message": None, "provinsi_data_kosong": [],
        "status_distribusi": status_counts, "timestamp": timestamp_ai,
        "matched_features": {"type": "FeatureCollection", "features": matched_features},
        "analysis_summary": analysis_summary, "raw_datasets": {},
    }


# ── Endpoints AI & BPS ────────────────────────────────────────────────────────
@api_view(['POST'])
def analyze_pangan_ai(request):
    try: tahun = int(request.data.get("tahun", 2025))
    except: tahun = 2025
    result = _run_ai_prediction(tahun, request.data.get("historical_data", {}))
    return Response(result, status=500 if "error" in result else 200)


@api_view(['POST'])
def analyze_pangan_bps(request):
    if not BPS_API_KEY:
        return Response({"error": "BPS Web API Key belum dikonfigurasi"}, status=500)
    try: tahun = int(request.data.get('tahun', 2025))
    except: tahun = 2025

    mode = request.data.get('mode', 'bps')
    if mode == 'ai':
        result = _run_ai_prediction(tahun, request.data.get("historical_data", {}))
        return Response(result, status=500 if "error" in result else 200)

    if tahun not in TAHUN_SUPPORTED:
        return Response({"error": f"Tahun {tahun} tidak didukung."}, status=400)

    try:
        analytics = PanganAnalytics(tahun=tahun)
        raw_data  = analytics.fetch_selected_data(INDIKATOR_DATASET_MAP['ALL'])

        padi_values,     padi_raw     = analytics.parse_padi_data(raw_data.get('PADI'))
        konsumsi_values, konsumsi_raw = analytics.parse_konsumsi_data(raw_data.get('KONSUMSI'), tahun)
        miskin_values,   miskin_raw   = analytics.parse_kemiskinan_data(raw_data.get('KEMISKINAN'))
        penduduk_values, penduduk_raw = analytics.parse_penduduk_data(raw_data.get('PENDUDUK'))

        for prov in penduduk_raw:
            penduduk_raw[prov]['pct_kemiskinan'] = miskin_values.get(prov)

        # Dedup DKI Jakarta
        for alias, canon in [('DKI JAKARTA','JAKARTA'), ('DAERAH KHUSUS IBUKOTA JAKARTA','JAKARTA')]:
            for ds in [padi_values, padi_raw, konsumsi_values, konsumsi_raw,
                       miskin_values, miskin_raw, penduduk_values, penduduk_raw]:
                if alias in ds and canon in ds: del ds[alias]
                elif alias in ds: ds[canon] = ds.pop(alias)

        # Hitung RPP dan PL
        rpp_values, pl_values = {}, {}
        for prov, padi in padi_values.items():
            prod, lp, pddk = padi.get('produksi'), padi.get('luas_panen'), penduduk_values.get(prov)
            if prod and pddk and pddk > 0:
                rpp_values[prov] = prod / pddk
                if padi_raw.get(prov): padi_raw[prov]['rpp'] = round(rpp_values[prov], 6)
            if prod and lp and lp > 0:
                pl_values[prov] = prod / lp

        # Daftar 38 provinsi canonical
        PROVINSI_CANONICAL_38 = [
            'ACEH','SUMATERA UTARA','SUMATERA BARAT','RIAU','JAMBI','SUMATERA SELATAN',
            'BENGKULU','LAMPUNG','KEPULAUAN BANGKA BELITUNG','KEPULAUAN RIAU','JAKARTA',
            'JAWA BARAT','JAWA TENGAH','DAERAH ISTIMEWA YOGYAKARTA','JAWA TIMUR','BANTEN',
            'BALI','NUSA TENGGARA BARAT','NUSA TENGGARA TIMUR','KALIMANTAN BARAT',
            'KALIMANTAN TENGAH','KALIMANTAN SELATAN','KALIMANTAN TIMUR','KALIMANTAN UTARA',
            'SULAWESI UTARA','SULAWESI TENGAH','SULAWESI SELATAN','SULAWESI TENGGARA',
            'GORONTALO','SULAWESI BARAT','MALUKU','MALUKU UTARA','PAPUA BARAT',
            'PAPUA BARAT DAYA','PAPUA','PAPUA SELATAN','PAPUA TENGAH','PAPUA PEGUNUNGAN',
        ]


        all_prov_set  = set(list(padi_values)+list(konsumsi_values)+list(miskin_values)+list(penduduk_values))
        all_provinces = PROVINSI_CANONICAL_38 + sorted(all_prov_set - set(PROVINSI_CANONICAL_38))

        # Hitung IK/IA
        ik_values = {p: d.get('protein')/57.0 for p,d in konsumsi_values.items() if d.get('protein')}
        ia_values = {p: 1 - v/100 for p,v in miskin_values.items()}

        rpp_norm_map = minmax_normalize(rpp_values)
        pl_norm_map  = minmax_normalize(pl_values)
        ik_norm_map  = minmax_normalize(ik_values)
        ia_norm_map  = minmax_normalize(ia_values)

        # Build province_map dari MongoDB
        province_map = {}
        for feat in mongo_db["batas_provinsi"].find({}, {'_id': 0}):
            for field in ['NAMOBJ','name','WADMPR','Provinsi']:
                v = feat.get('properties', {}).get(field)
                if v:
                    n = normalize_province_name(str(v).upper().strip())
                    province_map[n] = province_map[str(v).upper().strip()] = feat

        matched_features, analysis_summary = [], []
        status_counts, ikp_raw_output, provinsi_data_kosong = {"TINGGI":0,"SEDANG":0,"RENDAH":0}, {}, []

        for prov_name in all_provinces:
            has_padi = prov_name in padi_values
            has_k    = prov_name in konsumsi_values
            has_m    = prov_name in miskin_values
            has_p    = prov_name in penduduk_values
            has_any  = has_padi or has_k or has_m or has_p

            kosong_dims = []
            if not has_padi: kosong_dims.append('Produksi Padi')
            if not has_k:    kosong_dims.append('Konsumsi')
            if not has_m:    kosong_dims.append('Kemiskinan')
            if not has_p:    kosong_dims.append('Penduduk')
            rpp_n = rpp_norm_map.get(prov_name, 0.0)
            pl_n  = pl_norm_map.get(prov_name, 0.0)
            ik_n  = ik_norm_map.get(prov_name, 0.0)
            ia_n  = ia_norm_map.get(prov_name, 0.0)
            rpp_r = rpp_values.get(prov_name)
            pl_r  = pl_values.get(prov_name)
            ik_r  = ik_values.get(prov_name)
            ia_r  = ia_values.get(prov_name)

            ikp = calculate_ikp(rpp_n, pl_n, ik_n, ia_n)
            status, warna = classify_ikp(ikp)
            if has_any: status_counts[status] += 1
            if kosong_dims: provinsi_data_kosong.append({'provinsi': prov_name, 'dimensi_kosong': kosong_dims})

            analysis_summary.append({
                'provinsi': prov_name, 'indikator': 'ALL',
                'status': status if has_any else '-', 'warna': warna if has_any else '#94a3b8',
                'ikp': ikp if has_any else None,
                'rpp': rpp_r,
                'pl':  pl_r,
                'ik':  ik_r, 'ia': ia_r,
                'has_complete_data': not kosong_dims, 'dimensi_kosong': kosong_dims,
            })
            ikp_raw_output[prov_name] = {
                'provinsi': prov_name, 'rpp_norm': rpp_n if has_padi else None,
                'pl_norm': pl_n if has_padi else None, 'ik_norm': ik_n if has_k else None,
                'ia_norm': ia_n if has_m else None, 'ikp': ikp if has_any else None,
                'status': status if has_any else '-', 'warna': warna if has_any else '#94a3b8',
                'has_complete_data': not kosong_dims, 'dimensi_kosong': kosong_dims,
            }

            feat = province_map.get(prov_name) or next(
                (f for n,f in province_map.items() if prov_name in n or n in prov_name), None)
            if not feat or not has_any: continue

            padi_p = padi_values.get(prov_name, {})
            kons_p = konsumsi_values.get(prov_name, {})
            fc = feat.copy()
            fc['properties'] = {**fc.get('properties',{}), 'pangan_analysis': {
                'nama_provinsi': prov_name, 'indikator': 'ALL',
                'status': status, 'warna': warna, 'ikp': ikp,
                'rpp_norm': rpp_n, 'pl_norm': pl_n, 'ik_norm': ik_n, 'ia_norm': ia_n,
                'has_complete_data': not kosong_dims, 'dimensi_kosong': kosong_dims,
                'insights':    generate_pangan_insights(prov_name, rpp_r, pl_r, ik_r, ia_r, ikp, status),
                'rekomendasi': generate_pangan_recommendations(status, rpp_r, pl_r, ik_r, ia_r),
                'data_pangan': {
                    'produksi_padi': padi_p.get('produksi'), 'luas_panen': padi_p.get('luas_panen'),
                    'produktivitas': padi_p.get('produktivitas'),
                    'rpp': rpp_r, 'pl': pl_r,
                    'kalori': kons_p.get('kalori'), 'protein': kons_p.get('protein'),
                    'ik': ik_r, 'pct_kemiskinan': miskin_values.get(prov_name), 'ia': ia_r,
                    'jumlah_penduduk': penduduk_values.get(prov_name),
                },
            }}
            matched_features.append(fc)

        ada_kosong = len(provinsi_data_kosong) > 0
        alert = f"⚠️ Ada {len(provinsi_data_kosong)} dataset/provinsi data tidak lengkap." if ada_kosong else None

        return Response({
            'status': 'success', 'tahun': tahun,
            'source': "BPS Web API",
            'is_ai_prediction': False,
            'indikator': 'ALL', 'dataset_aktif': INDIKATOR_DATASET_MAP['ALL'],
            'total_provinsi': len(analysis_summary), 'total_dipetakan': len(matched_features),
            'total_data_kosong': len(provinsi_data_kosong), 'total_success': len(matched_features),
            'ada_data_kosong': ada_kosong, 'alert_message': alert,
            'provinsi_data_kosong': provinsi_data_kosong, 'status_distribusi': status_counts,
            'timestamp': analytics.timestamp_fetch,
            'matched_features': {'type': 'FeatureCollection', 'features': matched_features},
            'analysis_summary': analysis_summary,
            'raw_datasets': {
                'timestamp': analytics.timestamp_fetch, 'tahun': tahun, 'indikator': 'ALL',
                'PADI': padi_raw, 'KONSUMSI': konsumsi_raw, 'KEMISKINAN': miskin_raw,
                'PENDUDUK': penduduk_raw, 'IKP': ikp_raw_output,
            },
        })
    except Exception as e:
        import traceback; traceback.print_exc()
        return Response({'error': str(e)}, status=500)


# ── Endpoint: Info Model AI ───────────────────────────────────────────────────
@api_view(['GET'])
def get_ai_model_info(request):
    if not _load_ai_models():
        return Response({"loaded": False, "error": "Model tidak ditemukan.", "ai_model_dir": AI_MODEL_DIR})
    return Response({
        "loaded": True, "version": _AI_META.get("version"), "created_at": _AI_META.get("created_at"),
        "tahun_historis": _AI_META.get("tahun_historis"), "tahun_prediksi": _AI_META.get("tahun_prediksi"),
        "provinsi_count": len(_AI_META.get("provinsi_list", [])),
        "features_count": len(_AI_META.get("features", [])),
        "targets": _AI_META.get("targets"), "scores": _AI_META.get("scores"),
        "n_train_rows": _AI_META.get("n_train_rows"), "ai_model_dir": AI_MODEL_DIR,
    })


# ── CRUD Simpan / List / Detail / Hapus ──────────────────────────────────────
@api_view(['POST'])
def save_pangan_analysis(request):
    try:
        name = request.data.get('name', 'Analisis Ketahanan Pangan Tanpa Nama')
        data = request.data.get('analysis_data')
        if not data:
            return Response({"error": "Data analisis tidak ditemukan"}, status=400)
        aid = str(uuid.uuid4())
        mongo_db["pangan_analysis"].insert_one({"analysis_id": aid, "name": name, "type": "pangan",
                                                "timestamp": datetime.now().isoformat(), **data})
        return Response({"status": "success", "message": f"'{name}' berhasil disimpan", "analysis_id": aid})
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@api_view(['GET'])
def get_pangan_analysis_list(request):
    try:
        results = list(mongo_db["pangan_analysis"].find({}, {
            '_id': 0, 'analysis_id': 1, 'name': 1, 'timestamp': 1,
            'total_success': 1, 'status_distribusi': 1, 'tahun': 1,
            'indikator': 1, 'is_ai_prediction': 1, 'source': 1,
        }).sort('timestamp', -1))
        return Response({"status": "success", "count": len(results), "results": results})
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@api_view(['GET'])
def get_pangan_analysis_detail(request, analysis_id):
    try:
        result = mongo_db["pangan_analysis"].find_one({"analysis_id": analysis_id}, {'_id': 0})
        if not result:
            return Response({"error": "Analisis tidak ditemukan"}, status=404)
        return Response(result)
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@api_view(['DELETE'])
def delete_pangan_analysis(request, analysis_id):
    try:
        r = mongo_db["pangan_analysis"].delete_one({"analysis_id": analysis_id})
        if r.deleted_count == 0:
            return Response({"error": "Analisis tidak ditemukan"}, status=404)
        return Response({"status": "success", "message": "Analisis berhasil dihapus"})
    except Exception as e:
        return Response({"error": str(e)}, status=500)